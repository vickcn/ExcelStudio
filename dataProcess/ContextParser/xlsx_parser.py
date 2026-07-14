
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XLSX 解析器 - 符合 ContextParser 統一架構

此模組提供 Excel / Spreadsheet 檔案解析功能，遵循 parser_design.md 定義的核心功能：
- preview: 使用 pandas 將每個 worksheet 輸出為 Markdown 表格文字
- extract: 先以 sheet 為單位建立 unit_paras，並蒐集公式 / 圖像等 metadata
- segment: 在每個 sheet 中偵測實際 table 區塊，必要時批次做圖文解釋
- chunk: 依 row / column / cell 模式，從每個 table segment 展開最終 chunk
- process: extract -> segment -> chunk 一條龍流程
"""

from __future__ import annotations

import os
import sys
import io
import re
import csv
import json
import math
import base64
import shutil
import logging
import subprocess
import zipfile
import posixpath
import xml.etree.ElementTree as ET
from pathlib import Path
from collections import Counter, defaultdict
from typing import List, Dict, Any, Optional, Tuple, Set
from dataclasses import dataclass
from openpyxl.utils import get_column_letter
from copy import deepcopy as dcp

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pd = None
    PANDAS_AVAILABLE = False

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    np = None
    NUMPY_AVAILABLE = False

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    load_workbook = None
    OPENPYXL_AVAILABLE = False

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

try:
    from ContextParser.xlsx_parser_support import (
        parse_keywords_from_text,
        dedup_multi_prompts_by_llm,
        filter_meaningless_tags,
        analyze_images_via_batch_common,
        batch_extract_multi_prompts,
        merge_multi_prompts,
        attach_multi_prompts_meta,
        is_enable_multi_prompts,
        attach_segment_multi_prompts_meta,
    )
except Exception:
    try:
        from .xlsx_parser_support import (
            parse_keywords_from_text,
            dedup_multi_prompts_by_llm,
            filter_meaningless_tags,
            analyze_images_via_batch_common,
            batch_extract_multi_prompts,
            merge_multi_prompts,
            attach_multi_prompts_meta,
            is_enable_multi_prompts,
            attach_segment_multi_prompts_meta,
        )
    except Exception:
        parse_keywords_from_text = None
        dedup_multi_prompts_by_llm = None
        filter_meaningless_tags = None
        analyze_images_via_batch_common = None
        batch_extract_multi_prompts = None
        merge_multi_prompts = None
        attach_multi_prompts_meta = None
        is_enable_multi_prompts = None
        attach_segment_multi_prompts_meta = None

DATA_PROCESS_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if DATA_PROCESS_DIR not in sys.path:
    sys.path.insert(0, DATA_PROCESS_DIR)
try:
    from package import LOGger
    from package import dataframeprocedure as DFP
except Exception:
    try:
        import LOGger  # type: ignore
        import dataframeprocedure as DFP  # type: ignore
    except Exception:
        LOGger = None
        DFP = None

m_fn = os.path.basename(__file__).replace('.py', '')


def _setup_logger() -> logging.Logger:
    if LOGger is not None:
        log_dir = os.path.join(os.path.dirname(__file__), 'log')
        os.makedirs(log_dir, exist_ok=True)
        logger = LOGger.addloger(logfile=os.path.join(log_dir, 'context_parser_%t.log'))
        logger.error = lambda x, *args, colora=getattr(LOGger, 'FAIL', None), **kwargs: logger(x, *args, **kwargs, colora=colora)
        logger.warning = lambda x, *args, colora=getattr(LOGger, 'WARNING', None), **kwargs: logger(x, *args, **kwargs, colora=colora)
        logger.info = lambda x, *args, **kwargs: logger(x, *args, **kwargs)
        logger.debug = lambda x, *args, colora=getattr(LOGger, 'OKBLUE', None), **kwargs: logger(x, *args, **kwargs, colora=colora)
        logger.exception = lambda x, logfile='', **kwargs: LOGger.exception_process(x, logfile=logfile, **kwargs)
        logger.summary = lambda x, *args, colora=getattr(LOGger, 'OKCYAN', None), **kwargs: logger(x, *args, **kwargs, colora=colora)
        return logger

    logger = logging.getLogger(m_fn)
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        handler = logging.StreamHandler()
        handler.setFormatter(logging.Formatter('[%(levelname)s] %(message)s'))
        logger.addHandler(handler)
    return logger


m_logger = _setup_logger()

VISUAL_ANNOTATION_LINE_RE = re.compile(
    r"(?m)^\[(?:圖像說明|補充物件內容|頁面視覺說明|表格視覺說明)\|[^\]]+\][^\n]*\n?"
)
VALIDATION_ANNOTATION_LINE_RE = re.compile(
    r"(?m)^\[下拉選單\|[^\]]+\][^\n]*\n?"
)
VALIDATION_SECTION_HEADER_RE = re.compile(r"(?m)^## 下拉選單\s*\n?")


def _strip_visual_annotation_lines(text: str) -> str:
    if not isinstance(text, str) or not text:
        return ''
    return VISUAL_ANNOTATION_LINE_RE.sub('', text).strip()


def _strip_annotation_lines(text: str) -> str:
    if not isinstance(text, str) or not text:
        return ''
    cleaned = VALIDATION_ANNOTATION_LINE_RE.sub('', text)
    cleaned = VALIDATION_SECTION_HEADER_RE.sub('', cleaned)
    cleaned = VISUAL_ANNOTATION_LINE_RE.sub('', cleaned)
    return cleaned.strip()
m_config_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config.json')
m_config = {}
if LOGger is not None and os.path.exists(m_config_file):
    try:
        m_config = LOGger.load_json(m_config_file)
    except Exception:
        m_config = {}


def _log_prefix(color_name: str) -> str:
    if LOGger is None:
        return ''
    return getattr(LOGger, color_name, '') or ''


# ============================================================
# 輔助函式
# ============================================================

# dataclass reserved for next refactor; current pipeline still uses dict for backward compatibility
@dataclass
class HeaderInfo:
    has_header_row: bool
    has_header_col: bool
    header_rows: list[int]
    header_cols: list[int]
    headers: list[str]
    confidence: float


@dataclass
class IndexInfo:
    has_index_col: bool
    has_index_row: bool
    index_cols: list[int]
    index_rows: list[int]
    index_values: list[str]
    confidence: float


@dataclass
class TableMetadata:
    memo: list[str]
    formulas: list[dict]
    images: list[dict]
    merged_cells: list[dict]
    original_bbox: tuple[int, int, int, int]
    sheet_bbox: tuple[int, int, int, int] | None


@dataclass
class TableRegion:
    table_id: int
    sheet_name: str
    sheet_index: int

    bbox: tuple[int, int, int, int]
    shape: tuple[int, int]
    non_empty_count: int
    density: float

    table_df: Any
    values: list[list[Any]]
    markdown: str

    header: HeaderInfo
    index: IndexInfo
    metadata: TableMetadata

    region_type: str = "unknown"
    confidence: float = 0.0
    reason: str = ""


def _require_dependencies() -> None:
    if not PANDAS_AVAILABLE:
        raise ImportError('pandas 未安裝，請執行: pip install pandas')
    if not NUMPY_AVAILABLE:
        raise ImportError('numpy 未安裝，請執行: pip install numpy')
    if not OPENPYXL_AVAILABLE:
        raise ImportError('openpyxl 未安裝，請執行: pip install openpyxl')


class _MiniDFP:
    @staticmethod
    def rpathrpt(path: str) -> str:
        p = Path(path)
        stem = p.stem
        suffix = p.suffix
        parent = p.parent
        idx = 1
        candidate = p
        while candidate.exists():
            candidate = parent / f'{stem}_{idx}{suffix}'
            idx += 1
        return str(candidate)


def _rpathrpt(path: str) -> str:
    if DFP is not None and hasattr(DFP, 'rpathrpt'):
        return DFP.rpathrpt(path)
    return _MiniDFP.rpathrpt(path)



def _normalize_none_like(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str) and value.strip() == '':
        return None
    return value



def _truncate_text(text: str, max_len: Optional[int] = None) -> str:
    if max_len is None or max_len <= 0:
        return text
    if len(text) <= max_len:
        return text
    return text[:max_len] + '...'


def _cell_to_display_text(value: Any, max_len: Optional[int] = None, digit: int = 6) -> str:
    value = _normalize_none_like(value)
    if value is None:
        return ''
    if isinstance(value, (int, float)) or (NUMPY_AVAILABLE and isinstance(value, np.generic)):
        if DFP is not None and hasattr(DFP, 'parse'):
            try:
                return _truncate_text(str(DFP.parse(value, digit=digit)), max_len=max_len)
            except Exception:
                return _truncate_text(str(value), max_len=max_len)
    if isinstance(value, bytes):
        try:
            text = value.decode('utf-8', errors='ignore')
            text = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
            return _truncate_text(text, max_len=max_len)
        except Exception:
            return _truncate_text(str(value), max_len=max_len)
    text = str(value)
    text = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
    return _truncate_text(text, max_len=max_len)



def _safe_sheet_title(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]+', '_', name or 'Sheet')[:31] or 'Sheet'



def _read_text_with_fallbacks(file_path: str, encodings: Optional[List[str]] = None) -> str:
    encodings = encodings or ['utf-8-sig', 'utf-8', 'cp950', 'big5', 'latin1']
    last_error = None
    for enc in encodings:
        try:
            with open(file_path, 'r', encoding=enc, newline='') as f:
                return f.read()
        except Exception as e:
            last_error = e
    raise ValueError(f'無法讀取 CSV 文字內容: {last_error}')



def _detect_csv_dialect(text: str) -> csv.Dialect:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=',;\t|')
    except Exception:
        class _Default(csv.Dialect):
            delimiter = ','
            quotechar = '"'
            doublequote = True
            skipinitialspace = False
            lineterminator = '\n'
            quoting = csv.QUOTE_MINIMAL
        return _Default



def _to_serializable_df(df: 'pd.DataFrame') -> List[List[Any]]:
    safe_df = df.where(pd.notna(df), None)
    rows: List[List[Any]] = []
    for row in safe_df.values.tolist():
        rows.append([_normalize_none_like(v) for v in row])
    return rows



def _df_to_markdown(df: 'pd.DataFrame', max_rows: Optional[int] = None, max_cols: Optional[int] = None) -> str:
    if df is None:
        return ''
    view = df.copy()
    if max_rows is not None:
        view = view.head(max_rows)
    if max_cols is not None:
        view = view.iloc[:, :max_cols]
    try:
        return view.fillna('').to_markdown(index=False)
    except Exception:
        return view.fillna('').to_csv(index=False)


def _compress_markdown_blank_runs(markdown_text: str, blank_token_prefix: str = '[BLANK*') -> str:
    if not markdown_text or "|" not in markdown_text:
        return markdown_text or ""
    lines = markdown_text.splitlines()
    out_lines: List[str] = []
    sep_re = re.compile(r'^\|\s*[:\-\s|]+\|\s*$')
    for line in lines:
        stripped = line.strip()
        if not (stripped.startswith("|") and stripped.endswith("|")):
            out_lines.append(line)
            continue
        if sep_re.match(stripped):
            out_lines.append(line)
            continue
        inner = stripped[1:-1]
        cells = [cell.strip() for cell in inner.split("|")]
        compact_cells: List[str] = []
        blank_run = 0
        for cell in cells:
            if cell == "":
                blank_run += 1
                continue
            if blank_run > 0:
                compact_cells.append(f"{blank_token_prefix}{blank_run}]")
                blank_run = 0
            compact_cells.append(cell)
        if blank_run > 0:
            compact_cells.append(f"{blank_token_prefix}{blank_run}]")
        out_lines.append("| " + " | ".join(compact_cells) + " |")
    return "\n".join(out_lines)



def _normalize_to_xlsx(file_path: str, output_dir: Optional[str] = None) -> str:
    _require_dependencies()
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'檔案不存在: {file_path}')

    src = Path(file_path)
    suffix = src.suffix.lower()
    if suffix == '.xlsx':
        return str(src)
    if suffix == '.csv':
        return _convert_csv_to_xlsx(file_path, output_dir=output_dir)
    if suffix == '.xlsm':
        return _convert_xlsm_to_xlsx(file_path, output_dir=output_dir)
    if suffix == '.xls':
        return _convert_xls_to_xlsx(file_path, output_dir=output_dir)
    raise ValueError(f'不支援的檔案格式: {file_path}。僅支援 .csv / .xls / .xlsm / .xlsx')



def _prepare_output_xlsx_path(file_path: str, output_dir: Optional[str] = None) -> str:
    src = Path(file_path)
    parent = Path(output_dir) if output_dir else src.parent
    parent.mkdir(parents=True, exist_ok=True)
    target = parent / f'{src.stem}.xlsx'
    if str(target) == str(src):
        target = parent / f'{src.stem}_normalized.xlsx'
    if target.exists():
        src_mtime = os.path.getmtime(file_path)
        try:
            if os.path.getmtime(target) >= src_mtime:
                return str(target)
        except Exception:
            pass
        return _rpathrpt(str(target))
    return str(target)


def _convert_csv_to_xlsx(file_path: str, output_dir: Optional[str] = None) -> str:
    out_path = _prepare_output_xlsx_path(file_path, output_dir=output_dir)
    raw_text = _read_text_with_fallbacks(file_path)
    dialect = _detect_csv_dialect(raw_text)
    df = pd.read_csv(io.StringIO(raw_text), sep=dialect.delimiter, engine='python')
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    m_logger.info(f'[_convert_csv_to_xlsx] 轉換成功: {file_path} -> {out_path}')
    return out_path



def _convert_xlsm_to_xlsx(file_path: str, output_dir: Optional[str] = None) -> str:
    out_path = _prepare_output_xlsx_path(file_path, output_dir=output_dir)
    wb = load_workbook(file_path, data_only=False)
    out_wb = Workbook()
    if out_wb.active:
        out_wb.remove(out_wb.active)
    for ws in wb.worksheets:
        new_ws = out_wb.create_sheet(title=_safe_sheet_title(ws.title))
        for row in ws.iter_rows():
            for cell in row:
                new_ws[cell.coordinate].value = cell.value
        for merged in ws.merged_cells.ranges:
            try:
                new_ws.merge_cells(str(merged))
            except Exception:
                pass
    out_wb.save(out_path)
    m_logger.info(f'[_convert_xlsm_to_xlsx] 轉換成功: {file_path} -> {out_path}')
    return out_path



def _convert_xls_to_xlsx(file_path: str, output_dir: Optional[str] = None) -> str:
    out_path = _prepare_output_xlsx_path(file_path, output_dir=output_dir)
    src_path = Path(file_path)

    libreoffice_exe = shutil.which('libreoffice') or shutil.which('soffice')
    if not libreoffice_exe and m_config:
        libreoffice_exe = m_config.get('doc_conversion', {}).get('libreoffice_path', '') or libreoffice_exe
    if libreoffice_exe:
        cmd = [libreoffice_exe, '--headless', '--convert-to', 'xlsx', '--outdir', str(Path(out_path).parent), str(src_path)]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            generated = Path(out_path).parent / f'{src_path.stem}.xlsx'
            if result.returncode == 0 and generated.exists():
                if str(generated) != out_path:
                    shutil.move(str(generated), out_path)
                m_logger.info(f'[_convert_xls_to_xlsx] LibreOffice 轉換成功: {file_path} -> {out_path}')
                return out_path
            m_logger.warning(f'[_convert_xls_to_xlsx] LibreOffice 轉換失敗: {result.stderr}')
        except Exception as e:
            m_logger.warning(f'[_convert_xls_to_xlsx] LibreOffice 轉換失敗: {e}')

    try:
        import xlrd  # type: ignore
        book = xlrd.open_workbook(file_path)
        out_wb = Workbook()
        if out_wb.active:
            out_wb.remove(out_wb.active)
        for sheet in book.sheets():
            ws = out_wb.create_sheet(title=_safe_sheet_title(sheet.name))
            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    ws.cell(row=r + 1, column=c + 1, value=sheet.cell_value(r, c))
        out_wb.save(out_path)
        m_logger.info(f'[_convert_xls_to_xlsx] xlrd 轉換成功: {file_path} -> {out_path}')
        return out_path
    except Exception as e:
        raise ValueError(
            f'無法轉換 XLS 檔案 {file_path}: {e}。請安裝 LibreOffice 或 xlrd。'
        )



def _worksheet_to_dataframe(ws, include_formulas: bool = True) -> 'pd.DataFrame':
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    data: List[List[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_values = []
        for cell in row:
            value = cell.value
            if hasattr(cell, 'data_type') and cell.data_type == 'f' and include_formulas:
                value = value if value is not None else ''
            row_values.append(value)
        data.append(row_values)
    if not data:
        return pd.DataFrame()
    return pd.DataFrame(data)



def _trim_empty_df(df: 'pd.DataFrame') -> Tuple['pd.DataFrame', Optional[Tuple[int, int, int, int]]]:
    if df is None or df.empty:
        return pd.DataFrame(), None
    mask = df.applymap(lambda x: _normalize_none_like(x) is not None)
    if not mask.to_numpy().any():
        return pd.DataFrame(), None
    non_empty_rows = np.where(mask.any(axis=1).to_numpy())[0]
    non_empty_cols = np.where(mask.any(axis=0).to_numpy())[0]
    top = int(non_empty_rows[0])
    bottom = int(non_empty_rows[-1])
    left = int(non_empty_cols[0])
    right = int(non_empty_cols[-1])
    trimmed = df.iloc[top:bottom + 1, left:right + 1].copy()
    return trimmed, (top + 1, left + 1, bottom + 1, right + 1)


def _prune_empty_rows_cols(df: 'pd.DataFrame') -> 'pd.DataFrame':
    """移除資料框內全空白的列與欄，避免輸出大量空白區。"""
    if df is None or df.empty:
        return pd.DataFrame()
    mask = df.applymap(lambda x: _normalize_none_like(x) is not None)
    if not mask.to_numpy().any():
        return pd.DataFrame()
    keep_rows = mask.any(axis=1)
    keep_cols = mask.any(axis=0)
    return df.loc[keep_rows, keep_cols].copy()

def _extract_sheet_images(
        ws,
        include_images: bool = True,
        to_base64: bool = True,
        image_id_start: int = 0
    ) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    images = list(getattr(ws, '_images', []) or [])
    for idx, img in enumerate(images):
        info: Dict[str, Any] = {
            'local_index': idx,
            'sheet_name': ws.title,
        }
        image_id = None
        anchor = getattr(img, 'anchor', None)
        anchor_cell = None
        top_row = None
        left_col = None
        if anchor is not None:
            _from = getattr(anchor, '_from', None)
            if _from is not None:
                top_row = getattr(_from, 'row', None)
                left_col = getattr(_from, 'col', None)
                if top_row is not None and left_col is not None:
                    anchor_cell = f'{_colnum_to_letter(left_col + 1)}{top_row + 1}'
        info['anchor_cell'] = anchor_cell
        info['anchor_row'] = (top_row + 1) if top_row is not None else None
        info['anchor_col'] = (left_col + 1) if left_col is not None else None
        info['type'] = 'cell_image' if anchor_cell else 'sheet_image'

        placeholder = None
        if include_images:
            image_id = image_id_start + idx
            placeholder = _make_image_placeholder(image_id)
            info['id'] = image_id
            info['placeholder'] = placeholder

        raw_bytes = None
        try:
            if hasattr(img, '_data'):
                raw_bytes = img._data()
            elif hasattr(img, 'ref') and isinstance(img.ref, (bytes, bytearray)):
                raw_bytes = bytes(img.ref)
        except Exception:
            raw_bytes = None

        if raw_bytes is not None:
            info['size_bytes'] = len(raw_bytes)
            if to_base64:
                info['base64'] = base64.b64encode(raw_bytes).decode('utf-8')
            else:
                info['bytes'] = raw_bytes
        results.append(info)
    return results



def _extract_formula_cells(ws) -> List[Dict[str, Any]]:
    formulas: List[Dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            try:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    formulas.append({
                        'sheet_name': ws.title,
                        'cell': cell.coordinate,
                        'formula': cell.value,
                    })
            except Exception:
                continue
    return formulas


def _expand_sqref_cells(sqref: Any) -> List[str]:
    cells: List[str] = []
    raw = str(sqref or '').strip()
    if not raw:
        return cells
    for token in raw.split():
        token = token.strip()
        if not token:
            continue
        if ':' in token:
            start_ref, end_ref = token.split(':', 1)
            start = _parse_excel_cell_ref(start_ref)
            end = _parse_excel_cell_ref(end_ref)
            if not start or not end:
                continue
            r1, c1 = start
            r2, c2 = end
            if r1 > r2:
                r1, r2 = r2, r1
            if c1 > c2:
                c1, c2 = c2, c1
            for row_idx in range(r1, r2 + 1):
                for col_idx in range(c1, c2 + 1):
                    cells.append(f'{_colnum_to_letter(col_idx)}{row_idx}')
        else:
            parsed = _parse_excel_cell_ref(token)
            if parsed:
                cells.append(f'{_colnum_to_letter(parsed[1])}{parsed[0]}')
    return cells


def _parse_inline_list_options(formula1: Any) -> List[str]:
    text = str(formula1 or '').strip()
    if not text or text.startswith('='):
        return []
    if len(text) >= 2 and text[0] == text[-1] == '"':
        text = text[1:-1]
    options: List[str] = []
    for part in text.split(','):
        item = part.strip().strip('"')
        if item:
            options.append(item)
    return options


def _resolve_list_options_from_formula(wb, ws, formula1: Any) -> List[str]:
    text = str(formula1 or '').strip()
    if not text.startswith('='):
        return _parse_inline_list_options(text)
    expr = text[1:].strip()
    target_ws = ws
    range_part = expr
    sheet_match = re.match(r"^(?:'([^']+)'|([^!]+))!(.+)$", expr)
    if sheet_match:
        sheet_name = sheet_match.group(1) or sheet_match.group(2)
        range_part = sheet_match.group(3)
        if sheet_name and sheet_name in wb.sheetnames:
            target_ws = wb[sheet_name]
    range_part = range_part.replace('$', '')
    if ':' not in range_part:
        cell_val = target_ws[range_part].value
        if cell_val in (None, ''):
            return []
        return [str(cell_val).strip()]
    start_ref, end_ref = range_part.split(':', 1)
    start = _parse_excel_cell_ref(start_ref)
    end = _parse_excel_cell_ref(end_ref)
    if not start or not end:
        return []
    r1, c1 = start
    r2, c2 = end
    if r1 > r2:
        r1, r2 = r2, r1
    if c1 > c2:
        c1, c2 = c2, c1
    options: List[str] = []
    for row_idx in range(r1, r2 + 1):
        for col_idx in range(c1, c2 + 1):
            cell_val = target_ws.cell(row=row_idx, column=col_idx).value
            if cell_val in (None, ''):
                continue
            options.append(str(cell_val).strip())
    return options


def _extract_data_validations(ws, wb) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    dvs = getattr(ws, 'data_validations', None)
    if not dvs:
        return results
    for dv in getattr(dvs, 'dataValidation', []) or []:
        dv_type = str(getattr(dv, 'type', '') or '').strip().lower()
        formula1 = getattr(dv, 'formula1', None)
        formula2 = getattr(dv, 'formula2', None)
        sqref = str(getattr(dv, 'sqref', '') or '').strip()
        cells = _expand_sqref_cells(sqref)
        options: List[str] = []
        options_source = ''
        if dv_type == 'list':
            if str(formula1 or '').strip().startswith('='):
                options = _resolve_list_options_from_formula(wb, ws, formula1)
                options_source = 'formula_ref' if options else 'formula_ref_unresolved'
            else:
                options = _parse_inline_list_options(formula1)
                options_source = 'inline_list'
        results.append({
            'sheet_name': ws.title,
            'type': dv_type or None,
            'operator': getattr(dv, 'operator', None),
            'allow_blank': bool(getattr(dv, 'allow_blank', False)),
            'sqref': sqref,
            'cells': cells,
            'formula1': formula1,
            'formula2': formula2,
            'options': options,
            'options_source': options_source,
        })
    return results


def _build_validation_by_cell(data_validations: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    by_cell: Dict[str, Dict[str, Any]] = {}
    for item in data_validations or []:
        cell_entry = {
            'type': item.get('type'),
            'operator': item.get('operator'),
            'allow_blank': item.get('allow_blank'),
            'options': item.get('options') or [],
            'options_source': item.get('options_source'),
            'formula1': item.get('formula1'),
            'formula2': item.get('formula2'),
            'sqref': item.get('sqref'),
        }
        for cell_ref in item.get('cells') or []:
            by_cell[str(cell_ref).upper()] = dict(cell_entry)
    return by_cell


def _read_cell_display_value(sheet_df: Any, cell_ref: str) -> str:
    parsed = _parse_excel_cell_ref(cell_ref)
    if parsed is None or sheet_df is None or getattr(sheet_df, 'empty', True):
        return ''
    row, col = parsed
    try:
        value = sheet_df.iloc[row - 1, col - 1]
        return _cell_to_display_text(value, max_len=200)
    except Exception:
        return ''


def _validation_options_sig(entry: Dict[str, Any]) -> Any:
    options = entry.get('options') or []
    if options:
        return tuple(str(o) for o in options)
    formula1 = entry.get('formula1')
    if formula1:
        return ('formula', str(formula1))
    return tuple()


def _join_validation_options(options: List[Any]) -> str:
    return '、'.join(str(o) for o in (options or []) if str(o).strip())


def _validation_token_value(value: str) -> str:
    text = str(value or '').strip()
    return text.replace('|', '/').replace(']', ')')


def _cells_to_a1_range(cells: List[Dict[str, Any]]) -> str:
    if not cells:
        return ''
    rows = [int(c['row']) for c in cells]
    cols = [int(c['col']) for c in cells]
    min_r, max_r = min(rows), max(rows)
    min_c, max_c = min(cols), max(cols)
    return f"{_colnum_to_letter(min_c)}{min_r}:{_colnum_to_letter(max_c)}{max_r}"


def _validation_cells_in_table(
        table: Dict[str, Any],
        unit: Dict[str, Any],
        validation_by_cell: Dict[str, Dict[str, Any]],
        emitted: Optional[Set[str]] = None,
        include_same_row: bool = True,
    ) -> List[Dict[str, Any]]:
    top = int(table.get('top_row') or 0)
    bottom = int(table.get('bottom_row') or 0)
    left = int(table.get('left_col') or 0)
    right = int(table.get('right_col') or 0)
    sheet_df = unit.get('sheet_df')
    used = emitted if emitted is not None else set()
    cells: List[Dict[str, Any]] = []
    for cell_ref, entry in (validation_by_cell or {}).items():
        cell_key = str(cell_ref).upper()
        if cell_key in used:
            continue
        if str(entry.get('type', '')).lower() != 'list':
            continue
        parsed = _parse_excel_cell_ref(cell_ref)
        if not parsed:
            continue
        row, col = parsed
        in_bbox = row >= top and row <= bottom and col >= left and col <= right
        in_row = bool(include_same_row) and row >= top and row <= bottom
        if not in_bbox and not in_row:
            continue
        options = list(entry.get('options') or [])
        cells.append({
            'cell_ref': cell_key,
            'row': row,
            'col': col,
            'options': options,
            'options_sig': _validation_options_sig(entry),
            'options_source': entry.get('options_source'),
            'selected': _read_cell_display_value(sheet_df, cell_ref),
            'entry': entry,
            'in_bbox': in_bbox,
        })
    return cells


def _cluster_validation_cells(
        cells: List[Dict[str, Any]],
        min_cells: int = 2,
        example_max: int = 3,
    ) -> Dict[str, List[Any]]:
    consumed: Set[str] = set()
    row_uniform: List[Dict[str, Any]] = []
    col_uniform: List[Dict[str, Any]] = []
    row_mixed: List[Dict[str, Any]] = []
    col_mixed: List[Dict[str, Any]] = []

    by_row: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for cell in cells:
        by_row[int(cell['row'])].append(cell)

    for row_idx, row_cells in sorted(by_row.items()):
        if len(row_cells) < min_cells:
            continue
        sigs = {cell['options_sig'] for cell in row_cells}
        if len(sigs) == 1:
            row_uniform.append({'row': row_idx, 'cells': row_cells, 'options_sig': next(iter(sigs))})
            for cell in row_cells:
                consumed.add(cell['cell_ref'])
        elif len(sigs) > 1:
            row_mixed.append({'row': row_idx, 'cells': row_cells})
            for cell in row_cells:
                consumed.add(cell['cell_ref'])

    by_col: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for cell in cells:
        if cell['cell_ref'] in consumed:
            continue
        by_col[int(cell['col'])].append(cell)

    for col_idx, col_cells in sorted(by_col.items()):
        if len(col_cells) < min_cells:
            continue
        sigs = {cell['options_sig'] for cell in col_cells}
        if len(sigs) == 1:
            col_uniform.append({'col': col_idx, 'cells': col_cells, 'options_sig': next(iter(sigs))})
            for cell in col_cells:
                consumed.add(cell['cell_ref'])
        elif len(sigs) > 1:
            col_mixed.append({'col': col_idx, 'cells': col_cells})
            for cell in col_cells:
                consumed.add(cell['cell_ref'])

    scattered = [cell for cell in cells if cell['cell_ref'] not in consumed]
    return {
        'scattered': scattered,
        'row_uniform': row_uniform,
        'col_uniform': col_uniform,
        'row_mixed': row_mixed,
        'col_mixed': col_mixed,
        'example_max': example_max,
    }


def _format_validation_options_suffix(cell: Dict[str, Any]) -> str:
    options = cell.get('options') or []
    options_text = _join_validation_options(options)
    if options_text:
        return f'選項: {options_text}'
    if cell.get('options_source') == 'formula_ref_unresolved':
        return '選項: (公式參照未解析, options_source=formula_ref_unresolved)'
    return '選項: (未解析)'


def _format_validation_scattered_line(cell: Dict[str, Any]) -> str:
    options = cell.get('options') or []
    selected = _validation_token_value(cell.get('selected', ''))
    token_parts = [
        '下拉選單',
        'list',
        f"anchor={cell.get('cell_ref', '')}",
        f"option_count={len(options)}",
    ]
    if selected:
        token_parts.append(f"selected={selected}")
    return f"[{'|'.join(token_parts)}] {_format_validation_options_suffix(cell)}"


def _format_validation_uniform_line(group: Dict[str, Any], axis: str) -> str:
    cells = group.get('cells') or []
    if not cells:
        return ''
    sample = cells[0]
    options = sample.get('options') or []
    options_text = _join_validation_options(options)
    selected_values = [_validation_token_value(c.get('selected', '')) for c in cells if c.get('selected')]
    unique_selected = []
    for value in selected_values:
        if value and value not in unique_selected:
            unique_selected.append(value)
    if len(unique_selected) == 1:
        selected_note = f'目前皆為「{unique_selected[0]}」'
        selected_token = unique_selected[0]
    elif unique_selected:
        selected_note = f"目前選值: {'、'.join(unique_selected)}"
        selected_token = unique_selected[0]
    else:
        selected_note = '目前無選值'
        selected_token = ''
    axis_key = 'row' if axis == 'row' else 'col'
    axis_val = int(group.get(axis_key) or (cells[0]['row'] if axis == 'row' else cells[0]['col']))
    token_parts = [
        '下拉選單',
        'list',
        f"{axis_key}={axis_val}",
        f"range={_cells_to_a1_range(cells)}",
        f"option_count={len(options)}",
        f"cells={len(cells)}",
    ]
    if selected_token:
        token_parts.append(f"selected={selected_token}")
    axis_label = '全列' if axis == 'row' else '全欄'
    if options_text:
        suffix = f'{axis_label}共用選項: {options_text}；{selected_note}'
    else:
        suffix = f'{axis_label}共用下拉選單；{selected_note}'
    return f"[{'|'.join(token_parts)}] {suffix}"


def _format_validation_mixed_line(group: Dict[str, Any], axis: str, example_max: int = 3) -> str:
    cells = group.get('cells') or []
    if not cells:
        return ''
    axis_key = 'row' if axis == 'row' else 'col'
    axis_val = int(group.get(axis_key) or (cells[0]['row'] if axis == 'row' else cells[0]['col']))
    examples = []
    for cell in cells[:max(1, int(example_max or 3))]:
        option_count = len(cell.get('options') or [])
        selected = _validation_token_value(cell.get('selected', ''))
        if selected:
            examples.append(f"{cell.get('cell_ref')}({option_count}項,選「{selected}」)")
        else:
            examples.append(f"{cell.get('cell_ref')}({option_count}項)")
    token_parts = [
        '下拉選單',
        'list',
        f"{axis_key}={axis_val}",
        'pattern=mixed',
        f"cells={len(cells)}",
    ]
    return f"[{'|'.join(token_parts)}] 例: {'; '.join(examples)}"


def _build_validation_annotation_lines(
        table: Dict[str, Any],
        unit: Dict[str, Any],
        validation_by_cell: Dict[str, Dict[str, Any]],
        enable_validation_annotations: bool = True,
        validation_cluster_min_cells: int = 2,
        validation_mixed_example_max: int = 3,
        validation_include_same_row: bool = True,
        emitted_validation_cells: Optional[Set[str]] = None,
        **kwargs,
    ) -> List[str]:
    cells = _validation_cells_in_table(
        table,
        unit,
        validation_by_cell,
        emitted=emitted_validation_cells,
        include_same_row=validation_include_same_row,
    )
    if not cells:
        return []
    bbox_cells = [cell for cell in cells if cell.get('in_bbox')]
    row_assoc_cells = [cell for cell in cells if not cell.get('in_bbox')]
    clustered = _cluster_validation_cells(
        bbox_cells,
        min_cells=max(1, int(validation_cluster_min_cells or 2)),
        example_max=max(1, int(validation_mixed_example_max or 3)),
    )
    lines: List[str] = []
    for group in clustered.get('row_uniform') or []:
        line = _format_validation_uniform_line(group, axis='row')
        if line:
            lines.append(line)
    for group in clustered.get('col_uniform') or []:
        line = _format_validation_uniform_line(group, axis='col')
        if line:
            lines.append(line)
    example_max = int(clustered.get('example_max') or 3)
    for group in clustered.get('row_mixed') or []:
        line = _format_validation_mixed_line(group, axis='row', example_max=example_max)
        if line:
            lines.append(line)
    for group in clustered.get('col_mixed') or []:
        line = _format_validation_mixed_line(group, axis='col', example_max=example_max)
        if line:
            lines.append(line)
    for cell in clustered.get('scattered') or []:
        line = _format_validation_scattered_line(cell)
        if line:
            lines.append(line)
    for cell in row_assoc_cells:
        line = _format_validation_scattered_line(cell)
        if line:
            lines.append(line)
    if emitted_validation_cells is not None:
        for cell in cells:
            emitted_validation_cells.add(str(cell.get('cell_ref', '')).upper())
    return lines


def _append_validation_annotations_to_unit_text(
        unit: Dict[str, Any],
        **kwargs,
    ) -> None:
    """extract 階段：將下拉選單 annotation 寫入 unit_text（TB6 preview_text 同源）。"""
    validation_by_cell = unit.get('validation_by_cell') or {}
    if not validation_by_cell:
        return
    sheet_bbox = unit.get('sheet_bbox')
    sheet_df = unit.get('sheet_df')
    if sheet_bbox and len(sheet_bbox) >= 4:
        top, left, bottom, right = (int(sheet_bbox[0]), int(sheet_bbox[1]), int(sheet_bbox[2]), int(sheet_bbox[3]))
    elif sheet_df is not None and not getattr(sheet_df, 'empty', True):
        top, left = 1, 1
        bottom, right = int(sheet_df.shape[0]), int(sheet_df.shape[1])
    else:
        return
    table = {
        'table_id': 0,
        'top_row': top,
        'left_col': left,
        'bottom_row': bottom,
        'right_col': right,
    }
    lines = _build_validation_annotation_lines(
        table=table,
        unit=unit,
        validation_by_cell=validation_by_cell,
        enable_validation_annotations=True,
        **kwargs,
    )
    if not lines:
        return
    block = '## 下拉選單\n' + '\n'.join(lines)
    unit_text = str(unit.get('unit_text', '') or '')
    image_anchor = re.search(r'\n(\[(?:IMAGE_PLACEHOLDER|PDF_IMAGE)_\d+\])', unit_text)
    if image_anchor:
        pos = image_anchor.start()
        unit['unit_text'] = unit_text[:pos].rstrip() + '\n\n' + block + unit_text[pos:]
    else:
        unit['unit_text'] = unit_text.rstrip() + '\n\n' + block


def _colnum_to_letter(col_num: int) -> str:
    result = ''
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        result = chr(65 + rem) + result
    return result



def _make_image_placeholder(image_id: int) -> str:
    return f'[IMAGE_PLACEHOLDER_{image_id}]'


def _sheet_header(sheet_name: str, is_first: bool = False) -> str:
    header = f'#Sheet: {sheet_name}'
    return header if is_first else f'---\n\n{header}'


def _build_sheet_image_lines(sheet_images: List[Dict[str, Any]]) -> List[str]:
    lines: List[str] = []
    for img in sheet_images or []:
        placeholder = str(img.get('placeholder', '') or '').strip()
        if not placeholder:
            continue
        anchor_cell = str(img.get('anchor_cell', '') or '').strip()
        if anchor_cell:
            lines.append(f'{placeholder} [anchor={anchor_cell}]')
        else:
            lines.append(f'{placeholder} [anchor=floating]')
    return lines



def _norm_zip_path(path: str) -> str:
    p = (path or '').replace('\\', '/')
    p = posixpath.normpath(p)
    return p.lstrip('./')


def _resolve_target(base_path: str, target: str) -> str:
    return _norm_zip_path(posixpath.join(posixpath.dirname(base_path), target or ''))


def _safe_int(v: Optional[str], default: int = 0) -> int:
    try:
        return int(str(v or '').strip())
    except Exception:
        return default


def _anchor_to_dict(anchor_node: Any, ns: Dict[str, str]) -> Dict[str, int]:
    if anchor_node is None:
        return {'row': 1, 'col': 1, 'row_off': 0, 'col_off': 0}
    return {
        'row': _safe_int(anchor_node.findtext('xdr:row', default='0', namespaces=ns), 0) + 1,
        'col': _safe_int(anchor_node.findtext('xdr:col', default='0', namespaces=ns), 0) + 1,
        'row_off': _safe_int(anchor_node.findtext('xdr:rowOff', default='0', namespaces=ns), 0),
        'col_off': _safe_int(anchor_node.findtext('xdr:colOff', default='0', namespaces=ns), 0),
    }


def _a1(row_idx: int, col_idx: int) -> str:
    return f"{_colnum_to_letter(max(1, col_idx))}{max(1, row_idx)}"


def extract_ooxml_sheet_textboxes(file_path: str) -> Dict[str, Any]:
    ns = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'rel': 'http://schemas.openxmlformats.org/package/2006/relationships',
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    }

    result: Dict[str, Any] = {
        'ns': ns,
        'sheets': [],
    }
    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            workbook_xml = ET.fromstring(zf.read('xl/workbook.xml'))
            wb_rels_xml = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))

            wb_rel_map: Dict[str, str] = {}
            for rel in wb_rels_xml.findall('rel:Relationship', ns):
                rid = rel.attrib.get('Id')
                target = rel.attrib.get('Target')
                if rid and target:
                    wb_rel_map[rid] = _resolve_target('xl/workbook.xml', target)

            sheets: List[Tuple[str, str]] = []
            for sheet in workbook_xml.findall('main:sheets/main:sheet', ns):
                name = sheet.attrib.get('name', '')
                rid = sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
                sheet_path = wb_rel_map.get(rid, '')
                if name and sheet_path:
                    sheets.append((name, sheet_path))

            for sheet_name, sheet_path in sheets:
                if sheet_path not in zf.namelist():
                    continue
                sheet_xml = ET.fromstring(zf.read(sheet_path))
                drawing_ids = [n.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id') for n in sheet_xml.findall('main:drawing', ns)]
                drawing_ids = [x for x in drawing_ids if x]
                if not drawing_ids:
                    continue

                sheet_rels_path = _norm_zip_path(posixpath.join(posixpath.dirname(sheet_path), '_rels', posixpath.basename(sheet_path) + '.rels'))
                if sheet_rels_path not in zf.namelist():
                    continue
                sheet_rels_xml = ET.fromstring(zf.read(sheet_rels_path))
                sheet_rel_map: Dict[str, str] = {}
                for rel in sheet_rels_xml.findall('rel:Relationship', ns):
                    rid = rel.attrib.get('Id')
                    target = rel.attrib.get('Target')
                    if rid and target:
                        sheet_rel_map[rid] = _resolve_target(sheet_path, target)

                drawing_refs: List[Dict[str, Any]] = []
                for draw_rid in drawing_ids:
                    drawing_path = sheet_rel_map.get(draw_rid, '')
                    if not drawing_path or drawing_path not in zf.namelist():
                        continue
                    drawing_xml = ET.fromstring(zf.read(drawing_path))
                    drawing_refs.append({
                        'drawing_rel_id': draw_rid,
                        'drawing_path': drawing_path,
                        'drawing_xml': drawing_xml,
                    })
                if drawing_refs:
                    result['sheets'].append({
                        'sheet_name': sheet_name,
                        'sheet_path': sheet_path,
                        'drawing_refs': drawing_refs,
                    })
    except Exception as e:
        m_logger.warning(f"[extract_ooxml_sheet_textboxes] failed: {e}")
    return result


def _extract_xlsx_textboxes(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
    try:
        ooxml_data = extract_ooxml_sheet_textboxes(file_path) or {}
        ns = ooxml_data.get('ns') or {}
        sheets = ooxml_data.get('sheets') or []
        result: Dict[str, List[Dict[str, Any]]] = {}

        for sheet_info in sheets:
            sheet_name = str(sheet_info.get('sheet_name', '') or '')
            if not sheet_name:
                continue
            drawing_refs = sheet_info.get('drawing_refs') or []
            textboxes: List[Dict[str, Any]] = []
            for drawing_ref in drawing_refs:
                draw_rid = str(drawing_ref.get('drawing_rel_id', '') or '')
                drawing_xml = drawing_ref.get('drawing_xml')
                if drawing_xml is None:
                    continue
                anchors = list(drawing_xml.findall('xdr:twoCellAnchor', ns)) + list(drawing_xml.findall('xdr:oneCellAnchor', ns))
                for anchor in anchors:
                    frm = _anchor_to_dict(anchor.find('xdr:from', ns), ns)
                    to_node = anchor.find('xdr:to', ns)
                    to_dict = _anchor_to_dict(to_node, ns) if to_node is not None else dict(frm)
                    texts = [t.text.strip() for t in anchor.findall('.//a:t', ns) if t.text and t.text.strip()]
                    if not texts:
                        continue
                    textboxes.append({
                        'textbox_id': len(textboxes),
                        'drawing_rel_id': draw_rid,
                        'anchor_type': 'twoCellAnchor' if str(anchor.tag).endswith('twoCellAnchor') else 'oneCellAnchor',
                        'anchor_from': frm,
                        'anchor_to': to_dict,
                        'a1_range': f"{_a1(frm['row'], frm['col'])}:{_a1(to_dict['row'], to_dict['col'])}",
                        'text': '\n'.join(texts).strip(),
                        'line_count': len(texts),
                    })
            if textboxes:
                result[sheet_name] = textboxes
        return result
    except Exception as e:
        m_logger.warning(f"[_extract_xlsx_textboxes] failed: {e}")
        return {}


def _find_connected_components(mask: 'np.ndarray') -> List[Tuple[int, int, int, int]]:
    rows, cols = mask.shape
    visited = np.zeros_like(mask, dtype=bool)
    components: List[Tuple[int, int, int, int]] = []
    directions = [(-1, 0), (1, 0), (0, -1), (0, 1)]

    for r in range(rows):
        for c in range(cols):
            if not mask[r, c] or visited[r, c]:
                continue
            stack = [(r, c)]
            visited[r, c] = True
            min_r = max_r = r
            min_c = max_c = c
            while stack:
                cr, cc = stack.pop()
                min_r = min(min_r, cr)
                max_r = max(max_r, cr)
                min_c = min(min_c, cc)
                max_c = max(max_c, cc)
                for dr, dc in directions:
                    nr, nc = cr + dr, cc + dc
                    if 0 <= nr < rows and 0 <= nc < cols and mask[nr, nc] and not visited[nr, nc]:
                        visited[nr, nc] = True
                        stack.append((nr, nc))
            components.append((min_r, min_c, max_r, max_c))
    return components



def _detect_tables_from_sheet_df(sheet_df: 'pd.DataFrame', min_non_empty_cells: int = 1) -> List[Dict[str, Any]]:
    if sheet_df is None or sheet_df.empty:
        return []
    mask_df = sheet_df.applymap(lambda x: _normalize_none_like(x) is not None)
    mask = mask_df.to_numpy(dtype=bool)
    if not mask.any():
        return []

    boxes = _find_connected_components(mask)
    tables: List[Dict[str, Any]] = []
    for idx, (min_r, min_c, max_r, max_c) in enumerate(boxes):
        sub_df = sheet_df.iloc[min_r:max_r + 1, min_c:max_c + 1].copy()
        non_empty_count = int(sub_df.applymap(lambda x: _normalize_none_like(x) is not None).to_numpy().sum())
        if non_empty_count < min_non_empty_cells:
            continue
        tables.append({
            'table_id': idx,
            'top_row': min_r + 1,
            'left_col': min_c + 1,
            'bottom_row': max_r + 1,
            'right_col': max_c + 1,
            'shape': [int(sub_df.shape[0]), int(sub_df.shape[1])],
            'table_df': sub_df,
            'table_values': _to_serializable_df(sub_df),
            'table_markdown': _df_to_markdown(sub_df),
            'non_empty_count': non_empty_count,
        })
    tables.sort(key=lambda x: (x['top_row'], x['left_col']))
    for idx, t in enumerate(tables):
        t['table_id'] = idx
    return tables


def _tables_from_bboxes(sheet_df: 'pd.DataFrame', boxes: List[Dict[str, Any]], min_non_empty_cells: int = 1) -> List[Dict[str, Any]]:
    if sheet_df is None or sheet_df.empty:
        return []
    tables: List[Dict[str, Any]] = []
    for item in boxes or []:
        try:
            top = int(item['top_row'])
            left = int(item['left_col'])
            bottom = int(item['bottom_row'])
            right = int(item['right_col'])
        except Exception:
            continue
        if top < 1 or left < 1 or bottom < top or right < left:
            continue
        sub_df = sheet_df.iloc[top - 1:bottom, left - 1:right].copy()
        non_empty = int(sub_df.applymap(lambda x: _normalize_none_like(x) is not None).to_numpy().sum())
        if non_empty < min_non_empty_cells:
            continue
        tables.append({
            'table_id': len(tables),
            'top_row': top,
            'left_col': left,
            'bottom_row': bottom,
            'right_col': right,
            'shape': [int(sub_df.shape[0]), int(sub_df.shape[1])],
            'table_df': sub_df,
            'table_values': _to_serializable_df(sub_df),
            'table_markdown': _df_to_markdown(sub_df),
            'non_empty_count': non_empty,
        })
    return tables



def _resolve_table_detector_mode(
        table_detector: Optional[str] = None,
        feature_scan: bool = False,
    ) -> str:
    """Resolve table detection backend. feature_scan=True is kept as alias for inspector."""
    if table_detector is not None:
        mode = str(table_detector).strip().lower()
        if mode in ('inspector', 'excel_inspector'):
            return 'inspector'
        return 'native'
    if feature_scan:
        return 'inspector'
    return 'native'


def _load_inspector_table_detector():
    try:
        from excel_inspector.table_detector import detect_tables_from_sheet
        return detect_tables_from_sheet
    except Exception:
        try:
            from dataProcess.excel_inspector.table_detector import detect_tables_from_sheet
            return detect_tables_from_sheet
        except Exception:
            return None


def _detect_tables_for_sheet(
        sheet_name: str,
        sheet_non_empty_df: 'pd.DataFrame',
        *,
        table_detector: str = 'native',
        inspector_table_params: Optional[Dict[str, Any]] = None,
        min_non_empty_cells: int = 1,
    ) -> Tuple[List[Dict[str, Any]], str]:
    if table_detector != 'inspector':
        return _detect_tables_from_sheet_df(sheet_non_empty_df, min_non_empty_cells=min_non_empty_cells), 'native'

    inspector_detect = _load_inspector_table_detector()
    if inspector_detect is None:
        m_logger.warning('[extract] excel_inspector 不可用，fallback 至 native table detector')
        return _detect_tables_from_sheet_df(sheet_non_empty_df, min_non_empty_cells=min_non_empty_cells), 'native'

    params = inspector_table_params or {}
    min_non_empty = int(params.get('min_non_empty', params.get('min_non_empty_cells', 4)))
    min_density = float(params.get('min_density', 0.12))
    bridge_gap = int(params.get('bridge_gap', params.get('max_gap_cols', params.get('max_gap_rows', 1))))
    detected = inspector_detect(
        sheet_name,
        sheet_non_empty_df,
        min_non_empty=min_non_empty,
        min_density=min_density,
        bridge_gap=bridge_gap,
    )
    boxes = [{
        'top_row': t.get('top_row'),
        'left_col': t.get('left_col'),
        'bottom_row': t.get('bottom_row'),
        'right_col': t.get('right_col'),
    } for t in detected]
    tables = _tables_from_bboxes(sheet_non_empty_df, boxes, min_non_empty_cells=min_non_empty_cells)
    if not tables and sheet_non_empty_df is not None and not sheet_non_empty_df.empty:
        tables = _detect_tables_from_sheet_df(sheet_non_empty_df, min_non_empty_cells=min_non_empty_cells)
    return tables, 'inspector'


def _collect_floating_images(
        sheet_images: List[Dict[str, Any]],
        tables: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
    floating_images: List[Dict[str, Any]] = []
    for img in sheet_images or []:
        placeholder = img.get('placeholder')
        if not placeholder:
            continue
        in_table = False
        if img.get('type') == 'cell_image':
            row = img.get('anchor_row')
            col = img.get('anchor_col')
            if row is not None and col is not None:
                for table in tables or []:
                    if table['top_row'] <= row <= table['bottom_row'] and table['left_col'] <= col <= table['right_col']:
                        in_table = True
                        break
        if not in_table:
            floating_images.append(img)
    return floating_images


def _append_floating_image_units(
        unit_paras: List[Dict[str, Any]],
        sheet_name: str,
        sheet_index: int,
        floating_images: List[Dict[str, Any]],
    ) -> None:
    for img in floating_images or []:
        placeholder = str(img.get('placeholder', '') or '').strip()
        if not placeholder:
            continue
        unit_paras.append({
            'unit_text': '\n'.join([_sheet_header(sheet_name, is_first=False), placeholder]),
            'indent_level': 0,
            'order': len(unit_paras),
            'unit_type': 'sheet_image',
            'sheet_name': sheet_name,
            'sheet_index': sheet_index,
            'images': [img],
            'sheet_images': [img],
            'structure_chars': [
                {'type': 'sheet', 'sheet_name': sheet_name},
                {'type': 'image', 'image_id': img.get('id'), 'anchor_cell': img.get('anchor_cell')},
            ],
        })



def _guess_headers(table_df: 'pd.DataFrame') -> List[str]:
    if table_df is None or table_df.empty:
        return []
    first_row = table_df.iloc[0].tolist()
    headers: List[str] = []
    for i, v in enumerate(first_row):
        txt = _cell_to_display_text(v).strip()
        headers.append(txt if txt else f'col_{i + 1}')
    return headers



def _segment_images_for_table(table_box: Dict[str, Any], sheet_images: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    table_images: List[Dict[str, Any]] = []
    floating_images: List[Dict[str, Any]] = []
    for img in sheet_images or []:
        if img.get('type') == 'cell_image':
            row = img.get('anchor_row')
            col = img.get('anchor_col')
            if row is not None and col is not None:
                if table_box['top_row'] <= row <= table_box['bottom_row'] and table_box['left_col'] <= col <= table_box['right_col']:
                    table_images.append(img)
        else:
            floating_images.append(img)
    return table_images, floating_images



def _build_sheet_unit(sheet_name: str, sheet_index: int, sheet_df: 'pd.DataFrame', sheet_non_empty_df: 'pd.DataFrame',
                      sheet_bbox: Optional[Tuple[int, int, int, int]], formulas: List[Dict[str, Any]],
                      sheet_images: List[Dict[str, Any]]) -> Dict[str, Any]:
    render_df = sheet_non_empty_df if sheet_non_empty_df is not None and not sheet_non_empty_df.empty else sheet_df
    render_df = _prune_empty_rows_cols(render_df)
    sheet_md = _compress_markdown_blank_runs(_df_to_markdown(render_df))
    sheet_non_empty_md = _df_to_markdown(_prune_empty_rows_cols(sheet_non_empty_df)) if sheet_non_empty_df is not None else ''
    unit_text_parts = [_sheet_header(sheet_name, is_first=(sheet_index == 0))]
    if sheet_md:
        unit_text_parts.append(sheet_md)
    else:
        unit_text_parts.append('[EMPTY_SHEET]')
    return {
        'unit_text': '\n\n'.join(unit_text_parts),
        'indent_level': 0,
        'order': sheet_index,
        'structure_chars': [{'type': 'sheet', 'sheet_name': sheet_name}],
        'sheet_name': sheet_name,
        'sheet_index': sheet_index,
        'sheet_df': sheet_df,
        'sheet_values': _to_serializable_df(sheet_df) if sheet_df is not None else [],
        'sheet_markdown': sheet_md,
        'sheet_non_empty_df': sheet_non_empty_df,
        'sheet_non_empty_values': _to_serializable_df(sheet_non_empty_df) if sheet_non_empty_df is not None and not sheet_non_empty_df.empty else [],
        'sheet_non_empty_markdown': sheet_non_empty_md,
        'sheet_bbox': sheet_bbox,
        'formula_cells': formulas,
        'sheet_images': sheet_images,
    }



def _sanitize_metadata(metadata: Dict[str, Any]) -> Dict[str, Any]:
    display = {}
    for key, value in metadata.items():
        if key in {'sheets'}:
            summary = []
            for item in value:
                summary.append({
                    'sheet_name': item.get('sheet_name'),
                    'shape': item.get('shape'),
                    'non_empty_shape': item.get('non_empty_shape'),
                    'table_count': item.get('table_count'),
                    'formula_count': item.get('formula_count'),
                    'data_validation_count': item.get('data_validation_count'),
                    'image_count': item.get('image_count'),
                })
            display[key] = summary
        elif key == 'data_validations' and isinstance(value, list):
            display[key] = [{
                'sheet_name': item.get('sheet_name'),
                'type': item.get('type'),
                'sqref': item.get('sqref'),
                'cells': item.get('cells'),
                'options': item.get('options'),
                'options_source': item.get('options_source'),
            } for item in value]
        elif key == 'images' and isinstance(value, list):
            summarized = []
            for img in value:
                summarized.append({
                    'id': img.get('id'),
                    'placeholder': img.get('placeholder'),
                    'sheet_name': img.get('sheet_name'),
                    'anchor_cell': img.get('anchor_cell'),
                    'type': img.get('type'),
                    'size_bytes': img.get('size_bytes'),
                    'has_base64': 'base64' in img,
                    'has_bytes': 'bytes' in img,
                })
            display[key] = summarized
        else:
            display[key] = value
    return display


def _parse_excel_cell_ref(cell_ref: Any) -> Optional[Tuple[int, int]]:
    if not cell_ref:
        return None
    m = re.match(r'^([A-Z]+)(\d+)$', str(cell_ref).upper())
    if not m:
        return None
    col_letters, row_text = m.groups()
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - 64)
    return int(row_text), col


def _normalize_slice_config(xlsx_slice: Any) -> Dict[str, Any]:
    if not xlsx_slice:
        return {"strict": True, "rules": []}
    if hasattr(xlsx_slice, "dict") and callable(getattr(xlsx_slice, "dict")):
        xlsx_slice = xlsx_slice.dict()
    elif hasattr(xlsx_slice, "model_dump") and callable(getattr(xlsx_slice, "model_dump")):
        xlsx_slice = xlsx_slice.model_dump()
    if not isinstance(xlsx_slice, dict):
        raise ValueError("[XLSX_SLICE_INVALID] xlsx_slice 必須是物件")

    strict = bool(xlsx_slice.get("strict", True))
    raw_rules = xlsx_slice.get("rules", [])
    if raw_rules is None:
        raw_rules = []
    if not isinstance(raw_rules, list):
        raise ValueError("[XLSX_SLICE_INVALID] xlsx_slice.rules 必須是陣列")

    normalized_rules: List[Dict[str, Any]] = []
    for idx, raw_rule in enumerate(raw_rules):
        rule = raw_rule
        if hasattr(rule, "dict") and callable(getattr(rule, "dict")):
            rule = rule.dict()
        elif hasattr(rule, "model_dump") and callable(getattr(rule, "model_dump")):
            rule = rule.model_dump()
        if not isinstance(rule, dict):
            raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] 必須是物件")

        sheet_name = rule.get("sheet_name")
        sheet_index = rule.get("sheet_index")
        table_id = rule.get("table_id")
        row_start = rule.get("row_start")
        row_end = rule.get("row_end")

        if sheet_name in ("", None) and sheet_index is None:
            raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] 需提供 sheet_name 或 sheet_index")
        if row_start is None or row_end is None:
            raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] 需提供 row_start 與 row_end")

        try:
            row_start_i = int(row_start)
            row_end_i = int(row_end)
        except Exception:
            raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] row_start/row_end 需為整數")
        if row_start_i < 1 or row_end_i < 1:
            raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] row_start/row_end 需 >= 1")
        if row_end_i < row_start_i:
            raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] row_end 不可小於 row_start")

        sheet_index_i: Optional[int] = None
        if sheet_index is not None:
            try:
                sheet_index_i = int(sheet_index)
            except Exception:
                raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] sheet_index 需為整數")
            if sheet_index_i < 0:
                raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] sheet_index 需 >= 0")

        table_id_i: Optional[int] = None
        if table_id is not None:
            try:
                table_id_i = int(table_id)
            except Exception:
                raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] table_id 需為整數")
            if table_id_i < 0:
                raise ValueError(f"[XLSX_SLICE_INVALID] rules[{idx}] table_id 需 >= 0")

        normalized_rules.append({
            "sheet_name": str(sheet_name) if sheet_name not in (None, "") else None,
            "sheet_index": sheet_index_i,
            "table_id": table_id_i,
            "row_start": row_start_i,
            "row_end": row_end_i,
        })

    return {"strict": strict, "rules": normalized_rules}


def apply_extract_slice(extract_result: Dict[str, Any], xlsx_slice: Any) -> Dict[str, Any]:
    normalized = _normalize_slice_config(xlsx_slice)
    rules: List[Dict[str, Any]] = normalized.get("rules", [])
    strict: bool = bool(normalized.get("strict", True))
    if not rules:
        return extract_result

    unit_paras = extract_result.get("unit_paras")
    metadata = extract_result.get("metadata")
    if not isinstance(unit_paras, list) or not isinstance(metadata, dict):
        raise ValueError("[XLSX_SLICE_INVALID_RESPONSE] xlsx parser extract 結果不完整")

    sheets_meta = metadata.get("sheets", [])
    if not isinstance(sheets_meta, list):
        sheets_meta = []
    sheets_by_index: Dict[int, Dict[str, Any]] = {}
    sheets_by_name: Dict[str, Dict[str, Any]] = {}
    for sheet in sheets_meta:
        if not isinstance(sheet, dict):
            continue
        idx = sheet.get("sheet_index")
        name = sheet.get("sheet_name")
        if isinstance(idx, int):
            sheets_by_index[idx] = sheet
        if isinstance(name, str):
            sheets_by_name[name] = sheet

    units_by_index: Dict[int, Dict[str, Any]] = {}
    units_by_name: Dict[str, Dict[str, Any]] = {}
    for unit in unit_paras:
        if not isinstance(unit, dict):
            continue
        idx = unit.get("sheet_index")
        name = unit.get("sheet_name")
        if isinstance(idx, int):
            units_by_index[idx] = unit
        if isinstance(name, str):
            units_by_name[name] = unit

    warnings: List[str] = []
    sliced_units: List[Dict[str, Any]] = []
    slice_tables: List[Dict[str, Any]] = []
    applied_rules = 0

    for ridx, rule in enumerate(rules):
        target_sheet_meta: Optional[Dict[str, Any]] = None
        sheet_name = rule.get("sheet_name")
        sheet_index = rule.get("sheet_index")
        if sheet_name is not None:
            target_sheet_meta = sheets_by_name.get(sheet_name)
        if target_sheet_meta is None and sheet_index is not None:
            target_sheet_meta = sheets_by_index.get(sheet_index)

        if target_sheet_meta is None:
            msg = f"rules[{ridx}] 找不到指定 sheet"
            if strict:
                raise ValueError(f"[XLSX_SLICE_INVALID] {msg}")
            warnings.append(msg)
            continue

        target_sheet_index = target_sheet_meta.get("sheet_index")
        target_sheet_name = target_sheet_meta.get("sheet_name")
        target_unit = units_by_index.get(target_sheet_index) if isinstance(target_sheet_index, int) else None
        if target_unit is None and isinstance(target_sheet_name, str):
            target_unit = units_by_name.get(target_sheet_name)
        if target_unit is None:
            msg = f"rules[{ridx}] 找不到 sheet 對應內容"
            if strict:
                raise ValueError(f"[XLSX_SLICE_INVALID] {msg}")
            warnings.append(msg)
            continue

        source_df = target_unit.get("sheet_non_empty_df")
        if source_df is None:
            source_df = target_unit.get("sheet_df")
        if source_df is None or not hasattr(source_df, "iloc"):
            msg = f"rules[{ridx}] sheet 無法取得資料框"
            if strict:
                raise ValueError(f"[XLSX_SLICE_INVALID] {msg}")
            warnings.append(msg)
            continue

        tables_preview = target_sheet_meta.get("tables_preview", [])
        if not isinstance(tables_preview, list):
            tables_preview = []
        table_id = rule.get("table_id")
        if table_id is None:
            target_tables = [t for t in tables_preview if isinstance(t, dict)]
        else:
            target_tables = [t for t in tables_preview if isinstance(t, dict) and t.get("table_id") == table_id]

        if not target_tables:
            msg = f"rules[{ridx}] 找不到指定 table"
            if strict:
                raise ValueError(f"[XLSX_SLICE_INVALID] {msg}")
            warnings.append(msg)
            continue

        for table in target_tables:
            top_row = table.get("top_row")
            bottom_row = table.get("bottom_row")
            left_col = table.get("left_col")
            right_col = table.get("right_col")
            table_idx = table.get("table_id")
            if not all(isinstance(v, int) for v in [top_row, bottom_row, left_col, right_col]):
                msg = f"rules[{ridx}] table 座標不完整"
                if strict:
                    raise ValueError(f"[XLSX_SLICE_INVALID] {msg}")
                warnings.append(msg)
                continue

            table_height = bottom_row - top_row + 1
            req_row_start = int(rule["row_start"])
            req_row_end = int(rule["row_end"])
            if req_row_start > table_height:
                msg = f"rules[{ridx}] table_id={table_idx} row_start 超過表格高度"
                if strict:
                    raise ValueError(f"[XLSX_SLICE_INVALID] {msg}")
                warnings.append(msg)
                continue
            applied_row_end = min(req_row_end, table_height)
            if applied_row_end < req_row_start:
                msg = f"rules[{ridx}] table_id={table_idx} row 範圍無效"
                if strict:
                    raise ValueError(f"[XLSX_SLICE_INVALID] {msg}")
                warnings.append(msg)
                continue

            data_top = top_row - 1
            data_bottom = bottom_row
            data_left = left_col - 1
            data_right = right_col
            table_df = source_df.iloc[data_top:data_bottom, data_left:data_right].copy()
            sliced_df = table_df.iloc[req_row_start - 1:applied_row_end].copy()
            markdown = _df_to_markdown(sliced_df)

            selected_top_row = top_row + (req_row_start - 1)
            selected_bottom_row = top_row + (applied_row_end - 1)

            formulas = []
            for formula in target_unit.get("formula_cells", []) or []:
                if not isinstance(formula, dict):
                    continue
                parsed = _parse_excel_cell_ref(formula.get("cell"))
                if not parsed:
                    continue
                row_num, col_num = parsed
                if selected_top_row <= row_num <= selected_bottom_row and left_col <= col_num <= right_col:
                    formulas.append(formula)

            unit_text = "\n\n".join([
                _sheet_header(target_sheet_name, is_first=(int(target_sheet_index or 0) == 0)),
                f"## Table {table_idx}",
                markdown
            ]).strip()
            sliced_units.append({
                "unit_text": unit_text,
                "indent_level": 0,
                "order": len(sliced_units),
                "unit_type": "table_slice",
                "sheet_name": target_sheet_name,
                "sheet_index": target_sheet_index,
                "table_id": table_idx,
                "top_row": selected_top_row,
                "bottom_row": selected_bottom_row,
                "left_col": left_col,
                "right_col": right_col,
                "row_start": req_row_start,
                "row_end": applied_row_end,
                "table_markdown": markdown,
                "table_values": _to_serializable_df(sliced_df),
                "formula_cells": formulas,
            })
            slice_tables.append({
                "rule_index": ridx,
                "sheet_name": target_sheet_name,
                "sheet_index": target_sheet_index,
                "table_id": table_idx,
                "row_start": req_row_start,
                "row_end": applied_row_end,
                "table_height": table_height,
            })
            applied_rules += 1

    if not sliced_units:
        raise ValueError("[XLSX_SLICE_NO_MATCH] 依規則找不到可輸出的表格內容")

    text_content = '\n'.join([str(item.get('unit_text', '')).strip() for item in sliced_units if str(item.get('unit_text', '')).strip()])
    new_metadata = dict(metadata)
    new_metadata["slice_applied"] = True
    new_metadata["slice_rules"] = rules
    new_metadata["slice_warnings"] = warnings
    new_metadata["slice_tables"] = slice_tables
    new_metadata["slice_summary"] = {
        "requested_rules": len(rules),
        "applied_rules": applied_rules,
        "matched_tables": len(slice_tables),
        "strict": strict,
        "warnings_count": len(warnings),
    }

    new_result = dict(extract_result)
    new_result["unit_paras"] = sliced_units
    new_result["text"] = text_content
    new_result["metadata"] = new_metadata
    return new_result


# ============================================================
# 1. Preview（預覽）
# ============================================================

def preview(
        file_path: str,
        separator: str = '\n\n---\n\n',
        max_rows_per_sheet: Optional[int] = None,
        max_cols_per_sheet: Optional[int] = None,
        **kwargs
    ) -> str:
    _require_dependencies()
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'檔案不存在: {file_path}')

    normalized_path = _normalize_to_xlsx(file_path)
    m_logger.info(f'[preview] 開始預覽檔案: {normalized_path}')

    xls = pd.ExcelFile(normalized_path, engine='openpyxl')
    parts: List[str] = []
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine='openpyxl')
        part = [_sheet_header(sheet_name, is_first=(len(parts) == 0))]
        part.append(_df_to_markdown(df, max_rows=max_rows_per_sheet, max_cols=max_cols_per_sheet))
        parts.append('\n\n'.join([p for p in part if p]))
    return separator.join(parts)



def parse(file_path: str, **kwargs) -> str:
    return preview(file_path=file_path, **kwargs)


# ============================================================
# 2. Extract（提取）
# ============================================================

def extract(
        file_path: Optional[str] = None,
        text: Optional[str] = None,
        include_images: bool = True,
        image_placeholder: bool = True,
        to_base64: bool = True,
        include_formulas: bool = True,
        table_detector: Optional[str] = None,
        inspector_table_params: Optional[Dict[str, Any]] = None,
        feature_scan: bool = False,
        feature_scan_params: Optional[Dict[str, Any]] = None,
        include_textboxes: bool = False,
        **kwargs
    ) -> Dict[str, Any]:
    _require_dependencies()
    if text is not None and not file_path:
        return _extract_from_text(text)
    if not file_path:
        raise ValueError('extract() 需要 file_path 或 text 其中之一')
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'檔案不存在: {file_path}')

    resolved_table_detector = _resolve_table_detector_mode(table_detector, feature_scan)
    resolved_inspector_params = inspector_table_params if inspector_table_params is not None else feature_scan_params

    original_file_path = file_path
    normalized_path = _normalize_to_xlsx(file_path)
    m_logger.info(f'[extract] 開始提取 Excel 檔案: {normalized_path}')

    wb = load_workbook(normalized_path, data_only=False)
    metadata: Dict[str, Any] = {
        'file_path': original_file_path,
        'processed_file_path': normalized_path,
        'source_format': Path(original_file_path).suffix.lower().lstrip('.'),
        'sheet_count': len(wb.worksheets),
        'sheet_names': [ws.title for ws in wb.worksheets],
        'sheets': [],
        'images': [],
        'formula_cells': [],
        'data_validations': [],
        'has_images': False,
        'has_formulas': False,
        'has_data_validations': False,
        'has_textboxes': False,
        'image_count': 0,
        'formula_count': 0,
        'data_validation_count': 0,
        'textbox_count': 0,
        'total_tables': 0,
    }
    unit_paras: List[Dict[str, Any]] = []
    global_image_id = 0
    textboxes_by_sheet = _extract_xlsx_textboxes(normalized_path) if include_textboxes else {}

    for sheet_index, ws in enumerate(wb.worksheets):
        sheet_df = _worksheet_to_dataframe(ws, include_formulas=include_formulas)
        sheet_non_empty_df, sheet_bbox = _trim_empty_df(sheet_df)
        formulas = _extract_formula_cells(ws) if include_formulas else []
        sheet_validations = _extract_data_validations(ws, wb)
        validation_by_cell = _build_validation_by_cell(sheet_validations)
        sheet_images = _extract_sheet_images(
            ws,
            include_images=include_images and image_placeholder,
            to_base64=to_base64,
            image_id_start=global_image_id
        )
        global_image_id += len(sheet_images)
        if not image_placeholder:
            for img in sheet_images:
                img.pop('placeholder', None)
                img.pop('id', None)

        tables, detector_mode = _detect_tables_for_sheet(
            ws.title,
            sheet_non_empty_df,
            table_detector=resolved_table_detector,
            inspector_table_params=resolved_inspector_params,
            min_non_empty_cells=1,
        )
        floating_images = _collect_floating_images(sheet_images, tables)

        sheet_unit = _build_sheet_unit(
            sheet_name=ws.title,
            sheet_index=sheet_index,
            sheet_df=sheet_df,
            sheet_non_empty_df=sheet_non_empty_df,
            sheet_bbox=sheet_bbox,
            formulas=formulas,
            sheet_images=sheet_images,
        )
        sheet_unit['data_validations'] = sheet_validations
        sheet_unit['validation_by_cell'] = validation_by_cell
        sheet_unit['table_detector'] = detector_mode
        sheet_textboxes = textboxes_by_sheet.get(ws.title, [])
        if sheet_textboxes:
            sheet_unit['textboxes'] = sheet_textboxes
        _append_validation_annotations_to_unit_text(sheet_unit, **kwargs)
        unit_paras.append(sheet_unit)
        _append_floating_image_units(unit_paras, ws.title, sheet_index, floating_images)

        metadata['total_tables'] += len(tables)
        metadata['formula_cells'].extend(formulas)
        metadata['data_validations'].extend(sheet_validations)
        metadata['images'].extend(sheet_images)
        metadata['sheets'].append({
            'sheet_name': ws.title,
            'sheet_index': sheet_index,
            'shape': [int(sheet_df.shape[0]), int(sheet_df.shape[1])],
            'non_empty_shape': [int(sheet_non_empty_df.shape[0]), int(sheet_non_empty_df.shape[1])] if sheet_non_empty_df is not None and not sheet_non_empty_df.empty else [0, 0],
            'sheet_bbox': sheet_bbox,
            'table_count': len(tables),
            'formula_count': len(formulas),
            'data_validation_count': len(sheet_validations),
            'image_count': len(sheet_images),
            'tables_preview': [{
                'table_id': t['table_id'],
                'top_row': t['top_row'],
                'left_col': t['left_col'],
                'bottom_row': t['bottom_row'],
                'right_col': t['right_col'],
                'shape': t['shape'],
            } for t in tables],
            'table_detector': detector_mode,
            'textboxes': sheet_textboxes,
            'textbox_count': len(sheet_textboxes),
            'data_validations': sheet_validations,
            'validation_by_cell': validation_by_cell,
        })
        sheet_unit['detected_tables'] = [{
            'table_id': t['table_id'],
            'top_row': t['top_row'],
            'left_col': t['left_col'],
            'bottom_row': t['bottom_row'],
            'right_col': t['right_col'],
        } for t in tables]

    metadata['formula_count'] = len(metadata['formula_cells'])
    metadata['data_validation_count'] = len(metadata['data_validations'])
    metadata['image_count'] = len(metadata['images'])
    metadata['has_images'] = metadata['image_count'] > 0
    metadata['has_formulas'] = metadata['formula_count'] > 0
    metadata['has_data_validations'] = metadata['data_validation_count'] > 0
    metadata['textbox_count'] = int(sum(int(s.get('textbox_count', 0) or 0) for s in metadata.get('sheets', [])))
    metadata['has_textboxes'] = metadata['textbox_count'] > 0

    m_logger.info(
        f"[extract] 提取完成: sheet數={metadata['sheet_count']}, table數={metadata['total_tables']}, "
        f"公式數={metadata['formula_count']}, 圖像數={metadata['image_count']}, "
        f"資料驗證數={metadata['data_validation_count']}"
    )
    _apply_extract_multi_prompts_to_units(unit_paras, **kwargs)
    result = {
        'unit_paras': unit_paras,
        'metadata': metadata,
    }
    xlsx_slice = kwargs.get('xlsx_slice')
    if xlsx_slice:
        result = apply_extract_slice(result, xlsx_slice)
    return result



def _extract_from_text(text: str, **kwargs) -> Dict[str, Any]:
    df = pd.DataFrame([line.split('\t') for line in text.splitlines() if line.strip()])
    unit = {
        'unit_text': _df_to_markdown(df),
        'indent_level': 0,
        'order': 0,
        'structure_chars': [{'type': 'sheet', 'sheet_name': 'TextSheet'}],
        'sheet_name': 'TextSheet',
        'sheet_index': 0,
        'sheet_df': df,
        'sheet_values': _to_serializable_df(df),
        'sheet_markdown': _df_to_markdown(df),
        'sheet_non_empty_df': df,
        'sheet_non_empty_values': _to_serializable_df(df),
        'sheet_non_empty_markdown': _df_to_markdown(df),
        'sheet_bbox': (1, 1, int(df.shape[0]), int(df.shape[1]) if len(df.shape) > 1 else 1),
        'formula_cells': [],
        'sheet_images': [],
    }
    unit_paras = [unit]
    _apply_extract_multi_prompts_to_units(unit_paras, **kwargs)
    return {
        'unit_paras': unit_paras,
        'metadata': {
            'file_path': None,
            'processed_file_path': None,
            'source_format': 'text',
            'sheet_count': 1,
            'sheet_names': ['TextSheet'],
            'sheets': [],
            'images': [],
            'formula_cells': [],
            'has_images': False,
            'has_formulas': False,
            'image_count': 0,
            'formula_count': 0,
            'total_tables': 1,
        }
    }


# ============================================================
# 3. Segment（分段）
# ============================================================

def segment(
        unit_paras: List[Dict[str, Any]],
        suitable_char_count: int = 500,
        separator: str = '\n',
        metadata: Dict[str, Any] = None,
        enable_image_llm: bool = True,
        llm_provider: str = 'remote',
        llm_model: str = 'remote8b',
        llm_base_url: str = None,
        image_context_window: int = 200,
        max_images_per_batch: int = 50,
        image_prompt_template: str = None,
        image_llm_provider: str = 'openai',
        image_llm_model: str = 'gpt4o_chat',
        min_non_empty_cells: int = 1,
        table_detector: Optional[str] = None,
        feature_scan: bool = False,
        **kwargs
    ) -> List[Dict[str, Any]]:
    if not unit_paras:
        m_logger.warning('[segment] unit_paras 為空，返回空列表')
        return []

    resolved_table_detector = _resolve_table_detector_mode(table_detector, feature_scan)

    results: List[Dict[str, Any]] = []
    order = 0
    sheet_label_emitted: Set[Tuple[int, str]] = set()
    for unit in unit_paras:
        sheet_name = unit.get('sheet_name', '')
        sheet_index = unit.get('sheet_index', 0)
        sheet_key = (int(sheet_index) if isinstance(sheet_index, int) else 0, str(sheet_name or ''))
        if unit.get('unit_type') == 'sheet_image':
            raw_unit_text = unit.get('unit_text', '')
            raw_unit_text = raw_unit_text if isinstance(raw_unit_text, str) else str(raw_unit_text)
            image_placeholder_only = re.sub(r'(?m)^#\s*Sheet:\s*.*\n?', '', raw_unit_text).strip()
            if sheet_key in sheet_label_emitted:
                seg_text = image_placeholder_only
            else:
                seg_text = separator.join([p for p in [_sheet_header(sheet_name, is_first=(sheet_key[0] == 0)), image_placeholder_only] if p])
                sheet_label_emitted.add(sheet_key)
            results.append({
                'segment_id': f'seg_{order:04d}',
                'unit_text': seg_text,
                'text': seg_text,
                'indent_level': 0,
                'order': order,
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'segment_type': 'sheet_image',
                'top_row': None,
                'left_col': None,
                'bottom_row': None,
                'right_col': None,
                'table_df': pd.DataFrame(),
                'table_values': [],
                'table_markdown': '',
                'headers': [],
                'images': unit.get('images', []),
                'sheet_images': unit.get('sheet_images', []),
                'structure_chars': unit.get('structure_chars', [{'type': 'sheet', 'sheet_name': sheet_name}]),
            })
            order += 1
            continue
        sheet_df = unit.get('sheet_non_empty_df') if unit.get('sheet_non_empty_df') is not None else unit.get('sheet_df')
        if sheet_df is None or getattr(sheet_df, 'empty', True):
            raw_unit_text = unit.get('unit_text', '')
            raw_unit_text = raw_unit_text if isinstance(raw_unit_text, str) else str(raw_unit_text)
            body_text = re.sub(r'(?m)^#\s*Sheet:\s*.*\n?', '', raw_unit_text).strip()
            if sheet_key in sheet_label_emitted:
                seg_text = body_text
            else:
                seg_text = separator.join([p for p in [_sheet_header(sheet_name, is_first=(sheet_key[0] == 0)), body_text] if p])
                sheet_label_emitted.add(sheet_key)
            results.append({
                'segment_id': f'seg_{order:04d}',
                'unit_text': seg_text,
                'text': seg_text,
                'indent_level': 0,
                'order': order,
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'segment_type': 'empty_sheet',
                'top_row': None,
                'left_col': None,
                'bottom_row': None,
                'right_col': None,
                'table_df': pd.DataFrame(),
                'table_values': [],
                'table_markdown': '',
                'headers': [],
                'images': unit.get('sheet_images', []),
                'sheet_images': unit.get('sheet_images', []),
                'structure_chars': [{'type': 'sheet', 'sheet_name': sheet_name}],
            })
            order += 1
            continue

        unit_detector = str(unit.get('table_detector') or resolved_table_detector or 'native').lower()
        detected_tables = unit.get('detected_tables') if unit_detector == 'inspector' else None
        if isinstance(detected_tables, list) and detected_tables:
            tables = _tables_from_bboxes(sheet_df, detected_tables, min_non_empty_cells=min_non_empty_cells)
        else:
            tables = _detect_tables_from_sheet_df(sheet_df, min_non_empty_cells=min_non_empty_cells)
        if not tables:
            tables = [{
                'table_id': 0,
                'top_row': 1,
                'left_col': 1,
                'bottom_row': int(sheet_df.shape[0]),
                'right_col': int(sheet_df.shape[1]),
                'shape': [int(sheet_df.shape[0]), int(sheet_df.shape[1])],
                'table_df': sheet_df.copy(),
                'table_values': _to_serializable_df(sheet_df),
                'table_markdown': _df_to_markdown(sheet_df),
                'non_empty_count': int(sheet_df.applymap(lambda x: _normalize_none_like(x) is not None).to_numpy().sum()),
            }]

        emitted_validation_cells: Set[str] = set()
        for table in tables:
            table_images, sheet_images = _segment_images_for_table(table, unit.get('sheet_images', []))
            headers = _guess_headers(table['table_df'])
            structure_chars = [
                {'type': 'sheet', 'sheet_name': sheet_name},
                {
                    'type': 'table',
                    'table_id': table['table_id'],
                    'top_row': table['top_row'],
                    'left_col': table['left_col'],
                    'bottom_row': table['bottom_row'],
                    'right_col': table['right_col'],
                }
            ]
            for img in table_images:
                if img.get('placeholder'):
                    structure_chars.append({'type': 'image', 'image_id': img.get('id'), 'anchor_cell': img.get('anchor_cell')})
            text_parts: List[str] = []
            if sheet_key not in sheet_label_emitted:
                text_parts.append(_sheet_header(sheet_name, is_first=(sheet_key[0] == 0)))
                sheet_label_emitted.add(sheet_key)
            text_parts.extend([f'## Table {table["table_id"]}', table['table_markdown']])
            validation_lines = _build_validation_annotation_lines(
                table=table,
                unit=unit,
                validation_by_cell=unit.get('validation_by_cell') or {},
                emitted_validation_cells=emitted_validation_cells,
                **kwargs,
            )
            if validation_lines:
                text_parts.append('## 下拉選單')
                text_parts.extend(validation_lines)
            if table_images:
                image_line = ' '.join([img.get('placeholder', f'[IMAGE_{img.get("local_index", 0)}]') for img in table_images])
                text_parts.append(image_line)
            seg_text = separator.join([p for p in text_parts if p])
            results.append({
                'segment_id': f'seg_{order:04d}',
                'unit_text': seg_text,
                'text': seg_text,
                'indent_level': 0,
                'order': order,
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'segment_type': 'table',
                'table_id': table['table_id'],
                'top_row': table['top_row'],
                'left_col': table['left_col'],
                'bottom_row': table['bottom_row'],
                'right_col': table['right_col'],
                'table_df': table['table_df'],
                'table_values': table['table_values'],
                'table_markdown': table['table_markdown'],
                'headers': headers,
                'images': table_images,
                'sheet_images': sheet_images,
                'formula_cells': [f for f in unit.get('formula_cells', []) if _formula_in_table(f, table)],
                'structure_chars': structure_chars,
            })
            order += 1

        sheet_textboxes = unit.get('textboxes', [])
        if isinstance(sheet_textboxes, list) and sheet_textboxes:
            text_parts: List[str] = []
            if sheet_key not in sheet_label_emitted:
                text_parts.append(_sheet_header(sheet_name, is_first=(sheet_key[0] == 0)))
                sheet_label_emitted.add(sheet_key)
            text_parts.append('## Textboxes')
            for tb in sheet_textboxes:
                a1 = str(tb.get('a1_range', '') or '').strip()
                ttext = str(tb.get('text', '') or '').strip()
                if not ttext:
                    continue
                if a1:
                    text_parts.append(f"- ({a1}) {ttext}")
                else:
                    text_parts.append(f"- {ttext}")
            seg_text = separator.join([p for p in text_parts if p])
            results.append({
                'segment_id': f'seg_{order:04d}',
                'unit_text': seg_text,
                'text': seg_text,
                'indent_level': 0,
                'order': order,
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'segment_type': 'sheet_textboxes',
                'table_id': None,
                'top_row': None,
                'left_col': None,
                'bottom_row': None,
                'right_col': None,
                'table_df': pd.DataFrame(),
                'table_values': [],
                'table_markdown': '',
                'headers': [],
                'images': [],
                'sheet_images': unit.get('sheet_images', []),
                'formula_cells': [],
                'textboxes': sheet_textboxes,
                'structure_chars': [{'type': 'sheet', 'sheet_name': sheet_name}, {'type': 'textboxes'}],
            })
            order += 1

    m_logger.info(f'[segment] 分段完成: {len(results)} 個 table segments')

    if enable_image_llm and metadata and metadata.get('has_images'):
        _analyze_images_via_batch(
            segments=results,
            metadata=metadata,
            llm_provider=image_llm_provider,
            llm_model=image_llm_model,
            llm_base_url=llm_base_url,
            enable_image_llm=enable_image_llm,
            image_context_window=image_context_window,
            max_images_per_batch=max_images_per_batch,
            image_prompt_template=image_prompt_template,
        )
        _annotate_xlsx_image_labels_in_segments(results, metadata)
    _propagate_unit_meta_to_xlsx_segments(unit_paras, results)
    return results



def _formula_in_table(formula_info: Dict[str, Any], table: Dict[str, Any]) -> bool:
    cell = formula_info.get('cell')
    if not cell:
        return False
    m = re.match(r'([A-Z]+)(\d+)$', str(cell))
    if not m:
        return False
    letters, row_txt = m.groups()
    row = int(row_txt)
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - 64)
    return table['top_row'] <= row <= table['bottom_row'] and table['left_col'] <= col <= table['right_col']



def _analyze_images_via_batch(
        segments: List[Dict[str, Any]],
        metadata: Dict[str, Any],
        llm_provider: str = 'openai',
        llm_model: str = 'gpt4o_chat',
        llm_base_url: str = None,
        enable_image_llm: bool = True,
        image_context_window: int = 200,
        max_images_per_batch: int = 50,
        image_prompt_template: str = None,
        **kwargs
    ) -> Dict[str, Any]:
    if analyze_images_via_batch_common is None:
        m_logger.warning('[_analyze_images_via_batch] shared helper unavailable')
        return {}
    return analyze_images_via_batch_common(
        segments=segments,
        metadata=metadata,
        llm_provider=llm_provider,
        llm_model=llm_model,
        llm_base_url=llm_base_url,
        enable_image_llm=enable_image_llm,
        image_context_window=image_context_window,
        max_images_per_batch=max_images_per_batch,
        image_prompt_template=image_prompt_template,
        placeholder_pattern=r'\[IMAGE_PLACEHOLDER_(\d+)\]',
        placeholder_replacements=['[IMAGE_PLACEHOLDER_{image_id}]'],
        text_keys=['unit_text', 'call_prompt', 'text'],
        config=m_config,
        logger=m_logger
    )


def _annotate_xlsx_image_labels_in_segments(
        segments: List[Dict[str, Any]],
        metadata: Dict[str, Any],
    ) -> None:
    if not segments or not isinstance(metadata, dict):
        return
    images = metadata.get('images')
    if not isinstance(images, list) or not images:
        return

    image_meta_map: Dict[str, Dict[str, Any]] = {}
    for img in images:
        if not isinstance(img, dict):
            continue
        iid = img.get('id')
        if iid is None:
            continue
        image_meta_map[str(iid)] = img

    if not image_meta_map:
        return

    token_re = re.compile(r'\[圖像說明\|(?:id|image_id)=(\d+)(?:\|(?:hash|image_hash)=[0-9a-fA-F]{8,64})?\]')

    def _replace_token(match: re.Match) -> str:
        iid = match.group(1)
        img = image_meta_map.get(iid)
        if not isinstance(img, dict):
            return f'[圖像說明|id={iid}]'
        img_type = str(img.get('type') or 'unknown').strip()
        anchor = str(img.get('anchor_cell') or '').strip() or 'floating'
        image_hash = str(img.get('image_hash') or '').strip()
        hash_suffix = f'|hash={image_hash}' if image_hash else ''
        return f'[圖像說明|{img_type}|anchor={anchor}|id={iid}{hash_suffix}]'

    for segment in segments:
        for text_key in ('unit_text', 'text', 'call_prompt'):
            text = segment.get(text_key)
            if not isinstance(text, str) or '[圖像說明|id=' not in text:
                continue
            segment[text_key] = token_re.sub(_replace_token, text)


# ============================================================
# 4. Chunk（分塊）
# ============================================================


class BaseChunker:
    def __init__(self, parse_mode: str, max_cell_chars: int, digit: int, metadata: Optional[Dict[str, Any]], skip_dedup: bool = False, skip_keyword_gen: bool = False):
        self.parse_mode = parse_mode
        self.max_cell_chars = max_cell_chars
        self.digit = digit
        self.metadata = metadata
        self.skip_dedup = skip_dedup
        self.skip_keyword_gen = skip_keyword_gen
        self._headers_cache: Dict[int, List[str]] = {}  # 快取 headers（使用 DataFrame id 作為鍵）

    def _format_cell(self, value: Any) -> str:
        return _cell_to_display_text(value, max_len=self.max_cell_chars, digit=self.digit)

    def _build(self, chunk_text: str, seg: Dict[str, Any], chunk_order: int,
               row_index: Optional[int], col_index: Optional[int], cell_ref: Optional[str]) -> Dict[str, Any]:
        return _build_chunk_from_text(
            chunk_text=chunk_text,
            seg=seg,
            parse_mode=self.parse_mode,
            chunk_order=chunk_order,
            metadata=self.metadata,
            row_index=row_index,
            col_index=col_index,
            cell_ref=cell_ref,
            skip_dedup=self.skip_dedup,
            skip_keyword_gen=self.skip_keyword_gen,
        )

    def _get_headers(self, table_df: 'pd.DataFrame', cached_headers: Optional[List[str]] = None) -> List[str]:
        """取得 headers，優先使用快取的"""
        if cached_headers is not None:
            return cached_headers
        # 使用 DataFrame 的 id 作為快取鍵
        df_id = id(table_df)
        if df_id in self._headers_cache:
            return self._headers_cache[df_id]
        headers = _guess_headers(table_df)
        self._headers_cache[df_id] = headers
        return headers

    def generate(
        self,
        seg: Dict[str, Any],
        table_df: 'pd.DataFrame',
        chunk_order: int,
        index_signals: Optional[Dict[str, Any]] = None,
        cached_headers: Optional[List[str]] = None
    ) -> Tuple[List[Dict[str, Any]], int]:
        raise NotImplementedError


class RowChunker(BaseChunker):
    def generate(
        self,
        seg: Dict[str, Any],
        table_df: 'pd.DataFrame',
        chunk_order: int,
        index_signals: Optional[Dict[str, Any]] = None,
        cached_headers: Optional[List[str]] = None
    ) -> Tuple[List[Dict[str, Any]], int]:
        signals = index_signals or {}
        index_row = bool(signals.get('index_row', False))
        index_col = bool(signals.get('index_col', False))
        headers = signals.get('row_headers') if index_row else self._get_headers(table_df, cached_headers)
        if index_row and not headers:
            headers = [
                _cell_to_display_text(v, max_len=self.max_cell_chars, digit=self.digit).strip()
                for v in table_df.iloc[0].tolist()
            ]
        headers = headers or []
        chunks: List[Dict[str, Any]] = []
        top_row = seg.get('top_row', 1)
        row_start = 1 if index_row else 0
        col_start = 1 if index_col else 0
        for ridx in range(row_start, table_df.shape[0]):
            values = table_df.iloc[ridx].tolist()
            parts = []
            for i in range(col_start, len(values)):
                header_name = headers[i] if i < len(headers) else ''
                if header_name is not None:
                    header_name = str(header_name)
                if index_row and header_name.startswith('col_'):
                    header_name = ''
                if not header_name:
                    continue
                cell_text = self._format_cell(values[i])
                if cell_text != '':
                    parts.append(f'{header_name}={cell_text}')
            if not parts:
                continue
            row_text = ' | '.join(parts)
            actual_row = top_row + ridx
            chunks.append(self._build(row_text, seg, chunk_order, actual_row, None, None))
            chunk_order += 1
        return chunks, chunk_order


class ColumnChunker(BaseChunker):
    def generate(
        self,
        seg: Dict[str, Any],
        table_df: 'pd.DataFrame',
        chunk_order: int,
        index_signals: Optional[Dict[str, Any]] = None,
        cached_headers: Optional[List[str]] = None
    ) -> Tuple[List[Dict[str, Any]], int]:
        signals = index_signals or {}
        index_row = bool(signals.get('index_row', False))
        index_col = bool(signals.get('index_col', False))
        headers = signals.get('row_headers') if index_row else self._get_headers(table_df, cached_headers)
        if index_row and not headers:
            headers = [
                _cell_to_display_text(v, max_len=self.max_cell_chars, digit=self.digit).strip()
                for v in table_df.iloc[0].tolist()
            ]
        headers = headers or []
        chunks: List[Dict[str, Any]] = []
        left_col = seg.get('left_col', 1)

        if index_col:
            row_start = 1 if index_row else 0
            row_headers = (signals.get('col_headers') or [])[row_start:]
            if not row_headers:
                col_values = table_df.iloc[row_start:, 0].tolist()
                row_headers = [
                    _cell_to_display_text(v, max_len=self.max_cell_chars, digit=self.digit).strip()
                    for v in col_values
                ]
            for cidx in range(1, table_df.shape[1]):
                values = table_df.iloc[row_start:, cidx].tolist()
                parts = []
                for r_i, v in enumerate(values):
                    header_name = row_headers[r_i] if r_i < len(row_headers) else ''
                    if header_name is not None:
                        header_name = str(header_name)
                    if not header_name:
                        continue
                    cell_text = self._format_cell(v)
                    if cell_text != '':
                        parts.append(f'{header_name}={cell_text}')
                if not parts:
                    continue
                col_name = headers[cidx] if cidx < len(headers) and headers[cidx] else f'col_{cidx + 1}'
                col_text = f'{col_name}: ' + ' | '.join(parts)
                actual_col = left_col + cidx
                chunks.append(self._build(col_text, seg, chunk_order, None, actual_col, None))
                chunk_order += 1
            return chunks, chunk_order

        for cidx in range(table_df.shape[1]):
            values = table_df.iloc[:, cidx].tolist()
            non_empty_values = []
            for v in values:
                cell_text = self._format_cell(v)
                if cell_text != '':
                    non_empty_values.append(cell_text)
            if not non_empty_values:
                continue
            col_name = headers[cidx] if cidx < len(headers) else f'col_{cidx + 1}'
            col_text = f'{col_name}: ' + ' | '.join(non_empty_values)
            actual_col = left_col + cidx
            chunks.append(self._build(col_text, seg, chunk_order, None, actual_col, None))
            chunk_order += 1
        return chunks, chunk_order


class CellChunker(BaseChunker):
    def generate(
        self,
        seg: Dict[str, Any],
        table_df: 'pd.DataFrame',
        chunk_order: int,
        index_signals: Optional[Dict[str, Any]] = None,
        cached_headers: Optional[List[str]] = None
    ) -> Tuple[List[Dict[str, Any]], int]:
        headers = self._get_headers(table_df, cached_headers)
        chunks: List[Dict[str, Any]] = []
        top_row = seg.get('top_row', 1)
        left_col = seg.get('left_col', 1)
        for ridx in range(table_df.shape[0]):
            for cidx in range(table_df.shape[1]):
                val = table_df.iat[ridx, cidx]
                text = self._format_cell(val)
                if text == '':
                    continue
                actual_row = top_row + ridx
                actual_col = left_col + cidx
                cell_ref = f'{_colnum_to_letter(actual_col)}{actual_row}'
                header = headers[cidx] if cidx < len(headers) else f'col_{actual_col}'
                cell_text = f'{cell_ref} | {header}={text}'
                chunks.append(self._build(cell_text, seg, chunk_order, actual_row, actual_col, cell_ref))
                chunk_order += 1
        return chunks, chunk_order


class PreviewChunker(BaseChunker):
    def generate(
        self,
        seg: Dict[str, Any],
        table_df: 'pd.DataFrame',
        chunk_order: int,
        index_signals: Optional[Dict[str, Any]] = None,
        cached_headers: Optional[List[str]] = None
    ) -> Tuple[List[Dict[str, Any]], int]:
        if table_df is None or getattr(table_df, 'empty', True):
            return [], chunk_order
        try:
            view_df = table_df.copy()
            view_df = view_df.applymap(lambda v: self._format_cell(v))
            table_md = _df_to_markdown(view_df)
        except Exception:
            table_md = _df_to_markdown(table_df)
        chunks = [self._build(table_md, seg, chunk_order, None, None, None)]
        return chunks, chunk_order + 1


class DiagonalChunker(BaseChunker):
    def generate(
        self,
        seg: Dict[str, Any],
        table_df: 'pd.DataFrame',
        chunk_order: int,
        index_signals: Optional[Dict[str, Any]] = None,
        cached_headers: Optional[List[str]] = None
    ) -> Tuple[List[Dict[str, Any]], int]:
        if table_df is None or getattr(table_df, 'empty', True):
            return [], chunk_order
        chunks: List[Dict[str, Any]] = []
        rows, cols = table_df.shape
        top_row = seg.get('top_row', 1)
        left_col = seg.get('left_col', 1)
        for diag in range(rows + cols - 1):
            parts = []
            for ridx in range(rows):
                cidx = diag - ridx
                if cidx < 0 or cidx >= cols:
                    continue
                val = table_df.iat[ridx, cidx]
                text = self._format_cell(val)
                if text == '':
                    continue
                actual_row = top_row + ridx
                actual_col = left_col + cidx
                cell_ref = f'{_colnum_to_letter(actual_col)}{actual_row}'
                parts.append(f'{cell_ref}={text}')
            if not parts:
                continue
            diag_text = f'diag_{diag + 1}: ' + ' | '.join(parts)
            chunks.append(self._build(diag_text, seg, chunk_order, None, None, None))
            chunk_order += 1
        return chunks, chunk_order


def _parse_mode_and_reason_from_llm_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    if not text:
        return None, None
    stripped = text.strip()
    # 嘗試解析 JSON 回應
    try:
        data = json.loads(stripped)
        if isinstance(data, dict):
            mode = (data.get('mode') or data.get('parse_mode') or '').strip().lower()
            reason = data.get('reason')
            if mode in {'row', 'column', 'preview', 'diagonal', 'cell'}:
                return mode, reason
    except Exception:
        pass
    lowered = stripped.lower()
    mode = None
    for cand in ('row', 'column', 'preview', 'diagonal', 'cell'):
        if cand in lowered:
            mode = cand
            break
    reason = None
    # 嘗試從「理由:」或「原因:」之後擷取
    for key in ('理由', '原因', 'reason'):
        if key in lowered:
            idx = lowered.find(key)
            reason = stripped[idx + len(key):].lstrip('：: \n')
            break
    return mode, reason


def _clean_llm_json_object(text: str) -> str:
    """
        清理 LLM 輸出，擷取第一個完整 JSON 物件字串。
        舉例: '{"a":1}<|eot_id|>...' -> '{"a":1}'
    """
    if not text or not isinstance(text, str):
        return ''
    cleaned = text.strip()
    special_tokens = ['<|eot_id|>', '<|end_of_text|>', '<|endoftext|>', '<|im_end|>', '<|end|>']
    for token in special_tokens:
        cleaned = cleaned.replace(token, '')
    cleaned = cleaned.strip()

    # Prefer JSON inside code block if present
    code_block_match = re.search(r'```(?:json)?\s*(\{[\s\S]*?\})\s*```', cleaned, re.DOTALL)
    if code_block_match:
        return code_block_match.group(1).strip()

    start = cleaned.find('{')
    if start == -1:
        return ''

    brace_count = 0
    end = -1
    for i in range(start, len(cleaned)):
        ch = cleaned[i]
        if ch == '{':
            brace_count += 1
        elif ch == '}':
            brace_count -= 1
            if brace_count == 0:
                end = i
                break
    if end != -1:
        return cleaned[start:end + 1].strip()

    # Fallback: try to use last closing brace
    end = cleaned.rfind('}')
    if end != -1 and end > start:
        return cleaned[start:end + 1].strip()
    return ''


def _normalize_header_list(headers: Any, max_items: int = 12) -> List[Any]:
    if not isinstance(headers, list):
        return []
    cleaned = [h for h in headers if h not in (None, '')]
    if max_items and len(cleaned) > max_items:
        cleaned = cleaned[:max_items]
    return cleaned


def _is_sequential_numeric_list(headers: List[Any]) -> bool:
    nums: List[int] = []
    for v in headers:
        if isinstance(v, bool):
            return False
        if isinstance(v, int):
            nums.append(int(v))
            continue
        if isinstance(v, str) and re.fullmatch(r'-?\d+', v.strip()):
            nums.append(int(v.strip()))
            continue
        return False
    if len(nums) < 3:
        return False
    nums_sorted = sorted(nums)
    if len(nums_sorted) != len(set(nums_sorted)):
        return False
    return nums_sorted[-1] - nums_sorted[0] == len(nums_sorted) - 1


def _extract_bool_field(text: str, key: str) -> Optional[bool]:
    m = re.search(rf'"{re.escape(key)}"\s*:\s*(true|false)', text, re.IGNORECASE)
    if not m:
        return None
    return m.group(1).lower() == 'true'


def _extract_array_field(text: str, key: str) -> Optional[List[Any]]:
    key_pos = text.find(f'"{key}"')
    if key_pos == -1:
        return None
    start = text.find('[', key_pos)
    if start == -1:
        return None
    in_str = False
    escape = False
    depth = 0
    end = -1
    for i in range(start, len(text)):
        ch = text[i]
        if in_str:
            if escape:
                escape = False
            elif ch == '\\':
                escape = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
            continue
        if ch == '[':
            depth += 1
            continue
        if ch == ']':
            depth -= 1
            if depth == 0:
                end = i
                break
    if end == -1:
        return None
    arr_text = text[start:end + 1]
    for marker in ('"reason":', '"mode":', '"index_row":', '"index_col":'):
        if marker in arr_text:
            prefix = arr_text.split(marker)[0]
            prefix = re.sub(r',\s*$', '', prefix)
            arr_text = prefix + ']'
            break
    try:
        val = json.loads(arr_text)
        return val if isinstance(val, list) else None
    except Exception:
        items = []
        inner = arr_text.strip().lstrip('[').rstrip(']')
        for part in inner.split(','):
            part = part.strip()
            if not part:
                continue
            if part.startswith('"') and part.endswith('"'):
                items.append(part[1:-1])
            elif re.fullmatch(r'-?\d+', part):
                items.append(int(part))
            else:
                items.append(part)
        return items


def _parse_index_relaxed(text: str) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    index_row = _extract_bool_field(text, 'index_row')
    index_col = _extract_bool_field(text, 'index_col')
    row_headers = _extract_array_field(text, 'row_headers') or []
    col_headers = _extract_array_field(text, 'col_headers') or []
    reason_match = re.search(r'"reason"\s*:\s*"(.*?)"', text, re.DOTALL)
    reason = reason_match.group(1) if reason_match else None
    if index_row is None and index_col is None and not row_headers and not col_headers:
        return None
    index_row = bool(index_row)
    index_col = bool(index_col)
    row_headers = _normalize_header_list(row_headers)
    col_headers = _normalize_header_list(col_headers)
    if _is_sequential_numeric_list(row_headers):
        row_headers = []
        index_row = False
    if _is_sequential_numeric_list(col_headers):
        col_headers = []
        index_col = False
    if index_row and not row_headers:
        index_row = False
    if index_col and not col_headers:
        index_col = False
    return {
        'index_row': index_row,
        'index_col': index_col,
        'row_headers': row_headers,
        'col_headers': col_headers,
        'reason': reason,
    }


def _build_llm_table_prompt(
    table_md: str,
    header_signals: Optional[Dict[str, Any]] = None,
    numeric_ratio: Optional[float] = None
) -> str:
    signals_text = ''
    if header_signals:
        row_headers = header_signals.get('row_headers') or []
        col_headers = header_signals.get('col_headers') or []
        if len(row_headers) > 8:
            row_headers = row_headers[:8] + ['...']
        if len(col_headers) > 8:
            col_headers = col_headers[:8] + ['...']
        signals_text = (
            "自動偵測：\n"
            f"- index_row={header_signals.get('index_row')}, row_header_score={header_signals.get('row_header_score')}\n"
            f"- index_col={header_signals.get('index_col')}, col_header_score={header_signals.get('col_header_score')}\n"
            f"- row_headers={row_headers}\n"
            f"- col_headers={col_headers}\n"
        )
    if numeric_ratio is not None:
        signals_text += f"- numeric_ratio={numeric_ratio:.3f} (numeric_cells / non_empty_cells)\n"
    return (
        "你是資料整理的助手，請根據下列表格內容選擇最適合的記錄模式。\n"
        "可選模式：row、column、preview、diagonal。\n"
        "判斷依據：\n"
        "- row：在表格前幾 row 有明顯的標題式 header，對應下來每一 col 與 header 的值域屬性相關。\n"
        "- column：在表格前幾 column 有明顯的標題式 header，對應下來每一 column 與 header 的值域屬性相關。\n"
        "- diagonal：前幾 row、前幾 column 找的到完整的標題式名稱，對應每一個儲存格都是相關值域；或者沒有 row/column 標題式名稱。\n"
        "- preview：前幾種都不適合；若 numeric_ratio > 0.5，優先選 preview。\n"
        "請用 JSON 回覆：{\"mode\":\"<row|column|preview|diagonal>\",\"reason\":\"...\"}。\n\n"
        f"{signals_text}\n"
        f"表格內容：\n{table_md}\n"
    )


def _build_llm_index_prompt(table_md: str) -> str:
    return (
        "你是資料整理的助手，請判斷下列表格是否有明確的 index_row 與 index_col。\n"
        "定義：\n"
        "- index_row：前幾 row 有明顯的標題式 header，對應到每一欄的值域屬性。\n"
        "- index_col：前幾 column 有明顯的標題式 header，對應到每一列的值域屬性。\n"
        "注意：\n"
        "- 若第一列/欄只是連續數字（例如 0,1,2... 或 1,2,3...）、空白或純序號，視為 DataFrame 自動 index/欄位序號，不要當成 header。\n"
        "- row_headers / col_headers 應回傳有語意的標籤，不要回傳一長串連續整數。\n"
        "- row_headers / col_headers 最多列出 12 個即可。\n"
        "請用 JSON 回覆："
        "{\"index_row\":true/false,\"index_col\":true/false,"
        "\"row_headers\":[...],\"col_headers\":[...],\"reason\":\"...\"}\n\n"
        f"表格內容：\n{table_md}\n"
    )


def _build_llm_combined_prompt(
        table_md: str,
        header_signals: Optional[Dict[str, Any]] = None,
        numeric_ratio: Optional[float] = None
    ) -> str:
    """合併的 prompt：同時要求 index 和 mode 偵測"""
    signals_text = ''
    if header_signals:
        row_headers = header_signals.get('row_headers') or []
        col_headers = header_signals.get('col_headers') or []
        if len(row_headers) > 8:
            row_headers = row_headers[:8] + ['...']
        if len(col_headers) > 8:
            col_headers = col_headers[:8] + ['...']
        signals_text = (
            "自動偵測：\n"
            f"- index_row={header_signals.get('index_row')}, row_header_score={header_signals.get('row_header_score')}\n"
            f"- index_col={header_signals.get('index_col')}, col_header_score={header_signals.get('col_header_score')}\n"
            f"- row_headers={row_headers}\n"
            f"- col_headers={col_headers}\n"
        )
    if numeric_ratio is not None:
        signals_text += f"- numeric_ratio={numeric_ratio:.3f} (numeric_cells / non_empty_cells)\n"
        signals_text += "- if numeric_ratio > 0.5, prefer preview\n"
    return (
        "你是資料整理的助手，請同時判斷下列表格的 index 與最適合的記錄模式。\n\n"
        "【Index 判斷】\n"
        "- index_row：前幾 row 有明顯的標題式 header，對應到每一欄的值域屬性。\n"
        "- index_col：前幾 column 有明顯的標題式 header，對應到每一列的值域屬性。\n\n"
        "注意：\n"
        "- 若第一列/欄只是連續數字（例如 0,1,2... 或 1,2,3...）、空白或純序號，視為 DataFrame 自動 index/欄位序號，不要當成 header。\n"
        "- row_headers / col_headers 應回傳有語意的標籤，不要回傳一長串連續整數。\n\n"
        "- row_headers / col_headers 最多列出 12 個即可。\n\n"
        "【Mode 判斷】\n"
        "可選模式：row、column、cell、preview、diagonal。\n"
        "判斷依據：\n"
        "- row：在表格前幾 row 有明顯的標題式 header，且每一欄與 header 的值域屬性對應。\n"
        "- column：在表格前幾 column 有明顯的標題式 header，且每一欄與 header 的值域屬性對應。\n"
        "- cell：需要逐個儲存格獨立記錄，或表格結構不規則無法用 row/column 模式處理。\n"
        "- diagonal：表格資料按對角線方向組織（從左上到右下），同一對角線上的儲存格屬於同一類別或相關聯。\n"
        "- preview：以上皆不適合，直接將整個表格轉為 Markdown 格式。\n\n"
        f"{signals_text}\n"
        "請用 JSON 回覆："
        "{\"index_row\":true/false,\"index_col\":true/false,"
        "\"row_headers\":[...],\"col_headers\":[...],"
        "\"mode\":\"<row|column|cell|preview|diagonal>\",\"reason\":\"...\"}\n\n"
        f"表格內容：\n{table_md}\n"
    )


def _parse_index_from_llm_text(text: str) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    stripped = _clean_llm_json_object(text)
    if not stripped:
        return None
    try:
        data = json.loads(stripped)
        if not isinstance(data, dict):
            return None
        index_row = bool(data.get('index_row', False))
        index_col = bool(data.get('index_col', False))
        row_headers = data.get('row_headers') or []
        col_headers = data.get('col_headers') or []
        if not isinstance(row_headers, list):
            row_headers = []
        if not isinstance(col_headers, list):
            col_headers = []
        row_headers = _normalize_header_list(row_headers)
        col_headers = _normalize_header_list(col_headers)
        if _is_sequential_numeric_list(row_headers):
            row_headers = []
            index_row = False
        if _is_sequential_numeric_list(col_headers):
            col_headers = []
            index_col = False
        if index_row and not row_headers:
            index_row = False
        if index_col and not col_headers:
            index_col = False
        reason = data.get('reason')
        return {
            'index_row': index_row,
            'index_col': index_col,
            'row_headers': row_headers,
            'col_headers': col_headers,
            'reason': reason,
        }
    except Exception:
        return _parse_index_relaxed(stripped) or _parse_index_relaxed(text)


def _parse_combined_from_llm_text(text: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """解析合併的 LLM 回應，返回 (index_dict, mode)"""
    if not text:
        return None, None
    stripped = _clean_llm_json_object(text)
    if not stripped:
        return None, None
    try:
        data = json.loads(stripped)
        if not isinstance(data, dict):
            return None, None
        
        # 解析 index
        index_row = bool(data.get('index_row', False))
        index_col = bool(data.get('index_col', False))
        row_headers = data.get('row_headers') or []
        col_headers = data.get('col_headers') or []
        if not isinstance(row_headers, list):
            row_headers = []
        if not isinstance(col_headers, list):
            col_headers = []
        row_headers = _normalize_header_list(row_headers)
        col_headers = _normalize_header_list(col_headers)
        if _is_sequential_numeric_list(row_headers):
            row_headers = []
            index_row = False
        if _is_sequential_numeric_list(col_headers):
            col_headers = []
            index_col = False
        if index_row and not row_headers:
            index_row = False
        if index_col and not col_headers:
            index_col = False
        reason = data.get('reason')
        index_dict = {
            'index_row': index_row,
            'index_col': index_col,
            'row_headers': row_headers,
            'col_headers': col_headers,
            'reason': reason,
        }
        
        # 解析 mode
        mode = (data.get('mode') or '').strip().lower()
        if mode not in {'row', 'column', 'preview', 'diagonal', 'cell'}:
            mode = None
        
        return index_dict, mode
    except Exception:
        index_dict = _parse_index_relaxed(stripped) or _parse_index_relaxed(text)
        mode = None
        lowered = stripped.lower() if stripped else (text.lower() if isinstance(text, str) else '')
        for cand in ('row', 'column', 'preview', 'diagonal', 'cell'):
            if cand in lowered:
                mode = cand
                break
        return index_dict, mode


def _is_numeric_text(text: str) -> bool:
    if not text:
        return False
    candidate = text.strip().replace(',', '')
    if candidate == '':
        return False
    try:
        float(candidate)
        return True
    except Exception:
        return False


def _is_text_like(value: Any, text: str) -> bool:
    if not text:
        return False
    if isinstance(value, str):
        return any(ch.isalpha() or ('\u4e00' <= ch <= '\u9fff') for ch in text)
    return not _is_numeric_text(text)


def _compute_numeric_cell_ratio(table_df: 'pd.DataFrame', max_cell_chars: int, digit: int) -> float:
    if table_df is None or getattr(table_df, 'empty', True):
        return 0.0
    total = 0
    numeric = 0
    for row in table_df.values.tolist():
        for value in row:
            cell_text = _cell_to_display_text(value, max_len=max_cell_chars, digit=digit).strip()
            if cell_text == '':
                continue
            total += 1
            if _is_numeric_text(cell_text):
                numeric += 1
    if total == 0:
        return 0.0
    return numeric / total


def _analyze_header_signals(
        table_df: 'pd.DataFrame',
        max_cell_chars: int,
        digit: int,
        min_non_empty_cells: int = 1,
        header_threshold: float = 0.3,
    ) -> Dict[str, Any]:
    if table_df is None or getattr(table_df, 'empty', True):
        return {
            'index_row': False,
            'index_col': False,
            'row_header_score': 0.0,
            'col_header_score': 0.0,
            'row_headers': [],
            'col_headers': [],
        }

    rows, cols = table_df.shape
    if rows == 0 or cols == 0:
        return {
            'index_row': False,
            'index_col': False,
            'row_header_score': 0.0,
            'col_header_score': 0.0,
            'row_headers': [],
            'col_headers': [],
        }

    # row header signals (first row)
    row_values = table_df.iloc[0].tolist()
    row_texts = [_cell_to_display_text(v, max_len=max_cell_chars, digit=digit).strip() for v in row_values]
    row_non_empty = [t for t in row_texts if t]
    row_non_empty_ratio = len(row_non_empty) / max(1, cols)
    row_text_like = [
        _is_text_like(row_values[i], row_texts[i]) for i in range(len(row_texts)) if row_texts[i]
    ]
    row_text_ratio = (sum(1 for x in row_text_like if x) / max(1, len(row_non_empty))) if row_non_empty else 0.0
    row_unique_ratio = len(set(t.lower() for t in row_non_empty)) / max(1, len(row_non_empty)) if row_non_empty else 0.0

    row_fill_ratio = 0.0
    if rows > 1 and row_non_empty:
        col_fill = []
        for cidx, header_text in enumerate(row_texts):
            if not header_text:
                continue
            col_values = table_df.iloc[1:, cidx].tolist()
            col_texts = [_cell_to_display_text(v, max_len=max_cell_chars, digit=digit).strip() for v in col_values]
            non_empty = sum(1 for t in col_texts if t)
            col_fill.append(non_empty / max(1, len(col_texts)))
        if col_fill:
            row_fill_ratio = sum(col_fill) / len(col_fill)

    row_header_score = round((row_non_empty_ratio + row_text_ratio + row_unique_ratio + row_fill_ratio) / 4, 3)
    index_row = row_header_score >= header_threshold

    # column header signals (first column)
    col_values = table_df.iloc[:, 0].tolist()
    col_texts = [_cell_to_display_text(v, max_len=max_cell_chars, digit=digit).strip() for v in col_values]
    col_non_empty = [t for t in col_texts if t]
    col_non_empty_ratio = len(col_non_empty) / max(1, rows)
    col_text_like = [
        _is_text_like(col_values[i], col_texts[i]) for i in range(len(col_texts)) if col_texts[i]
    ]
    col_text_ratio = (sum(1 for x in col_text_like if x) / max(1, len(col_non_empty))) if col_non_empty else 0.0
    col_unique_ratio = len(set(t.lower() for t in col_non_empty)) / max(1, len(col_non_empty)) if col_non_empty else 0.0

    col_fill_ratio = 0.0
    if cols > 1 and col_non_empty:
        row_fill = []
        for ridx, header_text in enumerate(col_texts):
            if not header_text:
                continue
            row_vals = table_df.iloc[ridx, 1:].tolist()
            row_texts_rest = [_cell_to_display_text(v, max_len=max_cell_chars, digit=digit).strip() for v in row_vals]
            non_empty = sum(1 for t in row_texts_rest if t)
            row_fill.append(non_empty / max(1, len(row_texts_rest)))
        if row_fill:
            col_fill_ratio = sum(row_fill) / len(row_fill)

    col_header_score = round((col_non_empty_ratio + col_text_ratio + col_unique_ratio + col_fill_ratio) / 4, 3)
    index_col = col_header_score >= header_threshold

    return {
        'index_row': index_row,
        'index_col': index_col,
        'row_header_score': row_header_score,
        'col_header_score': col_header_score,
        'row_headers': row_non_empty,
        'col_headers': col_non_empty,
    }


def _precompute_table_markdowns(
        segments: List[Dict[str, Any]],
        max_cell_chars: int,
        digit: int
    ) -> Dict[int, str]:
    """預先轉換所有 segments 的 DataFrame 為 Markdown 並快取"""
    markdown_map: Dict[int, str] = {}
    for idx, seg in enumerate(segments):
        table_df = seg.get('table_df')
        if table_df is None or getattr(table_df, 'empty', True):
            continue
        try:
            view_df = table_df.copy()
            view_df = view_df.applymap(lambda v: _cell_to_display_text(v, max_len=max_cell_chars, digit=digit))
            table_md = _df_to_markdown(view_df)
        except Exception:
            table_md = _df_to_markdown(table_df)
        markdown_map[idx] = table_md
    return markdown_map


def _select_table_index_by_llm(
        segments: List[Dict[str, Any]],
        max_cell_chars: int,
        digit: int,
        llm_provider: str,
        llm_model: str,
        llm_base_url: Optional[str],
        markdown_map: Optional[Dict[int, str]] = None,
        **kwargs
    ) -> Dict[int, Dict[str, Any]]:
    if not llm_base_url:
        llm_base_url = m_config.get('llm', {}).get('base_url', 'http://10.1.3.127:7017') if m_config else 'http://10.1.3.127:7017'

    # 如果沒有提供快取的 markdown_map，則預先計算
    if markdown_map is None:
        markdown_map = _precompute_table_markdowns(segments, max_cell_chars, digit)

    prompts: List[str] = []
    seg_indices: List[int] = []
    for idx, table_md in markdown_map.items():
        prompts.append(_build_llm_index_prompt(table_md))
        seg_indices.append(idx)

    if not prompts:
        return {}

    try:
        import requests
    except Exception:
        m_logger.warning("[chunk][index] 無法載入 requests，回退為空 index")
        return {}

    payload = {
        'prompts': prompts,
        'provider': llm_provider,
        'model': llm_model,
        'max_tokens': kwargs.get('llm_index_max_tokens', 32),
        'temperature': kwargs.get('llm_index_temperature', 0.0),
        'parallel': True,
        'max_batch_size': 190,
    }
    try:
        batch_chat_url = f"{llm_base_url.rstrip('/')}/chat/batch"
        resp = requests.post(batch_chat_url, json=payload, timeout=kwargs.get('llm_index_timeout', 120))
        resp.raise_for_status()
        data = resp.json()
        results = data.get('results', [])
    except Exception as e:
        m_logger.warning(f"[chunk][index] index 判斷失敗，回退為空: {e}")
        return {}

    index_map: Dict[int, Dict[str, Any]] = {}
    error_count = 0
    parse_fail_count = 0
    success_count = 0
    
    for i, item in enumerate(results):
        idx = seg_indices[i] if i < len(seg_indices) else None
        if idx is None:
            continue
        if item.get('error'):
            error_count += 1
            m_logger.warning(f"[chunk][index] seg_index={idx} LLM 錯誤: {item.get('error')}")
            continue
        result_obj = item.get('result', {}) if isinstance(item, dict) else {}
        output = ''
        if isinstance(result_obj, dict):
            output = result_obj.get('output', '') or ''
        parsed = _parse_index_from_llm_text(output)
        if parsed is None:
            parse_fail_count += 1
            m_logger.warning(f"[chunk][index] seg_index={idx} 解析失敗，LLM 輸出: {output[:1000]}+...")
            continue
        index_map[idx] = parsed
        success_count += 1
        m_logger.info(
            f"[chunk][index] seg_index={idx} index_row={parsed.get('index_row')} "
            f"index_col={parsed.get('index_col')} reason={parsed.get('reason')}"
        )
    
    if not index_map:
        m_logger.warning(
            f"[chunk][index] 所有 {len(seg_indices)} 個 segments 的 index 偵測失敗 "
            f"(成功={success_count}, 錯誤={error_count}, 解析失敗={parse_fail_count})"
        )
    else:
        m_logger.info(
            f"[chunk][index] index 偵測完成: 成功={success_count}/{len(seg_indices)}, "
            f"錯誤={error_count}, 解析失敗={parse_fail_count}"
        )
    
    return index_map


def _select_table_index_and_mode_by_llm(
        segments: List[Dict[str, Any]],
        max_cell_chars: int,
        digit: int,
        llm_provider: str,
        llm_model: str,
        llm_base_url: Optional[str],
        markdown_map: Optional[Dict[int, str]] = None,
        **kwargs
    ) -> Tuple[Dict[int, Dict[str, Any]], Dict[int, str]]:
    """合併的 LLM 調用：同時返回 index 和 mode 偵測結果"""
    if not llm_base_url:
        llm_base_url = m_config.get('llm', {}).get('base_url', 'http://10.1.3.127:7017') if m_config else 'http://10.1.3.127:7017'

    # 如果沒有提供快取的 markdown_map，則預先計算
    if markdown_map is None:
        markdown_map = _precompute_table_markdowns(segments, max_cell_chars, digit)

    prompts: List[str] = []
    seg_indices: List[int] = []
    for idx, table_md in markdown_map.items():
        table_df = segments[idx].get('table_df') if idx < len(segments) else None
        numeric_ratio = _compute_numeric_cell_ratio(table_df, max_cell_chars, digit) if table_df is not None else 0.0
        m_logger.info(f"[chunk][combined] seg_index={idx} numeric_ratio={numeric_ratio:.3f}")
        prompts.append(_build_llm_combined_prompt(table_md, numeric_ratio=numeric_ratio))
        seg_indices.append(idx)

    if not prompts:
        return {}, {}

    try:
        import requests
    except Exception:
        m_logger.warning("[chunk][combined] 無法載入 requests，回退為空結果")
        return {}, {}

    payload = {
        'prompts': prompts,
        'provider': llm_provider,
        'model': llm_model,
        'max_tokens': kwargs.get('llm_combined_max_tokens', 64),
        'temperature': kwargs.get('llm_combined_temperature', 0.0),
        'parallel': True,
        'max_batch_size': 190,
    }
    try:
        batch_chat_url = f"{llm_base_url.rstrip('/')}/chat/batch"
        resp = requests.post(batch_chat_url, json=payload, timeout=kwargs.get('llm_combined_timeout', 120))
        resp.raise_for_status()
        data = resp.json()
        results = data.get('results', [])
    except Exception as e:
        m_logger.warning(f"[chunk][combined] 合併偵測失敗，回退為空: {e}")
        return {}, {}

    index_map: Dict[int, Dict[str, Any]] = {}
    mode_map: Dict[int, str] = {}
    error_count = 0
    parse_fail_count = 0
    success_count = 0

    for i, item in enumerate(results):
        idx = seg_indices[i] if i < len(seg_indices) else None
        if idx is None:
            continue
        if item.get('error'):
            error_count += 1
            m_logger.warning(f"[chunk][combined] seg_index={idx} LLM 錯誤: {item.get('error')}")
            continue
        result_obj = item.get('result', {}) if isinstance(item, dict) else {}
        output = ''
        if isinstance(result_obj, dict):
            output = result_obj.get('output', '') or ''
        index_dict, mode = _parse_combined_from_llm_text(output)
        if index_dict is None and mode is None:
            parse_fail_count += 1
            m_logger.warning(f"[chunk][combined] seg_index={idx} 解析失敗，LLM 輸出: {output[:200]}")
            continue
        if index_dict:
            index_map[idx] = index_dict
        if mode:
            mode_map[idx] = mode
        else:
            mode_map[idx] = 'row'  # 預設回退
        success_count += 1
        m_logger.info(
            f"[chunk][combined] seg_index={idx} index_row={index_dict.get('index_row') if index_dict else None} "
            f"index_col={index_dict.get('index_col') if index_dict else None} mode={mode_map.get(idx)}"
        )

    if not index_map and not mode_map:
        m_logger.warning(
            f"[chunk][combined] 所有 {len(seg_indices)} 個 segments 的合併偵測失敗 "
            f"(成功={success_count}, 錯誤={error_count}, 解析失敗={parse_fail_count})"
        )
    else:
        m_logger.info(
            f"[chunk][combined] 合併偵測完成: 成功={success_count}/{len(seg_indices)}, "
            f"錯誤={error_count}, 解析失敗={parse_fail_count}"
        )

    return index_map, mode_map


def _select_table_modes_by_llm(
        segments: List[Dict[str, Any]],
        max_cell_chars: int,
        digit: int,
        llm_provider: str,
        llm_model: str,
        llm_base_url: Optional[str],
        index_signals_map: Optional[Dict[int, Dict[str, Any]]] = None,
        markdown_map: Optional[Dict[int, str]] = None,
        **kwargs
    ) -> Dict[int, str]:
    if not llm_base_url:
        llm_base_url = m_config.get('llm', {}).get('base_url', 'http://10.1.3.127:7017') if m_config else 'http://10.1.3.127:7017'

    # 如果沒有提供快取的 markdown_map，則預先計算
    if markdown_map is None:
        markdown_map = _precompute_table_markdowns(segments, max_cell_chars, digit)

    prompts: List[str] = []
    seg_indices: List[int] = []
    for idx, table_md in markdown_map.items():
        signals = index_signals_map.get(idx) if index_signals_map else None
        table_df = segments[idx].get('table_df') if idx < len(segments) else None
        numeric_ratio = _compute_numeric_cell_ratio(table_df, max_cell_chars, digit) if table_df is not None else 0.0
        m_logger.info(f"[chunk][llm] seg_index={idx} numeric_ratio={numeric_ratio:.3f}")
        prompts.append(_build_llm_table_prompt(table_md, signals, numeric_ratio))
        seg_indices.append(idx)

    if not prompts:
        return {}

    try:
        import requests
    except Exception:
        m_logger.warning("[chunk][llm] 無法載入 requests，回退為 row 模式")
        return {idx: 'row' for idx in seg_indices}

    payload = {
        'prompts': prompts,
        'provider': llm_provider,
        'model': llm_model,
        'max_tokens': kwargs.get('llm_mode_max_tokens', 8),
        'temperature': kwargs.get('llm_mode_temperature', 0.0),
        'parallel': True,
        'max_batch_size': 190,
    }
    try:
        batch_chat_url = f"{llm_base_url.rstrip('/')}/chat/batch"
        resp = requests.post(batch_chat_url, json=payload, timeout=kwargs.get('llm_mode_timeout', 120))
        resp.raise_for_status()
        data = resp.json()
        results = data.get('results', [])
    except Exception as e:
        m_logger.warning(f"[chunk][llm] 模式判斷失敗，回退為 row: {e}")
        return {idx: 'row' for idx in seg_indices}

    mode_map: Dict[int, str] = {}
    for i, item in enumerate(results):
        idx = seg_indices[i] if i < len(seg_indices) else None
        if idx is None:
            continue
        if item.get('error'):
            mode_map[idx] = 'row'
            continue
        result_obj = item.get('result', {}) if isinstance(item, dict) else {}
        output = ''
        if isinstance(result_obj, dict):
            output = result_obj.get('output', '') or ''
        mode, reason = _parse_mode_and_reason_from_llm_text(output)
        mode = mode or 'row'
        if mode not in {'row', 'column', 'preview', 'diagonal', 'cell'}:
            mode = 'row'
        mode_map[idx] = mode
        if reason:
            m_logger.info(f"[chunk][llm] seg_index={idx} mode={mode} reason={reason}")
        else:
            m_logger.info(f"[chunk][llm] seg_index={idx} mode={mode}")
    return mode_map


_XLSX_KV_LABELS = (
    '規格', '品名', '型號', '料號', '零件號', '產品名稱', '產品', '名稱', '圖號', '版本',
)
_XLSX_STOPWORDS = frozenset({
    'table', 'sheet', 'blank', 'col', 'unknown', 'empty', 'empty_sheet',
    '表格', '進度', '工作表', '空白', '明安', 'adg', 'adgroup', 'domain', 'launchtech',
    '審核', '保存期限', '平均', '標準', '公差', '半成品', '成品', '毛胚', 'σ', 'sigma',
    'index', 'columns', 'shape', 'original_table', 'table_block', 'range', 'block_shape',
})
_XLSX_COMPANY_NOISE_RE = re.compile(
    r'(有限公司|股份有限公司|co\.?\s*,?\s*ltd|corp\.?|inc\.?|group|sporting\s+goods|adg|adgroup|明安)',
    re.IGNORECASE,
)
_XLSX_PART_CODE_RE = re.compile(
    r'\b[A-Z]{1,4}\d{3,}[A-Z0-9\-]*(?:\s+[A-Z]{1,3}\d{2,})?(?:\s*\([A-Z0-9\-]+\))?',
)
_XLSX_MEASURE_RE = re.compile(
    r'(?:±\s*)?\d+(?:\.\d+)?\s*(?:mm|cm|m|g|kg|度|%|mpa|kpa|μm|um|mil)\b|±\s*\d+(?:\.\d+)?',
    re.IGNORECASE,
)


def _xlsx_prompt_file_path(name: str = 'extract_keywords_xlsx.json') -> str:
    return os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        'prompt',
        name,
    )


def _is_xlsx_noise_token(token: str) -> bool:
    if token is None:
        return True
    s = str(token).strip()
    if not s or len(s) < 2:
        return True
    lower = s.lower()
    if lower.startswith('#sheet:') or lower.startswith('## table'):
        return True
    if s.startswith('[') and ('blank' in lower or 'original_table' in lower or 'table_block' in lower):
        return True
    if s.startswith('=') or s.startswith('|'):
        return True
    if re.match(r'^col_\d+$', lower):
        return True
    if re.match(r'^\d{1,2}$', s):
        return True
    if re.match(r'^[\d\s\.\,\:\;\|\+\-]+$', s) and not re.search(r'[a-zA-Z\u4e00-\u9fff]', s):
        return True
    if lower in _XLSX_STOPWORDS:
        return True
    if _XLSX_COMPANY_NOISE_RE.search(s):
        return True
    if re.match(r'^\[blank\*?\d*\]$', lower):
        return True
    return False


def _dedupe_keywords_preserve_order(keywords: List[str]) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for kw in keywords or []:
        if kw is None:
            continue
        s = str(kw).strip()
        if not s or _is_xlsx_noise_token(s):
            continue
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


def _filter_xlsx_keywords(keywords: List[str]) -> List[str]:
    if not keywords:
        return []
    filtered = _filter_table_noise_keywords(keywords)
    return _dedupe_keywords_preserve_order(filtered)


def _merge_xlsx_keyword_lists(*sources: List[str], max_keywords: int = 50) -> List[str]:
    merged: List[str] = []
    for source in sources:
        if not source:
            continue
        for item in source:
            if item is None:
                continue
            merged.append(str(item).strip())
    merged = _filter_xlsx_keywords(merged)
    if max_keywords > 0:
        return merged[:max_keywords]
    return merged


def _table_df_cell_text(table_df: 'pd.DataFrame', row: int, col: int) -> str:
    if table_df is None or table_df.empty:
        return ''
    try:
        return _cell_to_display_text(table_df.iloc[row, col]).strip()
    except Exception:
        return ''


def _extract_xlsx_kv_terms(table_df: 'pd.DataFrame', max_items: int) -> List[str]:
    if table_df is None or getattr(table_df, 'empty', True):
        return []
    terms: List[str] = []
    nrows, ncols = table_df.shape
    label_set = set(_XLSX_KV_LABELS)
    for r in range(min(nrows, 30)):
        for c in range(ncols):
            raw = _table_df_cell_text(table_df, r, c)
            if not raw:
                continue
            inline_m = re.match(r'^(' + '|'.join(_XLSX_KV_LABELS) + r')\s*[:：]\s*(.+)$', raw)
            if inline_m:
                val = inline_m.group(2).strip()
                if val and not _is_xlsx_noise_token(val):
                    terms.append(val)
                continue
            normalized = raw.rstrip('：:').strip()
            if normalized in label_set:
                for cc in range(c + 1, ncols):
                    val = _table_df_cell_text(table_df, r, cc)
                    if val and not _is_xlsx_noise_token(val):
                        terms.append(val)
                        break
    return _dedupe_keywords_preserve_order(terms)[:max_items]


def _extract_xlsx_header_index_terms(
        table_df: 'pd.DataFrame',
        headers: Optional[List[str]] = None,
        max_items: int = 20,
    ) -> List[str]:
    terms: List[str] = []
    if headers:
        for h in headers:
            txt = _cell_to_display_text(h).strip()
            if txt and not _is_xlsx_noise_token(txt) and not re.match(r'^col_\d+$', txt.lower()):
                terms.append(txt)
    if table_df is None or getattr(table_df, 'empty', True):
        return _dedupe_keywords_preserve_order(terms)[:max_items]
    nrows, ncols = table_df.shape
    for r in range(min(nrows, 8)):
        for c in range(ncols):
            txt = _table_df_cell_text(table_df, r, c)
            if not txt or _is_xlsx_noise_token(txt):
                continue
            if re.match(r'^(區|分|支|號)$', txt):
                continue
            if len(txt) >= 2 and (re.search(r'[\u4e00-\u9fff]', txt) or re.search(r'[A-Za-z]{2,}', txt)):
                if txt not in ('標準', '公差', '平均', '半成品', '毛胚', '成品'):
                    terms.append(txt)
    return _dedupe_keywords_preserve_order(terms)[:max_items]


def _extract_xlsx_spec_and_measure_terms(table_df: 'pd.DataFrame', max_items: int) -> List[str]:
    if table_df is None or getattr(table_df, 'empty', True):
        return []
    terms: List[str] = []
    for row in table_df.values.tolist():
        if not isinstance(row, list):
            continue
        for cell in row:
            txt = _cell_to_display_text(cell).strip()
            if not txt or _is_xlsx_noise_token(txt):
                continue
            for m in _XLSX_PART_CODE_RE.findall(txt):
                if m and not _is_xlsx_noise_token(m):
                    terms.append(m.strip())
            for m in _XLSX_MEASURE_RE.findall(txt):
                if m and not _is_xlsx_noise_token(m):
                    terms.append(m.strip())
            if re.search(r'±', txt) and len(txt) <= 24:
                terms.append(txt)
    return _dedupe_keywords_preserve_order(terms)[:max_items]


def _extract_xlsx_distinctive_terms(
        chunk_df: 'pd.DataFrame',
        full_table_df: 'pd.DataFrame',
        max_items: int = 15,
    ) -> List[str]:
    if (
        chunk_df is None or full_table_df is None
        or getattr(chunk_df, 'empty', True) or getattr(full_table_df, 'empty', True)
    ):
        return []
    if chunk_df.shape == full_table_df.shape and chunk_df.equals(full_table_df):
        return []

    def _collect_values(df: 'pd.DataFrame') -> List[str]:
        values: List[str] = []
        for row in df.values.tolist():
            if not isinstance(row, list):
                continue
            for cell in row:
                txt = _cell_to_display_text(cell).strip()
                if txt and not _is_xlsx_noise_token(txt):
                    values.append(txt)
        return values

    full_counter = Counter(_collect_values(full_table_df))
    chunk_values = _collect_values(chunk_df)
    distinctive: List[str] = []
    seen: Set[str] = set()
    for txt in chunk_values:
        key = txt.lower()
        if key in seen:
            continue
        freq = full_counter.get(txt, 0)
        if freq <= 2 or (freq <= 4 and len(txt) >= 4):
            seen.add(key)
            distinctive.append(txt)
    distinctive.sort(key=lambda x: (full_counter.get(x, 0), -len(x)))
    return _dedupe_keywords_preserve_order(distinctive)[:max_items]


def _extract_xlsx_multi_prompts(
        table_df: Optional['pd.DataFrame'],
        headers: Optional[List[str]] = None,
        full_table_df: Optional['pd.DataFrame'] = None,
        max_keywords: int = 50,
    ) -> List[str]:
    """表格索引詞 + 內容語義詞（規則抽取）。"""
    if table_df is None or getattr(table_df, 'empty', True):
        return []
    index_terms = _extract_xlsx_header_index_terms(table_df, headers=headers, max_items=max(10, max_keywords // 3))
    kv_terms = _extract_xlsx_kv_terms(table_df, max_items=max(10, max_keywords // 4))
    spec_terms = _extract_xlsx_spec_and_measure_terms(table_df, max_items=max(15, max_keywords // 2))
    distinctive = _extract_xlsx_distinctive_terms(table_df, full_table_df, max_items=max(8, max_keywords // 4))
    merged = _merge_xlsx_keyword_lists(
        index_terms,
        kv_terms,
        spec_terms,
        distinctive,
        max_keywords=max_keywords,
    )
    return merged


def _build_xlsx_llm_keyword_context(
        table_df: Optional['pd.DataFrame'],
        headers: Optional[List[str]] = None,
        full_table_df: Optional['pd.DataFrame'] = None,
        max_rows: int = 25,
    ) -> str:
    if table_df is None or getattr(table_df, 'empty', True):
        return ''
    index_terms = _extract_xlsx_header_index_terms(table_df, headers=headers, max_items=20)
    kv_terms = _extract_xlsx_kv_terms(table_df, max_items=10)
    distinctive = _extract_xlsx_distinctive_terms(table_df, full_table_df, max_items=12)
    preview_df = table_df.head(max_rows)
    try:
        preview_md = _compress_markdown_blank_runs(_df_to_markdown(preview_df))
    except Exception:
        preview_md = ''
    lines = [
        '[索引詞候選] ' + ', '.join(index_terms[:20]),
        '[規格/品名候選] ' + ', '.join(kv_terms[:10]),
    ]
    if distinctive:
        lines.append('[本區塊特徵值] ' + ', '.join(distinctive[:12]))
    if preview_md:
        lines.append('[表格預覽]\n' + preview_md[:4000])
    return '\n'.join(lines)


def _load_extract_keywords_prompt_config(for_xlsx: bool = False) -> Dict[str, Any]:
    prompt_name = 'extract_keywords_xlsx.json' if for_xlsx else 'extract_keywords.json'
    prompt_file = _xlsx_prompt_file_path(prompt_name)
    if not os.path.exists(prompt_file):
        return {}
    try:
        with open(prompt_file, 'r', encoding='utf-8-sig') as f:
            return json.load(f)
    except Exception:
        return {}


def _batch_extract_keywords_with_prompt(
        chunk_texts: List[str],
        llm_provider: str,
        llm_model: str,
        llm_base_url: Optional[str],
        batch_size: int = 50,
        max_keywords: Optional[int] = None,
        for_xlsx: bool = False,
    ) -> List[List[str]]:
    if not chunk_texts:
        return []
    try:
        import requests
    except Exception:
        return [[] for _ in chunk_texts]

    prompt_cfg = _load_extract_keywords_prompt_config(for_xlsx=for_xlsx)
    if not prompt_cfg:
        return [[] for _ in chunk_texts]

    system_prompt = prompt_cfg.get('system_prompt', '')
    user_prompt_template = prompt_cfg.get('user_prompt_template', '')
    generation_config = prompt_cfg.get('generation_config', {}) or {}
    max_tokens = int(generation_config.get('max_new_tokens', 150))
    temperature = float(generation_config.get('temperature', 0.3))
    if max_keywords is None:
        max_keywords = int(generation_config.get('max_keywords', 5))
    else:
        max_keywords = int(max_keywords)
    min_keyword_length = int(generation_config.get('min_keyword_length', 1))
    base_url = (llm_base_url or m_config.get('llm', {}).get('base_url', 'http://10.1.3.127:7017')).rstrip('/')
    batch_chat_url = f"{base_url}/chat/batch"

    all_keywords: List[List[str]] = []
    batch_size = max(1, int(batch_size))
    for start in range(0, len(chunk_texts), batch_size):
        batch_texts = chunk_texts[start:start + batch_size]
        prompts: List[str] = []
        for content in batch_texts:
            fmt_kwargs = {'content': content}
            if '{max_keywords}' in user_prompt_template:
                fmt_kwargs['max_keywords'] = max_keywords
            user_prompt = user_prompt_template.format(**fmt_kwargs)
            full_prompt = f"{system_prompt}\n\n{user_prompt}" if system_prompt else user_prompt
            prompts.append(full_prompt)

        payload = {
            'prompts': prompts,
            'provider': llm_provider,
            'model': llm_model,
            'max_tokens': max_tokens,
            'temperature': temperature,
            'system_prompt': system_prompt if system_prompt else None,
            'parallel': True,
            'max_batch_size': 190
        }
        try:
            response = requests.post(batch_chat_url, json=payload, timeout=300)
            response.raise_for_status()
            items = (response.json() or {}).get('results', [])
        except Exception:
            items = []

        if not items:
            all_keywords.extend([[] for _ in batch_texts])
            continue

        for item in items:
            keywords: List[str] = []
            if not item.get('error'):
                result_obj = item.get('result', {}) if isinstance(item, dict) else {}
                output = result_obj.get('output', '').strip() if isinstance(result_obj, dict) else ''
                if parse_keywords_from_text is not None:
                    keywords = parse_keywords_from_text(
                        output,
                        max_keywords=max_keywords,
                        min_keyword_length=min_keyword_length
                    )
            all_keywords.append(list(dict.fromkeys(keywords))[:max_keywords])

        if len(items) < len(batch_texts):
            all_keywords.extend([[] for _ in range(len(batch_texts) - len(items))])

    return all_keywords[:len(chunk_texts)]


def _flatten_table_values_for_keywords(table_df: 'pd.DataFrame') -> List[str]:
    if table_df is None or getattr(table_df, 'empty', True):
        return []
    values: List[str] = []
    for row in table_df.values.tolist():
        if not isinstance(row, list):
            continue
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            values.append(s)
    return values


def _build_keyword_source_text_from_values(values: List[str], max_items: int = 300) -> str:
    if not values:
        return ''
    cleaned: List[str] = []
    seen = set()
    for raw in values:
        s = str(raw).strip()
        if not s:
            continue
        if re.match(r'^[\\|\\:\\-\\+\\s\\*#_]+$', s):
            continue
        if len(s) > 200:
            s = s[:200].strip()
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(s)
        if len(cleaned) >= max_items:
            break
    return "\\n".join(cleaned)


def _xlsx_hierarchy_escape_field(value: Any, max_len: int = 480) -> str:
    """
    階層欄位內避免出現 '::' 分隔符；並截斷過長字串。
    """
    text = '' if value is None else str(value)
    text = text.replace('\r', ' ').replace('\n', ' ')
    text = text.replace('::', '\u2237\u2237').strip()
    if len(text) > max_len:
        return text[: max_len - 1] + '\u2026'
    return text


def _apply_extract_multi_prompts_to_units(unit_paras: List[Dict[str, Any]], **kwargs) -> None:
    if not unit_paras or attach_multi_prompts_meta is None:
        return
    if is_enable_multi_prompts is not None and not is_enable_multi_prompts(kwargs):
        return
    if is_enable_multi_prompts is None and kwargs.get('enable_multi_prompts') is False:
        return
    max_keywords = int(kwargs.get('max_multi_prompts_per_unit', kwargs.get('max_keywords_per_chunk', 50)) or 50)
    llm_min = int(kwargs.get('xlsx_llm_keyword_min', 8) or 8)
    use_llm = kwargs.get('xlsx_use_llm_keywords', True)
    llm_contexts: List[str] = []
    llm_unit_indices: List[int] = []

    for idx, unit in enumerate(unit_paras):
        table_df = unit.get('sheet_non_empty_df')
        if table_df is None or getattr(table_df, 'empty', True):
            table_df = unit.get('sheet_df')
        headers = _guess_headers(table_df) if table_df is not None and not getattr(table_df, 'empty', True) else []
        prompts = _extract_xlsx_multi_prompts(
            table_df,
            headers=headers,
            max_keywords=max_keywords,
        )
        if use_llm and len(prompts) < llm_min and batch_extract_multi_prompts is not None:
            ctx = _build_xlsx_llm_keyword_context(table_df, headers=headers)
            if ctx and len(ctx) >= 10:
                llm_contexts.append(ctx)
                llm_unit_indices.append(idx)
        attach_multi_prompts_meta(unit, prompts)

    if llm_contexts and batch_extract_multi_prompts is not None:
        prompts_list = batch_extract_multi_prompts(
            llm_contexts,
            llm_base_url=kwargs.get('llm_base_url'),
            llm_provider=kwargs.get('llm_provider', 'remote'),
            llm_model=kwargs.get('llm_model', 'remote8b'),
            prompt_file=_xlsx_prompt_file_path('extract_keywords_xlsx.json'),
        )
        for unit_idx, llm_kw in zip(llm_unit_indices, prompts_list):
            merged = _merge_xlsx_keyword_lists(
                merge_multi_prompts([unit_paras[unit_idx]]) if merge_multi_prompts else [],
                llm_kw or [],
                max_keywords=max_keywords,
            )
            attach_multi_prompts_meta(unit_paras[unit_idx], merged)


def _propagate_unit_meta_to_xlsx_segments(
        unit_paras: List[Dict[str, Any]],
        segments: List[Dict[str, Any]],
    ) -> None:
    if not unit_paras or not segments or attach_segment_multi_prompts_meta is None:
        return
    unit_by_sheet: Dict[Tuple[Any, Any], Dict[str, Any]] = {}
    for unit in unit_paras:
        key = (unit.get('sheet_index'), unit.get('sheet_name'))
        unit_by_sheet[key] = unit
    for seg in segments:
        key = (seg.get('sheet_index'), seg.get('sheet_name'))
        unit = unit_by_sheet.get(key)
        if unit:
            attach_segment_multi_prompts_meta(seg, [unit])


def _apply_xlsx_hierarchy_to_multi_prompts(
        chunks: List[Dict[str, Any]],
        root_metadata: Optional[Dict[str, Any]],
    ) -> None:
    """
    將每個 chunk 的 multi_prompts 展成：
    <檔案名稱>::<tab名稱>::<儲存格／關鍵字內容>
    """
    if not chunks:
        return
    rm = root_metadata if isinstance(root_metadata, dict) else {}
    raw_path = rm.get('file_path') or rm.get('processed_file_path') or ''
    file_disp = _xlsx_hierarchy_escape_field(os.path.basename(str(raw_path)) or 'unknown', max_len=240)
    for ch in chunks:
        if not isinstance(ch, dict):
            continue
        sheet_disp = _xlsx_hierarchy_escape_field(str(ch.get('sheet_name') or '') or 'unknown', max_len=240)
        kws = ch.get('multi_prompts')
        if not isinstance(kws, list) or not kws:
            continue
        new_kws: List[str] = []
        for kw in kws:
            if kw is None:
                continue
            cell_part = _xlsx_hierarchy_escape_field(str(kw).strip(), max_len=600)
            if not cell_part:
                continue
            new_kws.append(f'{file_disp}::{sheet_disp}::{cell_part}')
        ch['multi_prompts'] = new_kws


def chunk(
        segments: List[Dict[str, Any]],
        parse_mode: str = 'preview',
        max_cell_chars: int = 100,
        digit: int = 6,
        extract_kw_lbd: int = 10,
        llm_provider: str = 'remote',
        llm_model: str = 'remote8b',
        llm_base_url: str = None,
        metadata: Dict[str, Any] = None,
        use_llm_index: bool = True,
        **kwargs
    ) -> List[Dict[str, Any]]:
    if not segments:
        return []
    # 固定路線：有效表格整塊輸出為單一 chunk（避免再細切）
    parse_mode = (parse_mode or 'preview').lower().strip()
    parse_mode = 'preview'

    chunks: List[Dict[str, Any]] = []
    keyword_source_map: Dict[int, str] = {}
    rule_keywords_map: Dict[int, List[str]] = {}
    chunk_order = 0
    # XLSX 預設允許每個 chunk 更大的 multi_prompts 額度（可用 kwargs 覆蓋）
    max_multi_prompts_per_chunk = int(
        kwargs.get('max_multi_prompts_per_chunk', kwargs.get('max_keywords_per_chunk', 50)) or 50
    )
    if max_multi_prompts_per_chunk <= 0:
        max_multi_prompts_per_chunk = 50
    if max_multi_prompts_per_chunk > 200:
        max_multi_prompts_per_chunk = 200
    # 批次處理：先跳過生成和去重，最後統一處理
    use_batch_dedup = dedup_multi_prompts_by_llm is not None
    use_batch_keyword_gen = True
    max_chunk_rows = int(kwargs.get('max_chunk_rows', kwargs.get('chunk_max_rows', 100)) or 100)
    max_chunk_cols = int(kwargs.get('max_chunk_cols', kwargs.get('chunk_max_cols', 15)) or 15)
    if max_chunk_rows <= 0:
        max_chunk_rows = 100
    if max_chunk_cols <= 0:
        max_chunk_cols = 15

    # 先收集所有 chunks（不進行去重）
    for idx, seg in enumerate(segments):
        table_df = seg.get('table_df')
        if table_df is None or getattr(table_df, 'empty', True):
            chunk_item = _build_chunk_from_text(
                chunk_text=seg.get('unit_text', ''),
                seg=seg,
                parse_mode=parse_mode,
                chunk_order=chunk_order,
                metadata=metadata,
                row_index=None,
                col_index=None,
                cell_ref=None,
                skip_dedup=use_batch_dedup,  # 先跳過去重
                skip_keyword_gen=use_batch_keyword_gen,  # 先跳過生成
                max_multi_prompts_per_chunk=max_multi_prompts_per_chunk,
            )
            plain_text = _strip_annotation_lines(seg.get('unit_text', '') or '')
            rule_keywords_map[chunk_order] = _filter_xlsx_keywords(
                _simple_keywords(plain_text, max_keywords=max_multi_prompts_per_chunk)
            )
            keyword_source_map[chunk_order] = plain_text
            chunks.append(chunk_item)
            chunk_order += 1
            continue

        total_rows, total_cols = table_df.shape
        is_split = total_rows > max_chunk_rows or total_cols > max_chunk_cols
        top_row = int(seg.get('top_row') or 1)
        left_col = int(seg.get('left_col') or 1)

        if not is_split:
            chunk_text = seg.get('unit_text', '') or ''
            if not chunk_text:
                try:
                    chunk_text = _compress_markdown_blank_runs(_df_to_markdown(table_df))
                except Exception:
                    chunk_text = ''
            chunk_item = _build_chunk_from_text(
                chunk_text=chunk_text,
                seg=seg,
                parse_mode='preview',
                chunk_order=chunk_order,
                metadata=metadata,
                row_index=None,
                col_index=None,
                cell_ref=None,
                skip_dedup=use_batch_dedup,  # 先跳過去重
                skip_keyword_gen=use_batch_keyword_gen,  # 先跳過生成
                max_multi_prompts_per_chunk=max_multi_prompts_per_chunk,
            )
            rule_keywords_map[chunk_order] = _extract_xlsx_multi_prompts(
                table_df,
                headers=seg.get('headers'),
                max_keywords=max_multi_prompts_per_chunk,
            )
            keyword_source_map[chunk_order] = _strip_annotation_lines(
                _build_xlsx_llm_keyword_context(table_df, headers=seg.get('headers')) or chunk_text
            )
            chunks.append(chunk_item)
            chunk_order += 1
            continue

        full_columns = [str(c) for c in table_df.columns.tolist()]
        full_index = [str(i) for i in table_df.index.tolist()]
        full_shape = [int(total_rows), int(total_cols)]
        full_context = (
            f"[ORIGINAL_TABLE] shape={full_shape} "
            f"columns={json.dumps(full_columns, ensure_ascii=False)} "
            f"index={json.dumps(full_index, ensure_ascii=False)}"
        )

        for row_start in range(0, total_rows, max_chunk_rows):
            row_end = min(row_start + max_chunk_rows, total_rows)
            for col_start in range(0, total_cols, max_chunk_cols):
                col_end = min(col_start + max_chunk_cols, total_cols)
                block_df = table_df.iloc[row_start:row_end, col_start:col_end].copy()
                block_md = _compress_markdown_blank_runs(_df_to_markdown(block_df))
                block_top_row = top_row + row_start
                block_left_col = left_col + col_start
                block_bottom_row = top_row + row_end - 1
                block_right_col = left_col + col_end - 1
                block_range = (
                    f"{_colnum_to_letter(block_left_col)}{block_top_row}:"
                    f"{_colnum_to_letter(block_right_col)}{block_bottom_row}"
                )
                chunk_text = (
                    f"{full_context}\n"
                    f"[TABLE_BLOCK] range={block_range} "
                    f"block_shape={[int(row_end - row_start), int(col_end - col_start)]}\n"
                    f"{block_md}"
                )
                chunk_item = _build_chunk_from_text(
                    chunk_text=chunk_text,
                    seg=seg,
                    parse_mode='preview',
                    chunk_order=chunk_order,
                    metadata=metadata,
                    row_index=block_top_row,
                    col_index=block_left_col,
                    cell_ref=block_range,
                    skip_dedup=use_batch_dedup,  # 先跳過去重
                    skip_keyword_gen=use_batch_keyword_gen,  # 先跳過生成
                    max_multi_prompts_per_chunk=max_multi_prompts_per_chunk,
                )
                chunk_item['block_bbox'] = [block_top_row, block_left_col, block_bottom_row, block_right_col]
                chunk_item['original_table_shape'] = full_shape
                chunk_item['original_table_columns'] = full_columns
                chunk_item['original_table_index'] = full_index
                rule_keywords_map[chunk_order] = _extract_xlsx_multi_prompts(
                    block_df,
                    headers=seg.get('headers'),
                    full_table_df=table_df,
                    max_keywords=max_multi_prompts_per_chunk,
                )
                keyword_source_map[chunk_order] = _strip_annotation_lines(
                    _build_xlsx_llm_keyword_context(
                        block_df,
                        headers=seg.get('headers'),
                        full_table_df=table_df,
                    ) or chunk_text
                )
                chunks.append(chunk_item)
                chunk_order += 1

    xlsx_llm_keyword_min = int(kwargs.get('xlsx_llm_keyword_min', 8) or 8)
    use_xlsx_llm = kwargs.get('xlsx_use_llm_keywords', True)

    # 批次生成 keywords：規則抽取為主，不足時再以 xlsx 專用 prompt 補強
    if chunks:
        try:
            chunk_texts: List[str] = []
            chunk_indices_for_gen: List[int] = []
            for i, ch in enumerate(chunks):
                rule_kw = rule_keywords_map.get(i, [])
                upstream_kw = _filter_xlsx_keywords(merge_multi_prompts([ch]) if merge_multi_prompts else [])
                keywords = _merge_xlsx_keyword_lists(
                    rule_kw,
                    upstream_kw,
                    max_keywords=max_multi_prompts_per_chunk,
                )
                if keywords:
                    ch['multi_prompts'] = keywords
                if use_xlsx_llm and len(keywords) < xlsx_llm_keyword_min:
                    source_text = keyword_source_map.get(i, '') or ch.get('text', '') or ch.get('unit_text', '')
                    if source_text and len(source_text) >= 10:
                        chunk_texts.append(source_text)
                        chunk_indices_for_gen.append(i)

            if chunk_texts:
                batch_size = kwargs.get('keywords_batch_size', 50)
                generated_keywords: List[List[str]] = []
                for start in range(0, len(chunk_texts), batch_size):
                    batch_texts = chunk_texts[start:start + batch_size]
                    for text in batch_texts:
                        try:
                            generated = _batch_extract_keywords_with_prompt(
                                chunk_texts=[text],
                                llm_provider=llm_provider,
                                llm_model=llm_model,
                                llm_base_url=llm_base_url,
                                batch_size=1,
                                max_keywords=max_multi_prompts_per_chunk,
                                for_xlsx=True,
                            )
                            kw = generated[0] if generated else []
                            if not kw:
                                kw = _filter_xlsx_keywords(
                                    _simple_keywords(text, max_keywords=max_multi_prompts_per_chunk)
                                )
                            if filter_meaningless_tags is not None:
                                try:
                                    kw = _filter_xlsx_keywords(filter_meaningless_tags(kw) or kw)
                                except Exception:
                                    kw = _filter_xlsx_keywords(kw)
                            else:
                                kw = _filter_xlsx_keywords(kw)
                            generated_keywords.append(kw)
                        except Exception:
                            generated_keywords.append(
                                _filter_xlsx_keywords(
                                    _simple_keywords(text, max_keywords=max_multi_prompts_per_chunk)
                                )
                            )

                for idx, llm_kw in zip(chunk_indices_for_gen, generated_keywords):
                    merged = _merge_xlsx_keyword_lists(
                        chunks[idx].get('multi_prompts', []) or [],
                        llm_kw or [],
                        max_keywords=max_multi_prompts_per_chunk,
                    )
                    if merged:
                        chunks[idx]['multi_prompts'] = merged

                m_logger.info(f"[chunk] 批次生成 keywords 完成: {len(chunk_indices_for_gen)} 個 chunks 使用 LLM 補強")
        except Exception as e:
            m_logger.warning(f"[chunk] 批次生成 keywords 失敗，回退為單個處理: {e}")
            for i, ch in enumerate(chunks):
                rule_kw = rule_keywords_map.get(i, [])
                upstream_kw = _filter_xlsx_keywords(merge_multi_prompts([ch]) if merge_multi_prompts else [])
                keywords = _merge_xlsx_keyword_lists(
                    rule_kw,
                    upstream_kw,
                    max_keywords=max_multi_prompts_per_chunk,
                )
                if keywords:
                    ch['multi_prompts'] = keywords
                    continue
                chunk_text = keyword_source_map.get(i, '') or ch.get('text', '') or ch.get('unit_text', '')
                if chunk_text and len(chunk_text) >= 10:
                    try:
                        generated = _batch_extract_keywords_with_prompt(
                            chunk_texts=[chunk_text],
                            llm_provider=llm_provider,
                            llm_model=llm_model,
                            llm_base_url=llm_base_url,
                            batch_size=1,
                            max_keywords=max_multi_prompts_per_chunk,
                            for_xlsx=True,
                        )
                        llm_kw = generated[0] if generated else []
                        if not llm_kw:
                            llm_kw = _filter_xlsx_keywords(
                                _simple_keywords(chunk_text, max_keywords=max_multi_prompts_per_chunk)
                            )
                        ch['multi_prompts'] = _merge_xlsx_keyword_lists(
                            keywords,
                            llm_kw,
                            max_keywords=max_multi_prompts_per_chunk,
                        )
                    except Exception:
                        pass
    
    # 批次去重：收集所有 keywords 後批次處理
    if use_batch_dedup and chunks:
        try:
            # 收集所有 keywords 列表
            all_keywords_lists: List[List[str]] = []
            chunk_indices: List[int] = []  # 記錄每個 keywords 對應的 chunk index
            for i, ch in enumerate(chunks):
                keywords = ch.get('multi_prompts', [])
                if keywords and isinstance(keywords, list):
                    all_keywords_lists.append(keywords)
                    chunk_indices.append(i)
            
            # 批次去重（dedup_multi_prompts_by_llm 內部有 batch_size 限制，預設 8）
            if all_keywords_lists:
                deduped_lists = dedup_multi_prompts_by_llm(all_keywords_lists)
                if deduped_lists and len(deduped_lists) == len(all_keywords_lists):
                    # 將去重後的結果分配回對應的 chunks
                    for idx, deduped_kw in zip(chunk_indices, deduped_lists):
                        if isinstance(deduped_kw, list):
                            chunks[idx]['multi_prompts'] = deduped_kw
                        elif deduped_kw:
                            # 如果返回的不是 list，嘗試轉換
                            chunks[idx]['multi_prompts'] = [str(deduped_kw)] if deduped_kw else []
                    m_logger.info(f"[chunk] 批次去重完成: {len(all_keywords_lists)} 個 keywords 列表")
        except Exception as e:
            m_logger.warning(f"[chunk] 批次去重失敗，回退為單個處理: {e}")
            # 回退：對每個 chunk 單獨去重
            for ch in chunks:
                keywords = ch.get('multi_prompts', [])
                if keywords and isinstance(keywords, list):
                    try:
                        deduped = dedup_multi_prompts_by_llm([keywords])
                        if deduped and isinstance(deduped, list) and len(deduped) > 0:
                            ch['multi_prompts'] = deduped[0] if isinstance(deduped[0], list) else keywords
                    except Exception:
                        pass
    
    _apply_xlsx_hierarchy_to_multi_prompts(chunks, metadata)
    return chunks



def _build_chunk_from_text(
        chunk_text: str,
        seg: Dict[str, Any],
        parse_mode: str,
        chunk_order: int,
        metadata: Optional[Dict[str, Any]],
        row_index: Optional[int],
        col_index: Optional[int],
        cell_ref: Optional[str],
        skip_dedup: bool = False,
        skip_keyword_gen: bool = False,
        max_multi_prompts_per_chunk: int = 50,
    ) -> Dict[str, Any]:
    upstream_mp = merge_multi_prompts([seg]) if merge_multi_prompts else []
    keywords: List[str] = []
    # 如果跳過生成，使用空列表（將在批次處理中生成）
    if not skip_keyword_gen and not upstream_mp:
        if len(chunk_text) >= 10:
            try:
                generated = _batch_extract_keywords_with_prompt(
                    chunk_texts=[chunk_text],
                    llm_provider='remote',
                    llm_model='remote8b',
                    llm_base_url=None,
                    batch_size=1,
                    max_keywords=max_multi_prompts_per_chunk,
                )
                keywords = generated[0] if generated else []
            except Exception:
                keywords = []
        if not keywords:
            keywords = _simple_keywords(chunk_text, max_keywords=max_multi_prompts_per_chunk)
        if filter_meaningless_tags is not None:
            try:
                keywords = filter_meaningless_tags(keywords)
                keywords = _filter_xlsx_keywords(keywords)
            except Exception:
                keywords = _filter_xlsx_keywords(keywords)
        else:
            keywords = _filter_xlsx_keywords(keywords)
        if not skip_dedup and dedup_multi_prompts_by_llm is not None:
            try:
                keywords = dedup_multi_prompts_by_llm([keywords]) or keywords
            except Exception:
                pass
    
    # 確保 keywords 是 List[str]，扁平化嵌套列表
    flattened_keywords: List[str] = []
    for kw in keywords:
        if isinstance(kw, str):
            flattened_keywords.append(kw)
        elif isinstance(kw, (list, tuple)):
            flattened_keywords.extend(str(item) for item in kw if item)
        else:
            flattened_keywords.append(str(kw))
    keywords = flattened_keywords
    if merge_multi_prompts:
        final_multi_prompts = merge_multi_prompts([upstream_mp, keywords])
    else:
        final_multi_prompts = keywords or upstream_mp

    call_prompt = chunk_text
    chunk_item = {
        'chunk_id': f'chunk_{chunk_order:04d}',
        'order': chunk_order,
        'unit_text': chunk_text,
        'text': chunk_text,
        'call_prompt': call_prompt,
        'multi_prompts': final_multi_prompts,
        'indent_level': 0,
        'parse_mode': parse_mode,
        'sheet_name': seg.get('sheet_name'),
        'sheet_index': seg.get('sheet_index'),
        'segment_id': seg.get('segment_id'),
        'segment_type': seg.get('segment_type'),
        'table_id': seg.get('table_id'),
        'row_index': row_index,
        'column_index': col_index,
        'cell_ref': cell_ref,
        'table_bbox': [seg.get('top_row'), seg.get('left_col'), seg.get('bottom_row'), seg.get('right_col')],
        'headers': seg.get('headers', []),
        'images': seg.get('images', []),
        'sheet_images': seg.get('sheet_images', []),
        'formula_cells': seg.get('formula_cells', []),
        'metadata': {
            'sheet_name': seg.get('sheet_name'),
            'table_id': seg.get('table_id'),
            'parse_mode': parse_mode,
        }
    }
    if attach_multi_prompts_meta is not None:
        attach_multi_prompts_meta(chunk_item, upstream_mp)
    return chunk_item



def _simple_keywords(text: str, max_keywords: int = 10) -> List[str]:
    tokens = re.findall(r'[A-Za-z0-9_\-\u4e00-\u9fff]{2,}', text)
    seen = set()
    results = []
    for token in tokens:
        t = token.strip()
        if not t:
            continue
        key = t.lower()
        if key in seen:
            continue
        seen.add(key)
        results.append(t)
        if len(results) >= max_keywords:
            break
    return results


def _filter_table_noise_keywords(keywords: List[str]) -> List[str]:
    if not keywords:
        return []
    filtered: List[str] = []
    for keyword in keywords:
        if keyword is None:
            continue
        s = str(keyword).strip()
        if not s:
            continue
        lower = s.lower()
        if lower.startswith('#sheet:'):
            continue
        if re.match(r'^#{1,6}\s*table\s*\d*$', lower):
            continue
        if s.startswith('|') and s.endswith('|'):
            continue
        if re.match(r'^[\|\:\-\+\s]+$', s):
            continue
        filtered.append(s)
    return filtered


# ============================================================
# 5. Process（一條龍）
# ============================================================

def process(
    file_path: str,
    parse_mode: str = 'preview',
        max_cell_chars: int = 100,
        digit: int = 6,
        suitable_char_count: int = 500,
        include_images: bool = True,
        image_placeholder: bool = True,
        enable_image_llm: bool = True,
        llm_provider: str = 'remote',
        llm_model: str = 'remote8b',
        llm_base_url: str = None,
        image_llm_provider: str = 'openai',
        image_llm_model: str = 'gpt4o_chat',
        table_detector: Optional[str] = None,
        inspector_table_params: Optional[Dict[str, Any]] = None,
        feature_scan: bool = False,
        feature_scan_params: Optional[Dict[str, Any]] = None,
        include_textboxes: bool = False,
        **kwargs
    ) -> Dict[str, Any]:
    m_logger.info(f'[process] 開始處理檔案: {file_path}')
    extract_result = extract(
        file_path=file_path,
        include_images=include_images,
        image_placeholder=image_placeholder,
        table_detector=table_detector,
        inspector_table_params=inspector_table_params,
        feature_scan=feature_scan,
        feature_scan_params=feature_scan_params,
        include_textboxes=include_textboxes,
        **kwargs
    )
    unit_paras = extract_result['unit_paras']
    metadata = extract_result.get('metadata', {})

    segments = segment(
        unit_paras=unit_paras,
        suitable_char_count=suitable_char_count,
        metadata=metadata,
        enable_image_llm=enable_image_llm,
        llm_provider=llm_provider,
        llm_model=llm_model,
        llm_base_url=llm_base_url,
        image_llm_provider=image_llm_provider,
        image_llm_model=image_llm_model,
        table_detector=table_detector,
        feature_scan=feature_scan,
        **kwargs
    )
    chunks = chunk(
        segments=segments,
        parse_mode=parse_mode,
        max_cell_chars=max_cell_chars,
        digit=digit,
        llm_provider=llm_provider,
        llm_model=llm_model,
        llm_base_url=llm_base_url,
        metadata=metadata,
        **kwargs
    )

    stats = {
        'sheet_count': metadata.get('sheet_count', 0),
        'sheet_units': len(unit_paras),
        'segment_count': len(segments),
        'chunk_count': len(chunks),
        'table_count': metadata.get('total_tables', 0),
        'formula_count': metadata.get('formula_count', 0),
        'image_count': metadata.get('image_count', 0),
        'parse_mode': parse_mode,
    }

    return {
        'chunks': chunks,
        'unit_paras': unit_paras,
        'segments': segments,
        'metadata': metadata,
        'metadata_display': _sanitize_metadata(metadata),
        'stats': stats,
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python xlsx_parser.py <excel_file_path> [row|column|cell]')
        sys.exit(1)

    file_path = sys.argv[1]
    parse_mode = sys.argv[2] if len(sys.argv) > 2 else 'row'

    print('=== Preview ===')
    print(preview(file_path)[:2000])

    print('\n=== Process ===')
    result = process(file_path, parse_mode=parse_mode, enable_image_llm=False)
    print(json.dumps(result['stats'], ensure_ascii=False, indent=2))
    print('\n=== Metadata Summary ===')
    print(json.dumps(result['metadata_display'], ensure_ascii=False, indent=2))
