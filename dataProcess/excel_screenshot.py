#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Shared Excel range screenshot helpers for dataProcess."""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl.utils.cell import column_index_from_string, get_column_letter
from PIL import Image, ImageDraw, ImageFont


def _pick_font(size: int) -> ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/msjh.ttc",
        "C:/Windows/Fonts/mingliu.ttc",
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simsun.ttc",
    ]
    for font_path in candidates:
        try:
            return ImageFont.truetype(font_path, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _normalize_center(center_cell: str) -> Tuple[int, int]:
    text = center_cell.strip().upper()
    i = 0
    while i < len(text) and text[i].isalpha():
        i += 1
    if i == 0 or i == len(text):
        raise ValueError(f"Invalid center_cell '{center_cell}'")
    col = column_index_from_string(text[:i])
    row = int(text[i:])
    if col < 1 or row < 1:
        raise ValueError(f"Invalid center_cell '{center_cell}'")
    return col, row


def capture_excel_range_image(
    ws,
    center_cell: str,
    up: int,
    down: int,
    left: int,
    right: int,
    output_path: str,
) -> dict:
    """Render a worksheet area around center cell into a PNG image."""
    img, result = _render_excel_range_panel(ws, center_cell, up, down, left, right)
    target = Path(output_path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    img.save(target, format="PNG")
    result["path"] = str(target)
    return result


def _render_excel_range_panel(
    ws,
    center_cell: str,
    up: int,
    down: int,
    left: int,
    right: int,
) -> Tuple[Image.Image, Dict[str, Any]]:
    """Render one range panel and return image + metadata."""
    center_col, center_row = _normalize_center(center_cell)
    up = max(0, int(up))
    down = max(0, int(down))
    left = max(0, int(left))
    right = max(0, int(right))

    start_col = max(1, center_col - left)
    end_col = min(ws.max_column or center_col, center_col + right)
    start_row = max(1, center_row - up)
    end_row = min(ws.max_row or center_row, center_row + down)

    cols = end_col - start_col + 1
    rows = end_row - start_row + 1
    if cols <= 0 or rows <= 0:
        raise ValueError("Resolved screenshot range is empty.")

    cell_w = 100
    cell_h = 30
    header_h = 24
    row_label_w = 46
    pad = 8
    title_h = 34

    width = row_label_w + cols * cell_w + pad * 2
    height = title_h + header_h + rows * cell_h + pad * 2

    img = Image.new("RGB", (width, height), color="white")
    draw = ImageDraw.Draw(img)
    font = _pick_font(12)
    font_small = _pick_font(11)

    title = f"{ws.title} | {center_cell.upper()}"
    draw.text((pad, pad), title, fill="#003366", font=font)

    y0 = pad + title_h
    draw.rectangle([pad, y0, pad + row_label_w, y0 + header_h], fill="#EFEFEF", outline="black", width=1)

    for c in range(cols):
        col_no = start_col + c
        x1 = pad + row_label_w + c * cell_w
        x2 = x1 + cell_w
        draw.rectangle([x1, y0, x2, y0 + header_h], fill="#EFEFEF", outline="black", width=1)
        draw.text((x1 + 6, y0 + 5), get_column_letter(col_no), fill="#222222", font=font_small)

    for r in range(rows):
        row_no = start_row + r
        y1 = y0 + header_h + r * cell_h
        y2 = y1 + cell_h
        draw.rectangle([pad, y1, pad + row_label_w, y2], fill="#EFEFEF", outline="black", width=1)
        draw.text((pad + 6, y1 + 6), str(row_no), fill="#222222", font=font_small)

        for c in range(cols):
            col_no = start_col + c
            x1 = pad + row_label_w + c * cell_w
            x2 = x1 + cell_w
            draw.rectangle([x1, y1, x2, y2], fill="white", outline="black", width=1)
            value = ws.cell(row=row_no, column=col_no).value
            text = "" if value is None else str(value).replace("\n", " ")
            draw.text((x1 + 4, y1 + 6), text[:20], fill="#000000", font=font_small)

            if row_no == center_row and col_no == center_col:
                draw.rectangle([x1 + 1, y1 + 1, x2 - 1, y2 - 1], outline="#D60000", width=2)
    return (
        img,
        {
            "sheet": ws.title,
            "center_cell": f"{get_column_letter(center_col)}{center_row}",
            "range": f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}",
            "rows": rows,
            "cols": cols,
        },
    )


def capture_excel_ranges_figure(
    ws,
    tasks: List[Dict[str, Any]],
    output_path: str,
    ncols: int = 3,
    panel_gap: int = 16,
    panel_padding: int = 8,
) -> Dict[str, Any]:
    """Render multiple center-cell captures into one PNG figure."""
    if not isinstance(tasks, list) or not tasks:
        raise ValueError("'tasks' must be a non-empty list.")

    panel_gap = max(4, int(panel_gap))
    panel_padding = max(4, int(panel_padding))
    ncols = max(1, int(ncols))

    panels: List[Dict[str, Any]] = []
    max_w = 0
    max_h = 0

    for idx, task in enumerate(tasks, start=1):
        if not isinstance(task, dict):
            raise ValueError(f"tasks[{idx - 1}] must be an object.")
        center_cell = str(task.get("center_cell", "")).strip()
        if not center_cell:
            raise ValueError(f"tasks[{idx - 1}].center_cell is required.")
        up = int(task.get("up", 8))
        down = int(task.get("down", 8))
        left = int(task.get("left", 8))
        right = int(task.get("right", 8))

        panel_img, single_result = _render_excel_range_panel(
            ws=ws,
            center_cell=center_cell,
            up=up,
            down=down,
            left=left,
            right=right,
        )

        label = str(task.get("label", f"#{idx} {single_result['center_cell']}")).strip()
        panels.append(
            {
                "image": panel_img,
                "label": label,
                "center_cell": single_result["center_cell"],
                "range": single_result["range"],
                "rows": single_result["rows"],
                "cols": single_result["cols"],
            }
        )
        max_w = max(max_w, panel_img.width)
        max_h = max(max_h, panel_img.height)

    panel_title_h = 26
    panel_w = max_w + panel_padding * 2
    panel_h = max_h + panel_padding * 2 + panel_title_h
    n = len(panels)
    nrows = int(math.ceil(n / ncols))
    fig_w = ncols * panel_w + (ncols + 1) * panel_gap
    fig_h = nrows * panel_h + (nrows + 1) * panel_gap + 34

    fig = Image.new("RGB", (fig_w, fig_h), "white")
    draw = ImageDraw.Draw(fig)
    font = _pick_font(13)
    font_small = _pick_font(11)
    draw.text((panel_gap, 8), f"{ws.title} | batch captures: {n}", fill="#003366", font=font)

    items: List[Dict[str, Any]] = []
    for i, panel in enumerate(panels):
        r = i // ncols
        c = i % ncols
        x = panel_gap + c * (panel_w + panel_gap)
        y = panel_gap + 34 + r * (panel_h + panel_gap)
        draw.rectangle([x, y, x + panel_w, y + panel_h], outline="#A0A0A0", width=1)
        draw.text((x + panel_padding, y + 4), panel["label"], fill="#222222", font=font_small)

        panel_img = panel["image"]
        px = x + panel_padding
        py = y + panel_padding + panel_title_h
        fig.paste(panel_img, (px, py))

        items.append(
            {
                "index": i,
                "label": panel["label"],
                "center_cell": panel["center_cell"],
                "range": panel["range"],
                "rows": panel["rows"],
                "cols": panel["cols"],
                "panel_rect": {"x": x, "y": y, "w": panel_w, "h": panel_h},
            }
        )

    target = Path(output_path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    fig.save(target, format="PNG")
    return {
        "path": str(target),
        "sheet": ws.title,
        "count": n,
        "layout": {"rows": nrows, "cols": ncols, "panel_width": panel_w, "panel_height": panel_h},
        "items": items,
    }
