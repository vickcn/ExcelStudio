# -*- coding: utf-8 -*-
"""
Mark suspect cells in an Excel file using baseline discovered_rules.json.
No LLM calls are used.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except Exception as exc:  # pragma: no cover
    raise SystemExit("openpyxl not installed. Please install it: pip install openpyxl") from exc

from universal_table_detector import UniversalTableDetector
from window_scanner import WindowScanner
import pandas as pd

import rule_audit_io
import rule_audit_rules
import rule_audit_eval


def _load_main_dataframe(excel_path: Path) -> Tuple[Dict[str, Any], Any]:
    detector = UniversalTableDetector()
    results = detector.detect_tables_by_analysis(
        str(excel_path),
        table_mode="pure_numeric",
        use_llm=False,
    )
    if not results:
        raise RuntimeError("no table detected")
    result = results[0]
    dataframes = detector.extract_dataframes(str(excel_path), results)
    if not dataframes:
        raise RuntimeError("no dataframe extracted")
    return result, dataframes[0]


def _load_main_dataframe_from_full(full_df: pd.DataFrame) -> Tuple[Dict[str, Any], Any]:
    detector = UniversalTableDetector()
    results = detector._detect_tables_from_full(
        full_df,
        table_mode="pure_numeric",
        use_llm=False,
        prefer_local=False,
    )
    if not results:
        raise RuntimeError("no table detected")
    result = results[0]
    dataframes = detector.extract_dataframes_from_full(full_df, results)
    if not dataframes:
        raise RuntimeError("no dataframe extracted")
    return result, dataframes[0]


def _build_window_details(windows: List[Dict[str, Any]], scanner: WindowScanner) -> Tuple[List[Dict[str, Any]], int]:
    details: List[Dict[str, Any]] = []
    skipped = 0
    for w in windows:
        prompt = scanner.generate_prompt_data(w)
        if not prompt.get("has_numeric"):
            skipped += 1
            continue
        details.append({
            "window_info": {
                "window_id": w.get("window_id"),
                "start_loc_row_name": w.get("start_loc_row_name"),
                "excel_range": w.get("excel_range"),
                "position": w.get("position"),
            },
            "prompt_data": {
                "values": prompt.get("values"),
            },
        })
    return details, skipped


def _extract_indices(exprs: List[str]) -> List[Tuple[int, int]]:
    import re
    indices: List[Tuple[int, int]] = []
    for expr in exprs:
        if not isinstance(expr, str):
            continue
        for r, c in re.findall(r"\$\(\s*(-?\d+)\s*,\s*(-?\d+)\s*\)", expr):
            indices.append((int(r), int(c)))
        for r, c in re.findall(r"(?<!\$)\(\s*(-?\d+)\s*,\s*(-?\d+)\s*\)", expr):
            indices.append((int(r), int(c)))
        for r in re.findall(r"\$(\d+)", expr):
            indices.append((int(r), 0))
    return indices


def _col_to_excel(col_idx: int) -> str:
    # col_idx is 1-based
    if col_idx <= 0:
        return "A"
    letters: List[str] = []
    col = col_idx
    while col > 0:
        col, rem = divmod(col - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def audit_table_data(
    table: List[List[Any]],
    baseline_rules: Path,
    *,
    sheet_name: Optional[str] = None,
    row_name: Optional[str] = None,
    window_height: int = 3,
    window_width: int = 1,
    tolerance: float = 0.01,
    strict_row_match: bool = False,
) -> Dict[str, Any]:
    """以 JSON 表格進行稽核（不輸出 Excel 檔）。"""
    baseline_json = rule_audit_io.load_json(baseline_rules)
    baseline_rules_list = rule_audit_rules.extract_passed_rules(baseline_json)
    if not baseline_rules_list:
        raise RuntimeError("baseline rules empty")

    full_df = pd.DataFrame(table)
    table_result, main_df = _load_main_dataframe_from_full(full_df)
    values_range = table_result.get("values_range") or {}
    row_offset = int(values_range.get("start_row", 1)) - 1
    col_offset = int(values_range.get("start_col", 1)) - 1

    scanner = WindowScanner((window_height, window_width))
    windows = scanner.scan_dataframe(main_df)
    if row_name:
        windows = scanner.filter_windows_by_start_loc_row(windows, row_name)

    details, skipped = _build_window_details(windows, scanner)

    audit_report = rule_audit_eval.audit_windows_with_baseline(
        details,
        baseline_rules_list,
        row_name=row_name,
        tolerance=tolerance,
        strict_row_match=strict_row_match,
    )

    suspect_cells: set[Tuple[int, int]] = set()
    for fw in audit_report.get("failed_windows", []):
        pos = fw.get("position") or {}
        start_row = pos.get("start_row")
        start_col = pos.get("start_col")
        if start_row is None or start_col is None:
            continue
        for fr in fw.get("failed_rules", []):
            sides = fr.get("equation_sides")
            if not (isinstance(sides, list) and len(sides) == 2):
                continue
            indices = _extract_indices([str(sides[0]), str(sides[1])])
            for r, c in indices:
                if r < 0 or c < 0:
                    continue
                if r >= window_height or c >= window_width:
                    continue
                abs_row = row_offset + int(start_row) + r + 1
                abs_col = col_offset + int(start_col) + c + 1
                suspect_cells.add((abs_row, abs_col))

    suspect_cells_list: List[Dict[str, Any]] = []
    for r, c in sorted(suspect_cells):
        suspect_cells_list.append({
            "row": r,
            "col": c,
            "address": f"{_col_to_excel(c)}{r}",
        })

    return {
        "marked_cells": len(suspect_cells_list),
        "failed_windows": audit_report.get("summary", {}).get("failed_windows", 0),
        "output_excel": None,
        "suspect_sheet": sheet_name or table_result.get("sheet_name"),
        "suspect_cells": suspect_cells_list,
        "suspect_cells_count": len(suspect_cells_list),
    }


def _collect_row_stats(
    details: List[Dict[str, Any]],
    baseline_rules: List[Dict[str, Any]],
    *,
    tolerance: float,
    strict_row_match: bool,
) -> Dict[str, Dict[str, Any]]:
    row_stats: Dict[str, Dict[str, Any]] = {}
    for w in details:
        info = w.get("window_info") or {}
        row_name = info.get("start_loc_row_name") or "unknown"

        rules_to_use = [
            r for r in baseline_rules
            if r.get("_row_name") in (None, "", row_name)
        ]
        if not rules_to_use:
            if strict_row_match:
                continue
            rules_to_use = baseline_rules

        values = (w.get("prompt_data") or {}).get("values") or []

        stats = row_stats.setdefault(row_name, {
            "windows": 0,
            "rule_checks": 0,
            "passed_checks": 0,
            "failed_checks": 0,
        })
        stats["windows"] += 1

        for rule in rules_to_use:
            stats["rule_checks"] += 1
            eval_result = rule_audit_eval.evaluate_rule_on_window(
                rule,
                values,
                tolerance=tolerance,
            )
            if eval_result.get("ok"):
                stats["passed_checks"] += 1
            else:
                stats["failed_checks"] += 1

    for row_name, stats in row_stats.items():
        total = stats["rule_checks"]
        stats["pass_rate"] = stats["passed_checks"] / total if total else 0.0
    return row_stats

def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="Mark suspect cells in Excel using baseline discovered_rules.json",
    )
    parser.add_argument("--baseline-rules", required=True, type=Path, help="A 的 discovered_rules.json")
    parser.add_argument("--excel", required=True, type=Path, help="要檢查的 Excel 檔")
    parser.add_argument("--out-excel", required=True, type=Path, help="輸出 Excel 檔")
    parser.add_argument("--row-name", default=None, help="只檢查指定 start_loc_row_name")
    parser.add_argument("--window-height", type=int, default=3)
    parser.add_argument("--window-width", type=int, default=1)
    parser.add_argument("--tolerance", type=float, default=0.01)
    parser.add_argument("--strict-row-match", action="store_true", help="只套用相同 row 的規則")
    return parser

def main(args, ret: Dict[str, Any] = None) -> None:
    excel_path = args.excel
    if not excel_path.is_absolute():
        excel_path = Path.cwd() / excel_path
    if not excel_path.exists():
        raise SystemExit(f"excel not found: {excel_path}")

    baseline_json = rule_audit_io.load_json(args.baseline_rules)
    baseline_rules = rule_audit_rules.extract_passed_rules(baseline_json)
    if not baseline_rules:
        raise SystemExit("baseline rules empty")

    table_result, main_df = _load_main_dataframe(excel_path)
    values_range = table_result.get("values_range") or {}
    row_offset = int(values_range.get("start_row", 1)) - 1
    col_offset = int(values_range.get("start_col", 1)) - 1

    scanner = WindowScanner((args.window_height, args.window_width))
    windows = scanner.scan_dataframe(main_df)
    if args.row_name:
        windows = scanner.filter_windows_by_start_loc_row(windows, args.row_name)

    details, skipped = _build_window_details(windows, scanner)

    audit_report = rule_audit_eval.audit_windows_with_baseline(
        details,
        baseline_rules,
        row_name=args.row_name,
        tolerance=args.tolerance,
        strict_row_match=args.strict_row_match,
    )

    # row mapping: row_name -> baseline id
    row_to_id: Dict[str, str] = {}
    for key, group in (baseline_json or {}).items():
        if not isinstance(group, dict):
            continue
        rn = group.get("start_loc_row_name")
        if rn and rn not in row_to_id:
            row_to_id[rn] = str(key)

    row_stats = _collect_row_stats(
        details,
        baseline_rules,
        tolerance=args.tolerance,
        strict_row_match=args.strict_row_match,
    )

    print(f"\n=== Row Rule Pass Summary ===")
    print(f"file: {excel_path}")
    for row_name, stats in sorted(row_stats.items(), key=lambda x: x[0]):
        baseline_id = row_to_id.get(row_name, "n/a")
        print(
            f"- row_name={row_name} baseline_id={baseline_id} "
            f"windows={stats['windows']} "
            f"rule_checks={stats['rule_checks']} "
            f"pass_rate={stats['pass_rate']:.2%}"
        )

    # collect suspect cells (absolute Excel positions)
    suspect_cells: set[Tuple[int, int]] = set()
    for fw in audit_report.get("failed_windows", []):
        pos = fw.get("position") or {}
        start_row = pos.get("start_row")
        start_col = pos.get("start_col")
        if start_row is None or start_col is None:
            continue
        for fr in fw.get("failed_rules", []):
            sides = fr.get("equation_sides")
            if not (isinstance(sides, list) and len(sides) == 2):
                continue
            indices = _extract_indices([str(sides[0]), str(sides[1])])
            for r, c in indices:
                if r < 0 or c < 0:
                    continue
                if r >= args.window_height or c >= args.window_width:
                    continue
                abs_row = row_offset + int(start_row) + r + 1  # openpyxl is 1-based
                abs_col = col_offset + int(start_col) + c + 1
                suspect_cells.add((abs_row, abs_col))

    wb = load_workbook(excel_path)
    sheet_name = table_result.get("sheet_name")
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r, c in sorted(suspect_cells):
        ws.cell(row=r, column=c).fill = yellow

    out_path = args.out_excel
    if not out_path.is_absolute():
        out_path = Path.cwd() / out_path
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"\n=== Mark Summary ===")
    print(f"output_excel: {out_path}")
    print(f"marked_cells: {len(suspect_cells)}")
    print(f"failed_windows: {audit_report.get('summary', {}).get('failed_windows', 0)}")

    suspect_cells_list: List[Dict[str, Any]] = []
    for r, c in sorted(suspect_cells):
        suspect_cells_list.append({
            "row": r,
            "col": c,
            "address": f"{_col_to_excel(c)}{r}",
        })

    if isinstance(ret, dict):
        ret["marked_cells"] = len(suspect_cells)
        ret["failed_windows"] = audit_report.get("summary", {}).get("failed_windows", 0)
        ret["output_excel"] = str(out_path)
        ret["suspect_sheet"] = sheet_name or ws.title
        ret["suspect_cells"] = suspect_cells_list
        ret["suspect_cells_count"] = len(suspect_cells_list)


if __name__ == "__main__":
    parser = build_arg_parser()
    args = parser.parse_args()
    main(args)
