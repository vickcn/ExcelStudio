#!/usr/bin/env python3
"""ExcelStudio stdio server.

A JSON-lines based command server for Excel workbook automation built on
openpyxl. The server keeps workbook sessions in memory and exposes a stable
command envelope that can be called by agents, CLIs, or MCP-style wrappers.

Input line format:
    {"command": "create_workbook", "args": {...}, "request_id": "req-1"}

Output line format:
    {"ok": true, "command": "create_workbook", ...}

The implementation focuses on commands that are actually supported by
openpyxl and are reliable in a headless Python runtime.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
import traceback
import uuid
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from time import perf_counter
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

SERVER_VERSION = "1.0.0"
DEFAULT_ENCODING = "utf-8"


def utc_now_z() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


class XlsxStdioError(Exception):
    """Structured error for client-safe responses."""

    def __init__(self, code: str, message: str, details: Optional[Dict[str, Any]] = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass
class WorkbookSession:
    workbook_id: str
    workbook: Workbook
    path: Optional[str] = None
    created_at: str = field(default_factory=lambda: utc_now_z())
    last_saved_at: Optional[str] = None
    dirty: bool = False
    meta: Dict[str, Any] = field(default_factory=dict)
    audit_log: List[Dict[str, Any]] = field(default_factory=list)


class XlsxStdioServer:
    def __init__(self) -> None:
        self.sessions: Dict[str, WorkbookSession] = {}
        self.server_audit_log: List[Dict[str, Any]] = []
        self.command_specs: Dict[str, Dict[str, str]] = {
            "ping": {"group": "system", "summary": "Liveness probe."},
            "health_check": {"group": "system", "summary": "Runtime and dependency health."},
            "get_version": {"group": "system", "summary": "Server version metadata."},
            "list_commands": {"group": "system", "summary": "List supported commands."},
            "create_workbook": {"group": "workbook", "summary": "Create an in-memory workbook session."},
            "open_workbook": {"group": "workbook", "summary": "Open an xlsx/xlsm workbook from disk."},
            "save_workbook": {"group": "workbook", "summary": "Save workbook to its current path."},
            "save_workbook_as": {"group": "workbook", "summary": "Save workbook to a new path."},
            "close_workbook": {"group": "workbook", "summary": "Close a workbook session."},
            "list_open_workbooks": {"group": "workbook", "summary": "List workbook sessions in memory."},
            "get_workbook_info": {"group": "workbook", "summary": "Inspect workbook metadata and sheet list."},
            "list_sheets": {"group": "sheet", "summary": "List worksheet tabs."},
            "add_sheet": {"group": "sheet", "summary": "Create a worksheet."},
            "rename_sheet": {"group": "sheet", "summary": "Rename a worksheet."},
            "delete_sheet": {"group": "sheet", "summary": "Delete a worksheet."},
            "copy_sheet": {"group": "sheet", "summary": "Duplicate a worksheet."},
            "move_sheet": {"group": "sheet", "summary": "Reorder a worksheet."},
            "set_active_sheet": {"group": "sheet", "summary": "Set active worksheet."},
            "get_sheet_info": {"group": "sheet", "summary": "Inspect worksheet metadata."},
            "read_range": {"group": "range", "summary": "Read cells as matrix, records, or metadata."},
            "read_as_records": {"group": "range", "summary": "Read first row as header and return records."},
            "write_range": {"group": "range", "summary": "Write 2D values into a worksheet."},
            "clear_range": {"group": "range", "summary": "Clear values (optionally styles) in a range."},
            "append_rows": {"group": "range", "summary": "Append rows to worksheet bottom."},
            "insert_rows": {"group": "range", "summary": "Insert worksheet rows."},
            "delete_rows": {"group": "range", "summary": "Delete worksheet rows."},
            "insert_columns": {"group": "range", "summary": "Insert worksheet columns."},
            "delete_columns": {"group": "range", "summary": "Delete worksheet columns."},
            "get_used_range": {"group": "range", "summary": "Return actual populated bounds."},
            "preview_group_aggregate": {"group": "analysis", "summary": "Preview grouped aggregation and representative-row rules."},
            "group_aggregate": {"group": "analysis", "summary": "Run grouped aggregation and optionally write result to sheet."},
            "set_number_format": {"group": "format", "summary": "Apply number format to a range."},
            "set_alignment": {"group": "format", "summary": "Apply cell alignment to a range."},
            "set_font_style": {"group": "format", "summary": "Apply font styling to a range."},
            "set_fill_color": {"group": "format", "summary": "Apply fill color to a range."},
            "set_border": {"group": "format", "summary": "Apply border styling to a range."},
            "merge_cells": {"group": "format", "summary": "Merge a cell range."},
            "unmerge_cells": {"group": "format", "summary": "Unmerge a cell range."},
            "autofit_columns": {"group": "format", "summary": "Estimate column widths from content."},
            "freeze_panes": {"group": "format", "summary": "Set worksheet freeze panes."},
            "set_formula": {"group": "formula", "summary": "Set a formula on one cell or range."},
            "fill_formula": {"group": "formula", "summary": "Fill a translated formula across a range."},
            "get_formula": {"group": "formula", "summary": "Read formulas from a range."},
            "recalculate_workbook": {"group": "formula", "summary": "Flag workbook for full recalculation on open."},
            "create_table": {"group": "table", "summary": "Create an Excel structured table."},
            "list_tables": {"group": "table", "summary": "List tables in one or all sheets."},
            "set_filter": {"group": "table", "summary": "Apply worksheet autofilter to a range."},
            "clear_filter": {"group": "table", "summary": "Clear worksheet autofilter."},
            "set_data_validation": {"group": "validation", "summary": "Attach a data validation rule."},
            "set_dropdown_list": {"group": "validation", "summary": "Create list validation from values or formula."},
            "export_csv": {"group": "export", "summary": "Export a range or sheet to CSV."},
            "export_json": {"group": "export", "summary": "Export workbook or sheet data to JSON."},
            "get_audit_log": {"group": "audit", "summary": "Return in-memory audit log."},
            "write_audit_log": {"group": "audit", "summary": "Persist audit log to disk."},
        }

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def run_forever(self) -> int:
        for raw_line in sys.stdin:
            line = raw_line.strip()
            if not line:
                continue
            response = self.process_json_line(line)
            sys.stdout.write(json.dumps(response, ensure_ascii=False) + "\n")
            sys.stdout.flush()
        return 0

    def process_json_line(self, line: str) -> Dict[str, Any]:
        try:
            payload = json.loads(line)
        except json.JSONDecodeError as exc:
            return self._build_error(
                command="<parse>",
                request_id=None,
                error=XlsxStdioError("INVALID_JSON", f"Invalid JSON input: {exc}", {"line": line}),
                elapsed_ms=0,
            )
        return self.process_payload(payload)

    def process_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        started = perf_counter()
        command = payload.get("command")
        request_id = payload.get("request_id")
        args = payload.get("args", {})
        if not isinstance(args, dict):
            return self._build_error(
                command=command or "<missing>",
                request_id=request_id,
                error=XlsxStdioError("INVALID_ARGS", "'args' must be a JSON object."),
                elapsed_ms=self._elapsed_ms(started),
            )
        if not command or not isinstance(command, str):
            return self._build_error(
                command="<missing>",
                request_id=request_id,
                error=XlsxStdioError("MISSING_COMMAND", "Payload must include a string 'command'."),
                elapsed_ms=self._elapsed_ms(started),
            )
        handler = getattr(self, f"cmd_{command}", None)
        if handler is None:
            return self._build_error(
                command=command,
                request_id=request_id,
                error=XlsxStdioError(
                    "COMMAND_NOT_FOUND",
                    f"Unsupported command '{command}'.",
                    {"available_commands": sorted(self.command_specs)},
                ),
                elapsed_ms=self._elapsed_ms(started),
            )

        workbook_id = args.get("workbook_id")
        try:
            data = handler(args)
            elapsed_ms = self._elapsed_ms(started)
            response = {
                "ok": True,
                "command": command,
                "request_id": request_id,
                "data": self._json_safe(data),
                "warnings": [],
                "elapsed_ms": elapsed_ms,
            }
            if isinstance(data, dict) and "workbook_id" in data:
                response["workbook_id"] = data["workbook_id"]
            elif isinstance(workbook_id, str):
                response["workbook_id"] = workbook_id
            self._log_event(command, request_id, workbook_id, True, elapsed_ms, None, args)
            return response
        except XlsxStdioError as exc:
            elapsed_ms = self._elapsed_ms(started)
            self._log_event(command, request_id, workbook_id, False, elapsed_ms, exc, args)
            return self._build_error(command, request_id, exc, elapsed_ms)
        except Exception as exc:  # pragma: no cover - guardrail
            elapsed_ms = self._elapsed_ms(started)
            wrapped = XlsxStdioError(
                "INTERNAL_ERROR",
                str(exc),
                {"traceback": traceback.format_exc(limit=10)},
            )
            self._log_event(command, request_id, workbook_id, False, elapsed_ms, wrapped, args)
            return self._build_error(command, request_id, wrapped, elapsed_ms)

    # ------------------------------------------------------------------
    # System commands
    # ------------------------------------------------------------------
    def cmd_ping(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return {"message": "pong", "server_version": SERVER_VERSION, "utc_time": utc_now_z()}

    def cmd_health_check(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "status": "ok",
            "server_version": SERVER_VERSION,
            "python_version": sys.version.split()[0],
            "openpyxl_version": getattr(openpyxl, "__version__", None),
            "open_sessions": len(self.sessions),
            "cwd": os.getcwd(),
        }

    def cmd_get_version(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "server": "ExcelStudio xlsx_stdio",
            "version": SERVER_VERSION,
            "protocol": "jsonl-1",
            "runtime": {"python": sys.version.split()[0], "openpyxl": getattr(openpyxl, "__version__", None)},
        }

    def cmd_list_commands(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "count": len(self.command_specs),
            "commands": [
                {"name": name, "group": meta["group"], "summary": meta["summary"]}
                for name, meta in sorted(self.command_specs.items())
            ],
        }

    # ------------------------------------------------------------------
    # Workbook commands
    # ------------------------------------------------------------------
    def cmd_create_workbook(self, args: Dict[str, Any]) -> Dict[str, Any]:
        workbook = Workbook()
        workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        workbook_id = self._new_workbook_id(args.get("workbook_id"))
        default_title = args.get("default_sheet_name")
        if default_title:
            workbook.active.title = str(default_title)
        if args.get("remove_default_sheet"):
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            sheet_names = args.get("initial_sheets") or ["Sheet1"]
            for name in sheet_names:
                workbook.create_sheet(title=str(name))
        elif args.get("initial_sheets"):
            active = workbook.active
            first_name, *rest = [str(x) for x in args["initial_sheets"]]
            active.title = first_name
            for name in rest:
                workbook.create_sheet(title=name)
        session = WorkbookSession(workbook_id=workbook_id, workbook=workbook, path=args.get("path"))
        self.sessions[workbook_id] = session
        return self._session_summary(session)

    def cmd_open_workbook(self, args: Dict[str, Any]) -> Dict[str, Any]:
        path = self._required_str(args, "path")
        abs_path = str(Path(path).expanduser().resolve())
        if not Path(abs_path).exists():
            raise XlsxStdioError("FILE_NOT_FOUND", f"Workbook not found: {path}")
        workbook = load_workbook(
            filename=abs_path,
            read_only=bool(args.get("read_only", False)),
            data_only=bool(args.get("data_only", False)),
            keep_vba=bool(args.get("keep_vba", False)),
        )
        workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        workbook_id = self._new_workbook_id(args.get("workbook_id"))
        session = WorkbookSession(workbook_id=workbook_id, workbook=workbook, path=abs_path, dirty=False)
        self.sessions[workbook_id] = session
        return self._session_summary(session)

    def cmd_save_workbook(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        path = session.path or args.get("path")
        if not path:
            raise XlsxStdioError("PATH_REQUIRED", "Workbook has no bound path. Use save_workbook_as or provide 'path'.")
        return self._save_session(session, str(path))

    def cmd_save_workbook_as(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        path = self._required_str(args, "path")
        return self._save_session(session, path)

    def cmd_close_workbook(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        save_first = bool(args.get("save_before_close", False))
        if save_first:
            save_path = args.get("path") or session.path
            if not save_path:
                raise XlsxStdioError("PATH_REQUIRED", "No save path available before close.")
            self._save_session(session, str(save_path))
        workbook_id = session.workbook_id
        del self.sessions[workbook_id]
        return {"workbook_id": workbook_id, "closed": True}

    def cmd_list_open_workbooks(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return {"count": len(self.sessions), "workbooks": [self._session_summary(s) for s in self.sessions.values()]}

    def cmd_get_workbook_info(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        wb = session.workbook
        return {
            **self._session_summary(session),
            "sheet_count": len(wb.sheetnames),
            "sheet_names": list(wb.sheetnames),
            "active_sheet": wb.active.title if wb.worksheets else None,
            "defined_names": sorted([getattr(item, "name", str(item)) for item in getattr(wb.defined_names, "definedName", [])]),
        }

    # ------------------------------------------------------------------
    # Sheet commands
    # ------------------------------------------------------------------
    def cmd_list_sheets(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        wb = session.workbook
        return {
            "workbook_id": session.workbook_id,
            "active_sheet": wb.active.title if wb.worksheets else None,
            "sheets": [
                {"title": ws.title, "index": idx, "state": ws.sheet_state, "max_row": ws.max_row, "max_column": ws.max_column}
                for idx, ws in enumerate(wb.worksheets)
            ],
        }

    def cmd_add_sheet(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        title = self._required_str(args, "title")
        index = args.get("index")
        ws = session.workbook.create_sheet(title=title, index=index)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": {"title": ws.title, "index": session.workbook.index(ws)}}

    def cmd_rename_sheet(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args, key="sheet")
        new_title = self._required_str(args, "new_title")
        old_title = ws.title
        ws.title = new_title
        session.dirty = True
        return {"workbook_id": session.workbook_id, "old_title": old_title, "new_title": ws.title}

    def cmd_delete_sheet(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        wb = session.workbook
        if len(wb.sheetnames) <= 1:
            raise XlsxStdioError("LAST_SHEET", "Cannot delete the only remaining worksheet.")
        ws = self._get_worksheet(session, args, key="sheet")
        title = ws.title
        wb.remove(ws)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "deleted_sheet": title, "remaining_sheets": list(wb.sheetnames)}

    def cmd_copy_sheet(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        source = self._get_worksheet(session, args, key="source_sheet")
        copy_ws = session.workbook.copy_worksheet(source)
        if args.get("new_title"):
            copy_ws.title = str(args["new_title"])
        session.dirty = True
        return {"workbook_id": session.workbook_id, "source_sheet": source.title, "new_sheet": copy_ws.title}

    def cmd_move_sheet(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args, key="sheet")
        offset = int(args.get("offset", 0))
        session.workbook.move_sheet(ws, offset=offset)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "new_index": session.workbook.index(ws)}

    def cmd_set_active_sheet(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args, key="sheet")
        session.workbook.active = session.workbook.index(ws)
        return {"workbook_id": session.workbook_id, "active_sheet": ws.title, "active_index": session.workbook.index(ws)}

    def cmd_get_sheet_info(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args, key="sheet")
        return self._sheet_info(ws, session.workbook_id)

    # ------------------------------------------------------------------
    # Range commands
    # ------------------------------------------------------------------
    def cmd_read_range(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        mode = str(args.get("mode", "matrix"))
        cells = ws[cell_range]
        matrix = [[self._cell_value(cell) for cell in row] for row in cells]
        if mode == "matrix":
            data: Any = matrix
        elif mode == "records":
            data = self._matrix_to_records(matrix, include_empty=bool(args.get("include_empty", False)))
        elif mode == "with_meta":
            data = [[self._serialize_cell(cell) for cell in row] for row in cells]
        else:
            raise XlsxStdioError("INVALID_MODE", f"Unsupported read mode '{mode}'.", {"allowed": ["matrix", "records", "with_meta"]})
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "mode": mode, "values": data}

    def cmd_read_as_records(self, args: Dict[str, Any]) -> Dict[str, Any]:
        args = dict(args)
        args["mode"] = "records"
        return self.cmd_read_range(args)

    def cmd_write_range(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        values = args.get("values")
        if values is None or not isinstance(values, list):
            raise XlsxStdioError("INVALID_ARGS", "'values' must be a 2D JSON array.")
        matrix = self._normalize_matrix(values)
        if not matrix:
            raise XlsxStdioError("INVALID_ARGS", "'values' cannot be empty.")
        start_cell = args.get("start_cell")
        cell_range = args.get("range")
        if start_cell:
            min_col, min_row, _, _ = range_boundaries(f"{start_cell}:{start_cell}")
        elif cell_range:
            min_col, min_row, _, _ = range_boundaries(str(cell_range))
        else:
            raise XlsxStdioError("INVALID_ARGS", "Provide either 'start_cell' or 'range'.")
        written_rows = 0
        written_cols = max((len(r) for r in matrix), default=0)
        for r_idx, row in enumerate(matrix, start=min_row):
            written_rows += 1
            for c_offset, value in enumerate(row):
                cell = ws.cell(row=r_idx, column=min_col + c_offset)
                cell.value = value
        session.dirty = True
        end_col = min_col + max(written_cols - 1, 0)
        end_row = min_row + max(written_rows - 1, 0)
        written_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(end_col)}{end_row}"
        return {
            "workbook_id": session.workbook_id,
            "sheet": ws.title,
            "written_rows": written_rows,
            "written_cols": written_cols,
            "written_range": written_range,
        }

    def cmd_clear_range(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        clear_styles = bool(args.get("clear_styles", False))
        cleared = 0
        for row in ws[cell_range]:
            for cell in row:
                cell.value = None
                if clear_styles:
                    cell._style = copy(Cell(ws, row=1, column=1)._style)
                cleared += 1
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "cleared_cells": cleared, "clear_styles": clear_styles}

    def cmd_append_rows(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        rows = args.get("rows")
        if not isinstance(rows, list) or not rows:
            raise XlsxStdioError("INVALID_ARGS", "'rows' must be a non-empty 2D JSON array.")
        start_row = ws.max_row + 1 if ws.max_row or any(cell.value is not None for row in ws.iter_rows(1, 1, 1, 1) for cell in row) else 1
        for row in rows:
            if not isinstance(row, list):
                raise XlsxStdioError("INVALID_ARGS", "Each item in 'rows' must be a JSON array.")
            ws.append(row)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "start_row": start_row, "appended_rows": len(rows)}

    def cmd_insert_rows(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        row = int(args.get("row", 0))
        amount = int(args.get("amount", 1))
        if row <= 0 or amount <= 0:
            raise XlsxStdioError("INVALID_ARGS", "'row' and 'amount' must be positive integers.")
        ws.insert_rows(row, amount)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "row": row, "amount": amount}

    def cmd_delete_rows(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        row = int(args.get("row", 0))
        amount = int(args.get("amount", 1))
        if row <= 0 or amount <= 0:
            raise XlsxStdioError("INVALID_ARGS", "'row' and 'amount' must be positive integers.")
        ws.delete_rows(row, amount)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "row": row, "amount": amount}

    def cmd_insert_columns(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        column = self._coerce_column_index(args.get("column"))
        amount = int(args.get("amount", 1))
        if column <= 0 or amount <= 0:
            raise XlsxStdioError("INVALID_ARGS", "'column' and 'amount' must be positive.")
        ws.insert_cols(column, amount)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "column": column, "column_letter": get_column_letter(column), "amount": amount}

    def cmd_delete_columns(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        column = self._coerce_column_index(args.get("column"))
        amount = int(args.get("amount", 1))
        if column <= 0 or amount <= 0:
            raise XlsxStdioError("INVALID_ARGS", "'column' and 'amount' must be positive.")
        ws.delete_cols(column, amount)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "column": column, "column_letter": get_column_letter(column), "amount": amount}

    def cmd_get_used_range(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        used = self._actual_used_range(ws)
        return {"workbook_id": session.workbook_id, "sheet": ws.title, **used}

    # ------------------------------------------------------------------
    # Analysis commands
    # ------------------------------------------------------------------
    def cmd_preview_group_aggregate(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args, default_to_used=True)

        group_cols = self._parse_str_list_arg(args, "group_cols", required=True)
        time_priority = self._parse_str_list_arg(args, "time_priority", default=["Cdate", "檢驗日"])
        value_col = str(args.get("value_col", "PGA")).strip() or "PGA"
        include_empty_rows = bool(args.get("include_empty_rows", False))
        preview_limit = int(args.get("preview_limit", 10))
        if preview_limit < 0:
            raise XlsxStdioError("INVALID_ARGS", "'preview_limit' must be >= 0.")

        dataset = self._read_records_dataset(ws, cell_range, include_empty_rows=include_empty_rows)
        aggregate = self._aggregate_group_records(
            records=dataset["records"],
            headers=dataset["headers"],
            group_cols=group_cols,
            value_col=value_col,
            time_priority=time_priority,
        )

        preview_records = aggregate["records"][:preview_limit] if preview_limit else []
        return {
            "workbook_id": session.workbook_id,
            "sheet": ws.title,
            "range": cell_range,
            "source_rows": dataset["source_rows"],
            "records_rows": dataset["records_rows"],
            "group_cols": group_cols,
            "value_col": value_col,
            "time_priority": time_priority,
            "detected_time_cols": aggregate["detected_time_cols"],
            "used_time_mode": aggregate["used_time_mode"],
            "carry_columns": aggregate["carry_columns"],
            "warnings": aggregate["warnings"],
            "aggregated_rows": len(aggregate["records"]),
            "ordered_columns": aggregate["ordered_columns"],
            "preview_count": len(preview_records),
            "preview_records": preview_records,
        }

    def cmd_group_aggregate(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args, default_to_used=True)

        group_cols = self._parse_str_list_arg(args, "group_cols", required=True)
        time_priority = self._parse_str_list_arg(args, "time_priority", default=["Cdate", "檢驗日"])
        value_col = str(args.get("value_col", "PGA")).strip() or "PGA"
        include_empty_rows = bool(args.get("include_empty_rows", False))
        preview_limit = int(args.get("preview_limit", 10))
        if preview_limit < 0:
            raise XlsxStdioError("INVALID_ARGS", "'preview_limit' must be >= 0.")

        dataset = self._read_records_dataset(ws, cell_range, include_empty_rows=include_empty_rows)
        aggregate = self._aggregate_group_records(
            records=dataset["records"],
            headers=dataset["headers"],
            group_cols=group_cols,
            value_col=value_col,
            time_priority=time_priority,
        )

        write_sheet = args.get("write_sheet") or args.get("target_sheet")
        write_result: Optional[Dict[str, Any]] = None
        if write_sheet is not None:
            target_ws = self._get_or_create_sheet(
                session.workbook,
                str(write_sheet),
                replace_sheet=bool(args.get("replace_sheet", False)),
            )
            if bool(args.get("clear_sheet", True)):
                target_ws.delete_rows(1, target_ws.max_row)
            matrix = self._records_to_matrix(aggregate["records"], aggregate["ordered_columns"])
            write_result = self._write_matrix_to_sheet(
                ws=target_ws,
                start_cell=str(args.get("start_cell", "A1")),
                matrix=matrix,
            )
            session.dirty = True

        return_records = bool(args.get("return_records", False))
        response: Dict[str, Any] = {
            "workbook_id": session.workbook_id,
            "sheet": ws.title,
            "range": cell_range,
            "source_rows": dataset["source_rows"],
            "records_rows": dataset["records_rows"],
            "group_cols": group_cols,
            "value_col": value_col,
            "time_priority": time_priority,
            "detected_time_cols": aggregate["detected_time_cols"],
            "used_time_mode": aggregate["used_time_mode"],
            "carry_columns": aggregate["carry_columns"],
            "warnings": aggregate["warnings"],
            "aggregated_rows": len(aggregate["records"]),
            "ordered_columns": aggregate["ordered_columns"],
        }
        if write_result is not None:
            response["write_result"] = write_result
        if return_records:
            response["records"] = aggregate["records"]
        else:
            response["preview_records"] = aggregate["records"][:preview_limit] if preview_limit else []
            response["preview_count"] = len(response["preview_records"])
        return response

    # ------------------------------------------------------------------
    # Formatting commands
    # ------------------------------------------------------------------
    def cmd_set_number_format(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        fmt = self._required_str(args, "format_code")
        cell_range = self._resolve_range(ws, args)
        count = 0
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = fmt
                count += 1
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "format_code": fmt, "updated_cells": count}

    def cmd_set_alignment(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        count = 0
        for row in ws[cell_range]:
            for cell in row:
                cell.alignment = copy(cell.alignment) if cell.alignment else Alignment()
                cell.alignment = Alignment(
                    horizontal=args.get("horizontal", cell.alignment.horizontal if cell.alignment else None),
                    vertical=args.get("vertical", cell.alignment.vertical if cell.alignment else None),
                    wrap_text=args.get("wrap_text", cell.alignment.wrap_text if cell.alignment else None),
                    shrink_to_fit=args.get("shrink_to_fit", cell.alignment.shrink_to_fit if cell.alignment else None),
                    text_rotation=args.get("text_rotation", cell.alignment.textRotation if cell.alignment else 0),
                )
                count += 1
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "updated_cells": count}

    def cmd_set_font_style(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        count = 0
        for row in ws[cell_range]:
            for cell in row:
                existing = cell.font or Font()
                cell.font = Font(
                    name=args.get("name", existing.name),
                    size=args.get("size", existing.size),
                    bold=args.get("bold", existing.bold),
                    italic=args.get("italic", existing.italic),
                    underline=args.get("underline", existing.underline),
                    color=self._normalize_color(args.get("color", existing.color.rgb if getattr(existing.color, "rgb", None) else None)),
                )
                count += 1
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "updated_cells": count}

    def cmd_set_fill_color(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        color = self._normalize_color(self._required_str(args, "color"))
        fill_type = str(args.get("fill_type", "solid"))
        cell_range = self._resolve_range(ws, args)
        count = 0
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = PatternFill(fill_type=fill_type, start_color=color, end_color=color)
                count += 1
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "color": color, "updated_cells": count}

    def cmd_set_border(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        style = str(args.get("style", "thin"))
        color = self._normalize_color(args.get("color", "D9D9D9"))
        edges = args.get("edges") or ["left", "right", "top", "bottom"]
        if not isinstance(edges, list):
            raise XlsxStdioError("INVALID_ARGS", "'edges' must be an array of border sides.")
        side = Side(style=style, color=color)
        cell_range = self._resolve_range(ws, args)
        count = 0
        for row in ws[cell_range]:
            for cell in row:
                current = cell.border or Border()
                kwargs = {
                    "left": current.left,
                    "right": current.right,
                    "top": current.top,
                    "bottom": current.bottom,
                    "diagonal": current.diagonal,
                    "diagonalUp": current.diagonalUp,
                    "diagonalDown": current.diagonalDown,
                    "outline": current.outline,
                    "vertical": current.vertical,
                    "horizontal": current.horizontal,
                }
                for edge in edges:
                    if edge not in {"left", "right", "top", "bottom"}:
                        raise XlsxStdioError("INVALID_ARGS", f"Unsupported border edge '{edge}'.")
                    kwargs[edge] = side
                cell.border = Border(**kwargs)
                count += 1
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "edges": edges, "updated_cells": count}

    def cmd_merge_cells(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        ws.merge_cells(cell_range)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "merged_range": cell_range}

    def cmd_unmerge_cells(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        ws.unmerge_cells(cell_range)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "unmerged_range": cell_range}

    def cmd_autofit_columns(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        min_width = float(args.get("min_width", 8))
        max_width = float(args.get("max_width", 60))
        columns_arg = args.get("columns")
        target_columns: List[int]
        if columns_arg:
            if not isinstance(columns_arg, list):
                raise XlsxStdioError("INVALID_ARGS", "'columns' must be an array of letters or indices.")
            target_columns = [self._coerce_column_index(col) for col in columns_arg]
        else:
            used = self._actual_used_range(ws)
            if not used["has_data"]:
                target_columns = [1]
            else:
                target_columns = list(range(used["min_col"], used["max_col"] + 1))
        widths: Dict[str, float] = {}
        for col_idx in target_columns:
            max_len = 0
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if value is None:
                    continue
                display = self._display_string(value)
                max_len = max(max_len, len(display))
            width = min(max((max_len * 1.12) + 2, min_width), max_width)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            widths[col_letter] = round(width, 2)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "column_widths": widths}

    def cmd_freeze_panes(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        pane = args.get("cell")
        ws.freeze_panes = pane if pane else None
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "freeze_panes": ws.freeze_panes}

    # ------------------------------------------------------------------
    # Formula commands
    # ------------------------------------------------------------------
    def cmd_set_formula(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        if "cell" in args and "formula" in args:
            cell = ws[self._required_str(args, "cell")]
            formula = str(args["formula"])
            cell.value = formula
            updated = 1
            target = cell.coordinate
        elif "range" in args and "formulas" in args:
            formulas = self._normalize_matrix(args["formulas"])
            cell_range = self._resolve_range(ws, args)
            min_col, min_row, _, _ = range_boundaries(cell_range)
            updated = 0
            for r_idx, row in enumerate(formulas, start=min_row):
                for c_offset, formula in enumerate(row):
                    ws.cell(row=r_idx, column=min_col + c_offset).value = formula
                    updated += 1
            target = cell_range
        else:
            raise XlsxStdioError("INVALID_ARGS", "Provide ('cell' + 'formula') or ('range' + 'formulas').")
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "target": target, "updated_cells": updated}

    def cmd_fill_formula(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        source_cell_ref = self._required_str(args, "source_cell")
        target_range = self._required_str(args, "target_range")
        source_formula = ws[source_cell_ref].value
        if not isinstance(source_formula, str) or not source_formula.startswith("="):
            raise XlsxStdioError("NO_FORMULA", f"Source cell {source_cell_ref} does not contain a formula.")
        updated = 0
        for row in ws[target_range]:
            for cell in row:
                translated = Translator(source_formula, origin=source_cell_ref).translate_formula(cell.coordinate)
                cell.value = translated
                updated += 1
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "source_cell": source_cell_ref, "target_range": target_range, "updated_cells": updated}

    def cmd_get_formula(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        formulas: Dict[str, str] = {}
        for row in ws[cell_range]:
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[cell.coordinate] = cell.value
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "formulas": formulas, "count": len(formulas)}

    def cmd_recalculate_workbook(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        session.workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        session.dirty = True
        saved = False
        if args.get("save"):
            path = args.get("path") or session.path
            if not path:
                raise XlsxStdioError("PATH_REQUIRED", "No path available to save recalculation flags.")
            self._save_session(session, str(path))
            saved = True
        return {
            "workbook_id": session.workbook_id,
            "saved": saved,
            "note": "Workbook flagged for full recalculation by Excel-compatible engines on next open. openpyxl does not evaluate formulas itself.",
        }

    # ------------------------------------------------------------------
    # Table / filter commands
    # ------------------------------------------------------------------
    def cmd_create_table(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        table_name = self._required_str(args, "name")
        if table_name in ws.tables:
            raise XlsxStdioError("TABLE_EXISTS", f"Table '{table_name}' already exists in sheet '{ws.title}'.")
        tab = Table(displayName=table_name, ref=cell_range)
        style_name = str(args.get("style_name", "TableStyleMedium9"))
        tab.tableStyleInfo = TableStyleInfo(
            name=style_name,
            showFirstColumn=bool(args.get("show_first_column", False)),
            showLastColumn=bool(args.get("show_last_column", False)),
            showRowStripes=bool(args.get("show_row_stripes", True)),
            showColumnStripes=bool(args.get("show_column_stripes", False)),
        )
        ws.add_table(tab)
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "table": {"name": table_name, "range": cell_range, "style_name": style_name}}

    def cmd_list_tables(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        if args.get("sheet"):
            ws = self._get_worksheet(session, args)
            sheets = [ws]
        else:
            sheets = session.workbook.worksheets
        tables = []
        for ws in sheets:
            for name in ws.tables.keys():
                table = ws.tables[name]
                tables.append({"sheet": ws.title, "name": name, "range": getattr(table, "ref", ws.tables[name]), "style": getattr(getattr(table, "tableStyleInfo", None), "name", None)})
        return {"workbook_id": session.workbook_id, "tables": tables, "count": len(tables)}

    def cmd_set_filter(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        ws.auto_filter.ref = cell_range
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "filter_range": cell_range}

    def cmd_clear_filter(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        ws.auto_filter.ref = None
        session.dirty = True
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "filter_cleared": True}

    # ------------------------------------------------------------------
    # Data validation commands
    # ------------------------------------------------------------------
    def cmd_set_data_validation(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        cell_range = self._resolve_range(ws, args)
        validation_type = self._required_str(args, "validation_type")
        formula1 = args.get("formula1")
        formula2 = args.get("formula2")
        if formula1 is None and validation_type != "any":
            raise XlsxStdioError("INVALID_ARGS", "'formula1' is required for this validation type.")
        dv = DataValidation(
            type=validation_type,
            formula1=formula1,
            formula2=formula2,
            allow_blank=bool(args.get("allow_blank", True)),
            operator=args.get("operator"),
            showErrorMessage=bool(args.get("show_error_message", True)),
            errorTitle=args.get("error_title"),
            error=args.get("error_message"),
            promptTitle=args.get("prompt_title"),
            prompt=args.get("prompt_message"),
            showDropDown=bool(args.get("show_dropdown", False)),
        )
        ws.add_data_validation(dv)
        dv.add(cell_range)
        session.dirty = True
        return {
            "workbook_id": session.workbook_id,
            "sheet": ws.title,
            "range": cell_range,
            "validation_type": validation_type,
            "formula1": formula1,
            "formula2": formula2,
        }

    def cmd_set_dropdown_list(self, args: Dict[str, Any]) -> Dict[str, Any]:
        values = args.get("values")
        formula = args.get("formula")
        if values is None and formula is None:
            raise XlsxStdioError("INVALID_ARGS", "Provide either 'values' or 'formula' for dropdown list.")
        if values is not None:
            if not isinstance(values, list) or not values:
                raise XlsxStdioError("INVALID_ARGS", "'values' must be a non-empty array.")
            formula1 = '"' + ",".join(str(v) for v in values) + '"'
        else:
            formula1 = formula
        args = dict(args)
        args["validation_type"] = "list"
        args["formula1"] = formula1
        return self.cmd_set_data_validation(args)

    # ------------------------------------------------------------------
    # Export commands
    # ------------------------------------------------------------------
    def cmd_export_csv(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        ws = self._get_worksheet(session, args)
        path = self._required_str(args, "path")
        cell_range = self._resolve_range(ws, args, default_to_used=True)
        matrix = [[self._cell_value(cell) for cell in row] for row in ws[cell_range]]
        abs_path = self._ensure_parent_dir(path)
        with open(abs_path, "w", encoding=DEFAULT_ENCODING, newline="") as f:
            writer = csv.writer(f)
            for row in matrix:
                writer.writerow(["" if v is None else v for v in row])
        return {"workbook_id": session.workbook_id, "sheet": ws.title, "range": cell_range, "path": abs_path, "rows": len(matrix)}

    def cmd_export_json(self, args: Dict[str, Any]) -> Dict[str, Any]:
        session = self._get_session(args)
        path = self._required_str(args, "path")
        mode = str(args.get("mode", "matrix"))
        if args.get("sheet"):
            ws = self._get_worksheet(session, args)
            cell_range = self._resolve_range(ws, args, default_to_used=True)
            matrix = [[self._cell_value(cell) for cell in row] for row in ws[cell_range]]
            payload: Any
            if mode == "matrix":
                payload = {"sheet": ws.title, "range": cell_range, "values": matrix}
            elif mode == "records":
                payload = {"sheet": ws.title, "range": cell_range, "records": self._matrix_to_records(matrix, include_empty=bool(args.get("include_empty", False)))}
            else:
                raise XlsxStdioError("INVALID_MODE", f"Unsupported export mode '{mode}'.", {"allowed": ["matrix", "records"]})
        else:
            payload = {
                "workbook_id": session.workbook_id,
                "sheets": [
                    {
                        "sheet": ws.title,
                        "used_range": self._actual_used_range(ws),
                        "values": [[self._cell_value(cell) for cell in row] for row in ws[self._actual_used_range(ws)["range"]]] if self._actual_used_range(ws)["has_data"] else [],
                    }
                    for ws in session.workbook.worksheets
                ],
            }
        abs_path = self._ensure_parent_dir(path)
        with open(abs_path, "w", encoding=DEFAULT_ENCODING) as f:
            json.dump(self._json_safe(payload), f, ensure_ascii=False, indent=2)
        return {"workbook_id": session.workbook_id, "path": abs_path, "mode": mode}

    # ------------------------------------------------------------------
    # Audit commands
    # ------------------------------------------------------------------
    def cmd_get_audit_log(self, args: Dict[str, Any]) -> Dict[str, Any]:
        workbook_id = args.get("workbook_id")
        if workbook_id:
            session = self._get_session(args)
            log = session.audit_log
        else:
            log = self.server_audit_log
        limit = int(args.get("limit", len(log)))
        return {"count": min(len(log), limit), "events": log[-limit:]}

    def cmd_write_audit_log(self, args: Dict[str, Any]) -> Dict[str, Any]:
        workbook_id = args.get("workbook_id")
        if workbook_id:
            session = self._get_session(args)
            payload = {"workbook_id": session.workbook_id, "events": session.audit_log}
        else:
            payload = {"events": self.server_audit_log}
        path = self._required_str(args, "path")
        abs_path = self._ensure_parent_dir(path)
        with open(abs_path, "w", encoding=DEFAULT_ENCODING) as f:
            json.dump(self._json_safe(payload), f, ensure_ascii=False, indent=2)
        return {"path": abs_path, "event_count": len(payload["events"]), "scope": workbook_id or "server"}

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _elapsed_ms(self, started: float) -> int:
        return int(round((perf_counter() - started) * 1000))

    def _build_error(self, command: str, request_id: Optional[str], error: XlsxStdioError, elapsed_ms: int) -> Dict[str, Any]:
        return {
            "ok": False,
            "command": command,
            "request_id": request_id,
            "error": {"code": error.code, "message": error.message, "details": self._json_safe(error.details)},
            "elapsed_ms": elapsed_ms,
        }

    def _parse_str_list_arg(
        self,
        args: Dict[str, Any],
        key: str,
        default: Optional[Sequence[str]] = None,
        required: bool = False,
    ) -> List[str]:
        raw = args.get(key)
        values: List[str]
        if raw is None:
            values = [str(x).strip() for x in (default or []) if str(x).strip()]
        elif isinstance(raw, list):
            values = [str(x).strip() for x in raw if str(x).strip()]
        elif isinstance(raw, str):
            text = raw.strip()
            if not text:
                values = []
            elif "," in text:
                values = [part.strip() for part in text.split(",") if part.strip()]
            else:
                values = [text]
        else:
            raise XlsxStdioError("INVALID_ARGS", f"'{key}' must be a string list or comma-separated string.")

        if required and not values:
            raise XlsxStdioError("INVALID_ARGS", f"'{key}' must provide at least one column.")

        deduped: List[str] = []
        seen = set()
        for value in values:
            if value not in seen:
                deduped.append(value)
                seen.add(value)
        return deduped

    def _sanitize_headers_unique(self, header_row: Sequence[Any]) -> List[str]:
        headers: List[str] = []
        counts: Dict[str, int] = {}
        for idx, value in enumerate(header_row, start=1):
            base = self._sanitize_header(value, idx)
            count = counts.get(base, 0) + 1
            counts[base] = count
            headers.append(base if count == 1 else f"{base}_{count}")
        return headers

    def _read_records_dataset(self, ws, cell_range: str, include_empty_rows: bool = False) -> Dict[str, Any]:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if max_row < min_row or max_col < min_col:
            raise XlsxStdioError("INVALID_RANGE", f"Invalid range: {cell_range}")
        matrix: List[List[Any]] = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            matrix.append([self._cell_value(cell) for cell in row])
        if not matrix:
            raise XlsxStdioError("EMPTY_RANGE", f"Range '{cell_range}' has no rows.")

        headers = self._sanitize_headers_unique(matrix[0])
        records: List[Dict[str, Any]] = []
        for row in matrix[1:]:
            record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            if include_empty_rows or any(v is not None and v != "" for v in record.values()):
                records.append(record)
        return {
            "headers": headers,
            "records": records,
            "source_rows": len(matrix),
            "records_rows": len(records),
        }

    def _detect_time_like_columns_from_headers(self, headers: Sequence[str], time_priority: Sequence[str]) -> List[str]:
        out: List[str] = []
        seen = set()
        for col in time_priority:
            if col in headers and col not in seen:
                out.append(col)
                seen.add(col)
        for col in headers:
            lower = col.lower()
            if col in seen:
                continue
            if (
                "time" in lower
                or "date" in lower
                or "時間" in col
                or "日期" in col
                or col in {"檢驗日", "Cdate", "CDATE"}
            ):
                out.append(col)
                seen.add(col)
        return out

    def _to_datetime_safe(self, value: Any) -> Optional[datetime]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, time.min)
        if isinstance(value, Decimal):
            value = float(value)
        if isinstance(value, (int, float)):
            return None
        text = str(value).strip()
        if not text:
            return None

        candidates = [text]
        if text.endswith("Z"):
            candidates.append(text[:-1] + "+00:00")
        if "/" in text:
            candidates.append(text.replace("/", "-"))

        for candidate in candidates:
            try:
                return datetime.fromisoformat(candidate)
            except Exception:
                pass

        known_formats = [
            "%Y/%m/%d %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d",
            "%Y-%m-%d",
            "%Y%m%d",
        ]
        for fmt in known_formats:
            try:
                return datetime.strptime(text, fmt)
            except Exception:
                pass
        return None

    def _to_float_safe(self, value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        if isinstance(value, Decimal):
            value = float(value)
        try:
            number = float(value)
        except Exception:
            return None
        if not math.isfinite(number):
            return None
        return number

    def _order_aggregate_columns(
        self,
        headers: Sequence[str],
        group_cols: Sequence[str],
        detected_time_cols: Sequence[str],
        value_col: str,
        mean_col: str,
        std_col: str,
    ) -> List[str]:
        ordered: List[str] = []

        def add(col: str) -> None:
            if col and col not in ordered:
                ordered.append(col)

        for col in group_cols:
            add(col)
        for col in detected_time_cols:
            add(col)
        for col in [value_col, mean_col, std_col, "row_count"]:
            add(col)
        for col in headers:
            add(col)
        for col in [mean_col, std_col, "row_count"]:
            add(col)
        return ordered

    def _aggregate_group_records(
        self,
        records: Sequence[Dict[str, Any]],
        headers: Sequence[str],
        group_cols: Sequence[str],
        value_col: str,
        time_priority: Sequence[str],
    ) -> Dict[str, Any]:
        header_set = set(headers)
        missing = [col for col in list(group_cols) + [value_col] if col not in header_set]
        if missing:
            raise XlsxStdioError("MISSING_COLUMNS", "Required columns missing in selected range.", {"missing_columns": missing})

        detected_time_cols = self._detect_time_like_columns_from_headers(headers, time_priority)
        carry_columns = [c for c in headers if c not in set(group_cols) and c not in set(detected_time_cols) and c != value_col]

        row_items: List[Dict[str, Any]] = []
        has_parseable_time = False
        for row_order, record in enumerate(records):
            parsed_times = [self._to_datetime_safe(record.get(col)) for col in detected_time_cols]
            parsed_times = [dt for dt in parsed_times if dt is not None]
            earliest_time = min(parsed_times) if parsed_times else None
            if earliest_time is not None:
                has_parseable_time = True
            row_items.append({"record": record, "row_order": row_order, "earliest_time": earliest_time})

        use_time_rule = bool(detected_time_cols) and has_parseable_time
        used_time_mode = "earliest_of_all_time_cols" if use_time_rule else "first_row_fallback"
        warnings: List[str] = []
        if detected_time_cols and not has_parseable_time:
            warnings.append("Detected time-like columns but none are parseable; fallback to first row per group.")
        if not detected_time_cols:
            warnings.append("No time-like columns detected; fallback to first row per group.")
        if carry_columns:
            if use_time_rule:
                warnings.append("Non-key/non-time columns are taken from each group's earliest-time representative row.")
            else:
                warnings.append("Non-key/non-time columns are taken from each group's first row.")

        grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = {}
        for item in row_items:
            key = tuple(item["record"].get(col) for col in group_cols)
            grouped.setdefault(key, []).append(item)

        mean_col = f"{value_col}_mean"
        std_col = f"{value_col}_std"
        aggregated: List[Dict[str, Any]] = []
        for rows in grouped.values():
            if use_time_rule:
                representative = min(
                    rows,
                    key=lambda item: (
                        item["earliest_time"] is None,
                        item["earliest_time"] or datetime.max,
                        item["row_order"],
                    ),
                )
            else:
                representative = rows[0]

            out = dict(representative["record"])
            values = [self._to_float_safe(item["record"].get(value_col)) for item in rows]
            values = [v for v in values if v is not None]
            mean_value = (sum(values) / len(values)) if values else None
            if len(values) >= 2 and mean_value is not None:
                variance = sum((v - mean_value) ** 2 for v in values) / (len(values) - 1)
                std_value = math.sqrt(variance)
            else:
                std_value = None

            out[mean_col] = mean_value
            out[std_col] = std_value
            out["row_count"] = len(rows)
            aggregated.append(out)

        ordered_columns = self._order_aggregate_columns(
            headers=headers,
            group_cols=group_cols,
            detected_time_cols=detected_time_cols,
            value_col=value_col,
            mean_col=mean_col,
            std_col=std_col,
        )
        normalized_records = [{col: row.get(col) for col in ordered_columns} for row in aggregated]
        return {
            "records": normalized_records,
            "ordered_columns": ordered_columns,
            "detected_time_cols": detected_time_cols,
            "used_time_mode": used_time_mode,
            "carry_columns": carry_columns,
            "warnings": warnings,
        }

    def _records_to_matrix(self, records: Sequence[Dict[str, Any]], columns: Sequence[str]) -> List[List[Any]]:
        matrix: List[List[Any]] = [list(columns)]
        for record in records:
            matrix.append([record.get(col) for col in columns])
        return matrix

    def _get_or_create_sheet(self, workbook: Workbook, sheet_name: str, replace_sheet: bool = False):
        if sheet_name in workbook.sheetnames:
            if replace_sheet:
                old_ws = workbook[sheet_name]
                index = workbook.index(old_ws)
                workbook.remove(old_ws)
                return workbook.create_sheet(title=sheet_name, index=index)
            return workbook[sheet_name]
        return workbook.create_sheet(title=sheet_name)

    def _write_matrix_to_sheet(self, ws, start_cell: str, matrix: Sequence[Sequence[Any]]) -> Dict[str, Any]:
        if not matrix:
            raise XlsxStdioError("INVALID_ARGS", "Matrix is empty; nothing to write.")
        min_col, min_row, _, _ = range_boundaries(f"{start_cell}:{start_cell}")
        written_rows = 0
        written_cols = 0
        for r_idx, row in enumerate(matrix, start=min_row):
            written_rows += 1
            row_values = list(row)
            written_cols = max(written_cols, len(row_values))
            for c_offset, value in enumerate(row_values):
                ws.cell(row=r_idx, column=min_col + c_offset).value = value
        end_col = min_col + max(written_cols - 1, 0)
        end_row = min_row + max(written_rows - 1, 0)
        written_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(end_col)}{end_row}"
        return {
            "sheet": ws.title,
            "start_cell": start_cell,
            "written_rows": written_rows,
            "written_cols": written_cols,
            "written_range": written_range,
        }

    def _required_str(self, args: Dict[str, Any], key: str) -> str:
        value = args.get(key)
        if value is None or str(value).strip() == "":
            raise XlsxStdioError("INVALID_ARGS", f"'{key}' is required.")
        return str(value)

    def _new_workbook_id(self, requested: Optional[str] = None) -> str:
        if requested:
            if requested in self.sessions:
                raise XlsxStdioError("WORKBOOK_EXISTS", f"Workbook id '{requested}' already exists.")
            return str(requested)
        while True:
            candidate = f"wb_{uuid.uuid4().hex[:8]}"
            if candidate not in self.sessions:
                return candidate

    def _get_session(self, args: Dict[str, Any]) -> WorkbookSession:
        workbook_id = args.get("workbook_id")
        if not workbook_id:
            raise XlsxStdioError("WORKBOOK_REQUIRED", "'workbook_id' is required for this command.")
        session = self.sessions.get(str(workbook_id))
        if not session:
            raise XlsxStdioError("WORKBOOK_NOT_FOUND", f"Workbook session '{workbook_id}' does not exist.")
        return session

    def _get_worksheet(self, session: WorkbookSession, args: Dict[str, Any], key: str = "sheet"):
        wb = session.workbook
        sheet_ref = args.get(key)
        if sheet_ref is None:
            if not wb.worksheets:
                raise XlsxStdioError("SHEET_NOT_FOUND", "Workbook has no worksheets.")
            return wb.active
        if isinstance(sheet_ref, int):
            try:
                return wb.worksheets[sheet_ref]
            except IndexError as exc:
                raise XlsxStdioError("SHEET_NOT_FOUND", f"Worksheet index {sheet_ref} is out of range.") from exc
        if str(sheet_ref) not in wb.sheetnames:
            raise XlsxStdioError("SHEET_NOT_FOUND", f"Sheet '{sheet_ref}' does not exist.")
        return wb[str(sheet_ref)]

    def _resolve_range(self, ws, args: Dict[str, Any], default_to_used: bool = False) -> str:
        if args.get("range"):
            return str(args["range"])
        if args.get("start_cell") and args.get("end_cell"):
            return f"{args['start_cell']}:{args['end_cell']}"
        if default_to_used:
            used = self._actual_used_range(ws)
            return used["range"]
        raise XlsxStdioError("RANGE_REQUIRED", "Provide 'range' or ('start_cell' + 'end_cell').")

    def _coerce_column_index(self, value: Any) -> int:
        if isinstance(value, int):
            return value
        if isinstance(value, str):
            value = value.strip()
            if value.isdigit():
                return int(value)
            return column_index_from_string(value)
        raise XlsxStdioError("INVALID_ARGS", f"Invalid column reference: {value!r}")

    def _actual_used_range(self, ws) -> Dict[str, Any]:
        min_row = None
        max_row = None
        min_col = None
        max_col = None
        # Use worksheet bounds as an upper bound, then scan actual cell values.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    min_row = cell.row if min_row is None else min(min_row, cell.row)
                    max_row = cell.row if max_row is None else max(max_row, cell.row)
                    min_col = cell.column if min_col is None else min(min_col, cell.column)
                    max_col = cell.column if max_col is None else max(max_col, cell.column)
        if min_row is None:
            return {
                "has_data": False,
                "range": "A1:A1",
                "min_row": 1,
                "max_row": 1,
                "min_col": 1,
                "max_col": 1,
            }
        return {
            "has_data": True,
            "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
        }

    def _normalize_matrix(self, values: List[Any]) -> List[List[Any]]:
        matrix: List[List[Any]] = []
        for row in values:
            if isinstance(row, list):
                matrix.append(row)
            else:
                matrix.append([row])
        return matrix

    def _matrix_to_records(self, matrix: List[List[Any]], include_empty: bool = False) -> List[Dict[str, Any]]:
        if not matrix:
            return []
        headers = [self._sanitize_header(v, idx) for idx, v in enumerate(matrix[0], start=1)]
        records: List[Dict[str, Any]] = []
        for row in matrix[1:]:
            record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            if include_empty or any(v is not None and v != "" for v in record.values()):
                records.append(record)
        return records

    def _sanitize_header(self, value: Any, idx: int) -> str:
        text = str(value).strip() if value is not None else ""
        return text or f"column_{idx}"

    def _session_summary(self, session: WorkbookSession) -> Dict[str, Any]:
        wb = session.workbook
        return {
            "workbook_id": session.workbook_id,
            "path": session.path,
            "created_at": session.created_at,
            "last_saved_at": session.last_saved_at,
            "dirty": session.dirty,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": list(wb.sheetnames),
            "active_sheet": wb.active.title if wb.worksheets else None,
        }

    def _sheet_info(self, ws, workbook_id: str) -> Dict[str, Any]:
        return {
            "workbook_id": workbook_id,
            "sheet": ws.title,
            "index": ws.parent.index(ws),
            "state": ws.sheet_state,
            "dimension": ws.calculate_dimension(),
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "freeze_panes": ws.freeze_panes,
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "tables": [{"name": name, "range": table.ref} for name, table in ws.tables.items()],
        }

    def _save_session(self, session: WorkbookSession, path: str) -> Dict[str, Any]:
        abs_path = self._ensure_parent_dir(path)
        session.workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
        session.workbook.save(abs_path)
        session.path = abs_path
        session.last_saved_at = utc_now_z()
        session.dirty = False
        return {**self._session_summary(session), "saved": True}

    def _ensure_parent_dir(self, path: str) -> str:
        abs_path = str(Path(path).expanduser().resolve())
        Path(abs_path).parent.mkdir(parents=True, exist_ok=True)
        return abs_path

    def _display_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date, time)):
            return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
        return str(value)

    def _cell_value(self, cell: Cell) -> Any:
        value = cell.value
        if isinstance(value, Decimal):
            return float(value)
        return self._json_safe(value)

    def _serialize_cell(self, cell: Cell) -> Dict[str, Any]:
        return {
            "address": cell.coordinate,
            "value": self._cell_value(cell),
            "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
            "data_type": cell.data_type,
            "number_format": cell.number_format,
            "is_merged": any(cell.coordinate in rng for rng in cell.parent.merged_cells.ranges),
        }

    def _normalize_color(self, value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        color = str(value).strip().lstrip("#").upper()
        if len(color) == 6:
            return color
        if len(color) == 8:
            return color
        raise XlsxStdioError("INVALID_COLOR", f"Invalid color '{value}'. Use RGB hex like '4472C4'.")

    def _json_safe(self, value: Any) -> Any:
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (datetime, date, time)):
            return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
        if isinstance(value, Path):
            return str(value)
        if isinstance(value, dict):
            return {str(k): self._json_safe(v) for k, v in value.items()}
        if isinstance(value, (list, tuple, set)):
            return [self._json_safe(v) for v in value]
        return str(value)

    def _log_event(
        self,
        command: str,
        request_id: Optional[str],
        workbook_id: Optional[str],
        ok: bool,
        elapsed_ms: int,
        error: Optional[XlsxStdioError],
        args: Dict[str, Any],
    ) -> None:
        event = {
            "ts": utc_now_z(),
            "command": command,
            "request_id": request_id,
            "workbook_id": workbook_id,
            "ok": ok,
            "elapsed_ms": elapsed_ms,
            "error_code": error.code if error else None,
            "error_message": error.message if error else None,
            "args": self._json_safe(args),
        }
        self.server_audit_log.append(event)
        if workbook_id and workbook_id in self.sessions:
            self.sessions[workbook_id].audit_log.append(event)


# ----------------------------------------------------------------------
# CLI helpers
# ----------------------------------------------------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="ExcelStudio xlsx stdio server")
    parser.add_argument("--self-test", action="store_true", help="Run a minimal self-test and exit.")
    parser.add_argument("--demo", action="store_true", help="Print demo JSONL requests and exit.")
    return parser


def run_self_test() -> int:
    server = XlsxStdioServer()
    requests = [
        {"command": "create_workbook", "args": {"workbook_id": "wb_demo", "default_sheet_name": "Data"}, "request_id": "1"},
        {
            "command": "write_range",
            "args": {
                "workbook_id": "wb_demo",
                "sheet": "Data",
                "start_cell": "A1",
                "values": [["Name", "Dept", "Score"], ["Alice", "IT", 95], ["Bob", "HR", 88]],
            },
            "request_id": "2",
        },
        {
            "command": "set_number_format",
            "args": {"workbook_id": "wb_demo", "sheet": "Data", "range": "C2:C3", "format_code": "0"},
            "request_id": "3",
        },
        {
            "command": "save_workbook_as",
            "args": {"workbook_id": "wb_demo", "path": "/mnt/data/excelstudio/xlsx_stdio_self_test.xlsx"},
            "request_id": "4",
        },
    ]
    for req in requests:
        response = server.process_payload(req)
        print(json.dumps(response, ensure_ascii=False, indent=2))
        if not response.get("ok"):
            return 1
    print("[OK] self-test completed")
    return 0


def print_demo_requests() -> int:
    demo = [
        {"command": "ping", "args": {}, "request_id": "req_ping"},
        {"command": "create_workbook", "args": {"workbook_id": "wb_001", "default_sheet_name": "Sheet1"}, "request_id": "req_create"},
        {
            "command": "write_range",
            "args": {
                "workbook_id": "wb_001",
                "sheet": "Sheet1",
                "start_cell": "A1",
                "values": [["姓名", "部門", "分數"], ["王小明", "IT", 95]],
            },
            "request_id": "req_write",
        },
        {"command": "save_workbook_as", "args": {"workbook_id": "wb_001", "path": "/tmp/demo.xlsx"}, "request_id": "req_save"},
    ]
    for item in demo:
        print(json.dumps(item, ensure_ascii=False))
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    ns = parser.parse_args(argv)
    if ns.self_test:
        return run_self_test()
    if ns.demo:
        return print_demo_requests()
    server = XlsxStdioServer()
    return server.run_forever()


if __name__ == "__main__":
    raise SystemExit(main())
